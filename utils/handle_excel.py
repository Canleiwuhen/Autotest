# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from xlutils.copy import copy
from utils.logger import logger_t as logger
# import win32com.client as win32

class OperationExcel:
    def __init__(self, file_name=None, sheet_name=None):
        # logger.debug("testcase sheet_name is : {}", sheet_name)
        self.sheet_name = sheet_name
        if file_name:
            self.file_name = file_name
        else:
            # logger.debug("file_name 为空, 读取默认值！")
            # self.file_name = '../bgi_test/data/test_case.xlsx'
            raise ValueError("file_name为空，请传入正确的文件路径！")
        try:
            # logger.debug("testcase file_name is : {}", self.file_name)
            self.workbook = load_workbook(self.file_name)
        except OSError as reason:
            logger.error("读取excel错误，请检查文件: {}!", self.file_name)
            logger.error("错误原因：{}", reason)
            raise OSError(f"读取excel错误，请检查文件: {self.file_name}!错误原因：{reason}")

    def get_data(self):
        """
        通过 sheetname 读取sheet
        """
        try:
            ws = self.workbook[self.sheet_name]
            return ws
        except Exception as e:
            logger.error("sheet_name 不存在，请检查：{}", self.sheet_name)
            logger.error("错误原因：{}", e)
            raise Exception(f"sheet_name不存在，请检查: {self.sheet_name}!错误原因：{e}")

    def get_sheet_index(self):
        """
        通过 sheetname 获取sheet_index
        """
        try:
            sheets = self.workbook.sheetnames
            return sheets.index(self.sheet_name)
        except Exception as e:
            logger.error("sheet_name 不存在，请检查：{}", self.sheet_name)
            logger.error("错误原因：{}", e)
            raise Exception(f"sheet_name不存在，请检查: {self.sheet_name}!错误原因：{e}")

    def get_lines(self):
        """
        获取单元格的行数
        """
        tables = self.get_data()
        row_num = tables.max_row
        for i in range(row_num):
            # 当最大行的前3列均为空时认为该行为空行
            if self.get_cell_value(row_num-1, 0) is None and self.get_cell_value(row_num-1, 1) is None and\
                    self.get_cell_value(row_num-1, 2) is None:
                row_num -= 1
        return row_num

    def get_cols(self):
        """
        获取单元格的列数
        """
        tables = self.get_data()
        col_num = tables.max_column
        for i in range(col_num):
            # 当最大行的前3列均为空时认为该行为空行
            if self.get_cell_value(0, col_num-1) is None and self.get_cell_value(1, col_num-1) is None and\
                    self.get_cell_value(2, col_num-1) is None:
                col_num -= 1
        return col_num

    def get_cell_value(self, row, col):
        """
        获取某一个单元格的内容  openpyxl库的row和col是从1开始  解决历史问题 把row和col都加了1
        """
        return self.get_data().cell(row+1, col+1).value

    def write_value(self, row, col, value):
        """
        使用 openpyxl 写入excel数据  可以保留文件格式
        row,col,value
        """
        wb = load_workbook(self.file_name)
        sheet_data = wb[self.sheet_name]
        sheet_data.cell(row+1, col+1, value)
        wb.save(self.file_name)

    def write_values(self, sheet_id, cols, value):
        """
        写入多列数据 ---------待定
        row,col,value
        """
        lines = self.get_lines()
        read_data = load_workbook(self.file_name)
        write_data = copy(read_data)
        sheet_data = write_data.get_sheet(sheet_id)
        for col in cols:
            for row in range(1, lines):
                sheet_data.write(row, int(col), value)
        write_data.save(self.file_name)

    def get_row_data(self, row):
        """
        根据行号，找到该行的内容
        """
        tables = self.get_data()
        row_data = tables.row_values(row)
        return row_data

    def get_col_data(self, col_id):
        """
        获取某一列的内容
        """
        if col_id is not None:
            cols = self.get_data().col_values(col_id)
        else:
            cols = self.get_data().col_values(0)
        return cols

    #
    def get_cell_type(self, row, col):
        """
        获取单元格类型  cell_type 0 空单元格, 1  字符串,  2  float
        """
        cell_type = self.get_data().cell_type(row, col)
        return cell_type

    # 通过win32库模拟Windows启动Excel应用程序打开Excel文件后保存
    # def open_excel_by_win32(self, file_name=None):
    #     if not file_name:
    #         file_name = self.file_name
    #
    #     # 启动Excel应用程序
    #     excel = win32.gencache.EnsureDispatch('Excel.Application')
    #     # 设置Excel可见性
    #     excel.Visible = True
    #     # 打开一个包含VBA的Excel文件
    #     workbook = excel.Workbooks.Open(filename_new)
    #     # 接下来你可以对workbook进行操作
    #     # 关闭保存工作簿
    #     workbook.Save()
    #     # 关闭Excel应用程序
    #     excel.Quit()
    #     # 清理资源
    #     del workbook
    #     del excel


if __name__ == '__main__':
    # filename = r'../testcase/samplecenter/inventory_audit/batch_save_template.xlsx'
    filename = r'Z:\qa-goalkeeper\apiforward\testcase\samplecenter\samplecenter\abc.xlsx'
    filename_new = r'Z:\qa-goalkeeper\apiforward\testcase\samplecenter\samplecenter\入库审核_批量审核_new.xlsx'
    # sheetname = 'Sheet1'
    # oe = OperationExcel(file_name=filename, sheet_name=sheetname)
    # oe.write_value(1, 0, "24X082000074")

    # from openpyxl import load_workbook
    # wb = load_workbook(filename, data_only=True)
    # ws = wb.active
    # ws['A2'] = '24X082000075'
    # ws['B2'] = '24SZDNA02-0044'
    # ws['C2'] = 'A02'
    # wb.save(filename_new)
    # import pandas as pd
    # df = pd.read_excel(filename, sheet_name='Sheet1')
    # df.loc[0, '样本/产物编号'] = "24X082000075"
    # with pd.ExcelWriter(filename_new, engine='openpyxl') as writer:
    #     df.to_excel(writer, sheet_name='Sheet1', index=False)



