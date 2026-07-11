import copy
import json
import math
import os
import random
import re
import time
from copy import deepcopy
from http.client import responses
from typing import final
from datetime import datetime

import requests
from faker import Faker
from numpy.lib.function_base import select
# from Crypto.Random.random import sample
# from celery.worker.consumer.mingle import exception
from openpyxl import load_workbook
import ddddocr
from click import confirm
from cryptography.x509 import random_serial_number
from select import error

from data_generate.nifty import nifty_zkp_template, update_db
from data_generate.nifty.handle_cnv import handle_cnv_list
from data_generate.nifty.update_db import UpdateDB
from utils.handle_db import HandleDB
from utils.tools import sep, get_project_path, rsa_encrpt, data_to_image, replace_none, create_sample, \
    create_expressnum, page_rsa_encrpt
from utils.handle_yaml import GetConfig
from utils.request import Requests
from urllib.parse import urlencode
from utils.logger import logger_nifty_dg as logger
from data_generate.samplecenter.datagenerate import DataGenerate
from testcase.samplecenter.conftest import login


class NiftydataGenerate:
    def __init__(self, token, area_code=None, istest=None):
        # current_directory = os.getcwd()
        # demo_path = os.path.dirname(os.path.dirname(current_directory))
        self.result_data = []
        self.chip_num = None
        if area_code:
            self.area_code = area_code
        else:
            self.area_code = "A020"
        self.sequencing_task_code = None # 上机测序任务号
        # istest用于判断是造数还是自动化测试，为True时读取默认入参走主流程自动化测试，否则读取前端动态输入的参数造数
        if not istest:
            record_id = os.environ.get('LAST_RECORD_ID')
            data = {
                "id": record_id
            }
            url = "http://127.0.0.1:8087/get_data"
            response = requests.get(url, params=urlencode(data))
            var_list = response.json()["message"]
            # with open('var_dir/nifty/var.txt', 'r', encoding='utf-8') as file:
            #     var_list = []
            #     for line in file:
            #         var_list.append(line.strip())
        else:
            with open('var_dir/nifty/var_default.txt', 'r') as file:
            # with open('../../var_dir/nifty/var_default.txt', 'r', encoding='utf-8') as file:
                var_list = []
                for line in file:
                    var_list.append(line.strip())
        # with open('var_dir/nifty/nifty_var.txt', 'r') as file:
        # with open('var_dir/nifty/var.txt', 'r') as file:
        #     var_list = []
        #     for line in file:
        #         var_list.append(line.strip())
        # wb = load_workbook(filename='../../upload_file/nifty_input.xlsx')
        # self.config_info = {"indexConfig":"NIFTY1-48", "sequencePlatform": "MGISEQ-2000", "sequenceType": "SE35+10", "data":"6M", "isChipCode": "", "chipCode":"" ,"machineNum":"M080101"}
        self.config_info ={
            "indexConfig": var_list[7],
            "sequencePlatform": var_list[8],
            "sequenceType": var_list[9],
            "data": var_list[6],
            "isChipCode": var_list[10],
            "chipCode": var_list[11],
            "machineNum": "M080101",
            "env": var_list[12]
        }
        self.jkfdealorder_plate_code = None # 建库孔板号
        self.bmg_task_code = None  # BMG任务号
        self.jkfdealorder_task_code = None # 建库任务号
        self.makednb_task_code = None # makeDNB任务号
        # wb = load_workbook(filename=r'Z:\gitlab\apiforward\upload_file\nifty_input.xlsx')
        # sheet = wb['sample_nifty']
        # sample_excel_list = []
        # for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
        #     sample_excel_list.append(row)
        self.sample_excel_list = eval(var_list[0])
        now_date = datetime.now()
        date_str = now_date.strftime('%Y%m%d')
        time_str = now_date.strftime('%H%M%S')
        datatime_str = now_date.strftime('%Y-%m-%d %H:%M:%S')
        self.date_str = date_str
        self.time_str = time_str
        self.datatime_str =datatime_str
        self.zjob_code = None
        self.zjcwbh = None     # 质检产物编号
        self.token = token
        self.configname = "nifty_config.yaml"
        # self.datafactory_res = Requests(configname='samplecenter_config.yaml', baseurl='data_factory_url')
        if self.config_info["env"] == "HK":
            self.configname = "niftyHK_config.yaml"
        else:
            self.configname = "nifty_config.yaml"
        self.nifty_res = Requests(configname=self.configname, baseurl='test_url',
                                         headers={
                                             "content-type": "application/x-www-form-urlencoded",
                                             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0",
                                             "accept-encoding": "gzip, deflate",
                                             "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
                                             "Accept": "*/*"})
        self.nifty_api_res = Requests(configname=self.configname, baseurl='save_nifty_result',
                                         headers={"content-type": "application/x-www-form-urlencoded"})
        # self.samplecenter_res = Requests(configname='samplecenter_config.yaml', baseurl='test_url',
        #                                  headers={"content-type": "application/x-www-form-urlencoded"})
        # self.expressnum = ""
        self.container_prefix = f"{int(time.time())}"
        self.specifications = 96
        self.sample = []
        self.abnormal_input = var_list[3]    # 是否异常 0否 1是
        self.pooling_scheme = None
        self.dlhhorder_scheme = None
        self.lane_index_info = [{"lane":i["Lane"],"index":i["index"]} for i in self.sample_excel_list]
        self.qc_info = [
            {"cnv_band": i["CNV_info"], "cnv": i["CNV_section_info"], "test13": i["test13"], "test18": i["test18"], "test21": i["test21"], "test_sex": i["test_sex"],
             "test_auto": i["test_auto"], "note3": i["note3"], "note2": i["note2"], "disease": i["disease"], "qc": i["qc"], "report_tag": i["report_tag"],
             "product_no": i["productNum"], "fetus_type": i["tireType"]} for i in self.sample_excel_list]
        self.chr_info = [
            {
                "chr": i["chr"],
                "chrTest": i["chrTest"],
                "filterFlag": i["filterFlag"],
                "fra": i["fra"],
                "risk": i["risk"],
                "t": i["t"],
                "zScore": i["zScore"],
            }
            for i in self.sample_excel_list
        ]

    def get_sample_detail(self, token, sample_id):
        '''
        # 查询获取身份证号和姓名
        :param token: 
        :param sample_id:
        :return id_num:
        :return name:
        '''
        res = Requests(configname= self.configname, baseurl="mybgi_url",
                       headers={"content-type": "application/x-www-form-urlencoded"})
        data = {
            "userName": "huxiaofeng_A020",
            "token": token,
            "datas": [
                {
                    "dataId": "3fa86594d6ec420b8bcc8ac900eb8e46",
                    "dataType": "MYBGI-医学页面数据大权限",
                    "field": "",
                    "sign": "",
                    "option": "",
                    "low": "",
                    "high": "",
                    "userId": ""
                },
                {
                    "dataId": "4f4e9c2f20e14569a12c2d42d20792b5",
                    "dataType": "深圳信息审核组",
                    "field": "",
                    "sign": "",
                    "option": "",
                    "low": "",
                    "high": "",
                    "userId": ""
                },
                {
                    "dataId": "8fe91e19f13c43b2a68678406ddbae63",
                    "dataType": "管理员",
                    "field": "",
                    "sign": "",
                    "option": "",
                    "low": "",
                    "high": "",
                    "userId": ""
                },
                {
                    "dataId": "ea72c7ef32874cada50c6ac707d780af",
                    "dataType": "鲸云管理员",
                    "field": "",
                    "sign": "",
                    "option": "",
                    "low": "",
                    "high": "",
                    "userId": ""
                },
                {
                    "dataId": "MOD0FYokSw",
                    "dataType": "科服超级管理员",
                    "field": "",
                    "sign": "",
                    "option": "",
                    "low": "",
                    "high": "",
                    "userId": ""
                },
                {
                    "dataId": "SiAy5R2zOt",
                    "dataType": "前端管理员",
                    "field": "",
                    "sign": "",
                    "option": "",
                    "low": "",
                    "high": "",
                    "userId": ""
                }
            ],
            "pageId": page_rsa_encrpt("menu_2_1", "sam/queryMysamplesNew"),
            "SPRAS": "1",
            "info": {
                "ZSAMPLE": [
                    sample_id,
                    "EQ"
                ],
                "pageSize": [
                    "500",
                    "EQ"
                ],
                "pageIndex": [
                    "1",
                    "EQ"
                ]
            }
        }
        response = res.post_request("/mybgi/web/webintf.do?method=sam/queryMysamplesNew",
                                    data=urlencode(data))
        if len(response.json()["rows"])>0:
            logger.info("查询样本的送检单号成功!")
            insp_id = response.json()["rows"][-1]["ZSJDID"]
            data = {
                "userName": "huxiaofeng_A020",
                "token": token,
                "datas": [
                    {
                        "dataId": "3fa86594d6ec420b8bcc8ac900eb8e46",
                        "dataType": "MYBGI-医学页面数据大权限",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "4f4e9c2f20e14569a12c2d42d20792b5",
                        "dataType": "深圳信息审核组",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "8fe91e19f13c43b2a68678406ddbae63",
                        "dataType": "管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "ea72c7ef32874cada50c6ac707d780af",
                        "dataType": "鲸云管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "MOD0FYokSw",
                        "dataType": "科服超级管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "SiAy5R2zOt",
                        "dataType": "前端管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    }
                ],
                "pageId": page_rsa_encrpt("menu_2_1", "sam/sampleInfo"),
                "SPRAS": "1",
                "sampleInfo": {"ZSJDID": insp_id, "ZTEMPBS": "Nifty", "ZITEMCP": ""}
            }
            response = res.post_request("/mybgi/web/webintf.do?method=sam/sampleInfo",
                                        data=urlencode(data))
            if len(response.json()["rows"])>0:
                logger.info("查询样本详情成功!")
                id_num = response.json()["rows"][1]["ZIDCARD"]
                name = response.json()["rows"][1]["ZSAMPLENAME"]
            else:
                logger.error("查询样本详情失败!{}", response.json())
                raise Exception("查询样本详情失败!")
            return {"id_num": id_num, "name": name}
        else:
            logger.error("查询样本的送检单号失败!{}", response.json())
            raise Exception("查询样本的送检单号失败!")

    def random_idcard(self,gender="女"):
        """
        随机生成身份证
        :return:
        """
        fake = Faker(["zh_CN"])
        str1 = fake.ssn()
        if gender == '男':
            # 确保第17位是奇数
            while int(str1[16]) % 2 == 0:
                str1 = fake.ssn()
        elif gender == '女':
            # 确保第17位是偶数
            while int(str1[16]) % 2 != 0:
                str1 = fake.ssn()
        else:
            raise ValueError("性别参数必须是 '男' 或 '女'")
        logger.info("生成随机身份证号：{}", str1)
        return str1

    def random_phone_number(self):
        """
        随机生成手机号码
        :return:
        """
        fake = Faker(["zh_CN"])
        phone_number = fake.phone_number()
        logger.info("生成随机手机号码：{}", phone_number)
        return phone_number

    def random_name(self):
        """
        随机生成姓名
        :return:
        """
        fake = Faker(["zh_CN"])
        name = fake.name_female()
        logger.info("生成随机姓名：{}", name)
        return name

    def sumbit_sample(self):
        """
        造数工具提交送检单，返回已提交送检单的样本列表，目前一次调用只有一个样本
        :return:
        """
        key = "MFwwDQYJKoZIhvcNAQEBBQADSwAwSAJBAMmRhnJLei0SR/d6UkdpCgJHvF+3ygzhVh0CfPtJpSAX4SFPt75eXGw0VVKPrGQ1+FIsJF3dCHi/dq4SSHrI2fsCAwEAAQ=="
        res = Requests(configname=self.configname, baseurl="mybgi_url", headers={"content-type": "application/x-www-form-urlencoded"})
        # 获取验证码
        time_13 = int(time.time() * 1000)
        param = {"time": time_13}
        content = res.get_request("/mybgi/vcode.do", params=param).content
        cur_path = get_project_path()
        path_tmp = [cur_path, "img/mybgi_code.png"]
        img_path = sep(path_tmp)
        data_to_image(content, img_path)
        # 识别图片中的验证码
        ocr = ddddocr.DdddOcr()
        with open(img_path, 'rb') as f:
            img_bytes = f.read()
        code = ocr.classification(img_bytes)

        # 构造登录接口入参
        data = {"userName": "huxiaofeng_A020",
                "passWord": rsa_encrpt(key,"huxiaofeng_A020"),
                "vcode": code,
                "method": "getNewUamsUserAuthorityUser",
                "hasObj": "2",
                "language": "chinese",
                "shortMessageVerification": "3"
                }
        # 执行接口请求
        login_res = res.post_request("/mybgi/user/authorityUser.do", data=data)
        if login_res.status_code == 200 and login_res.json()["code"] == "200" and login_res.json()["msg"] == "200":
            logger.info(f"mybgi登录成功")
            new_token = (login_res.json()["rows"][0])["token"]
        elif login_res.json()["msg"] == "验证码不正确!":  # 验证码错误重试登录
            return self.sumbit_sample()
        for i in range(len(self.sample_excel_list)):
            if self.sample_excel_list[i]["sampleType"] =="S051-全血":
                zyblx_code= "S051"
            if self.sample_excel_list[i]["sampleType"] =="S052-血浆":
                zyblx_code = "S052"
            matnr = self.sample_excel_list[i]["productNum"]
            # 格式化日期到毫秒
            now_date = datetime.now()
            datatime_m_str = now_date.strftime('%Y%m%d%H%M%S%f')
            zcatalo = self.date_str[2:4] + "B" + self.date_str[4:8] + datatime_m_str[8:16]
            # while True:
            #     zcatalo = self.date_str[2:4] + "B" + self.date_str[4:8] + str(random.randint(1000, 9999))

            #     if response['code'] =="200" and response['msg']=="没有查询到数据！":
            #         break
            special_zcatalo = self.sample_excel_list[i]["special_sample_id"]
            data = {
                "userName": "huxiaofeng_A020",
                "token": new_token,
                "datas": [
                    {
                        "dataId": "3fa86594d6ec420b8bcc8ac900eb8e46",
                        "dataType": "MYBGI-医学页面数据大权限",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "4f4e9c2f20e14569a12c2d42d20792b5",
                        "dataType": "深圳信息审核组",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "8fe91e19f13c43b2a68678406ddbae63",
                        "dataType": "管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "ea72c7ef32874cada50c6ac707d780af",
                        "dataType": "鲸云管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "MOD0FYokSw",
                        "dataType": "科服超级管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    },
                    {
                        "dataId": "SiAy5R2zOt",
                        "dataType": "前端管理员",
                        "field": "",
                        "sign": "",
                        "option": "",
                        "low": "",
                        "high": "",
                        "userId": ""
                    }
                ],
                "pageId": page_rsa_encrpt("menu_4_10", "sam/queryMysamplesNew"),
                "SPRAS": "1",
                "info": {
                    "ZSAMPLE": [
                        special_zcatalo,
                        "EQ"
                    ],
                    "pageSize": [
                        "20",
                        "EQ"
                    ],
                    "pageIndex": [
                        "1",
                        "EQ"
                    ]
                }
            }
            response = res.post_request("/mybgi/web/webintf.do?method=sam/queryMysamplesNew",
                                        data=urlencode(data)).json()
            if response['code'] =="200" and response['msg']=="没有查询到数据！":
                if self.sample_excel_list[i]["special_sample_id"].endswith(('R', 'RR', 'RRR')):
                    original_zcatalo = self.sample_excel_list[i]["special_sample_id"].rstrip('R')
                    data = {
                        "userName": "huxiaofeng_A020",
                        "token": new_token,
                        "datas": [
                            {
                                "dataId": "3fa86594d6ec420b8bcc8ac900eb8e46",
                                "dataType": "MYBGI-医学页面数据大权限",
                                "field": "",
                                "sign": "",
                                "option": "",
                                "low": "",
                                "high": "",
                                "userId": ""
                            },
                            {
                                "dataId": "4f4e9c2f20e14569a12c2d42d20792b5",
                                "dataType": "深圳信息审核组",
                                "field": "",
                                "sign": "",
                                "option": "",
                                "low": "",
                                "high": "",
                                "userId": ""
                            },
                            {
                                "dataId": "8fe91e19f13c43b2a68678406ddbae63",
                                "dataType": "管理员",
                                "field": "",
                                "sign": "",
                                "option": "",
                                "low": "",
                                "high": "",
                                "userId": ""
                            },
                            {
                                "dataId": "ea72c7ef32874cada50c6ac707d780af",
                                "dataType": "鲸云管理员",
                                "field": "",
                                "sign": "",
                                "option": "",
                                "low": "",
                                "high": "",
                                "userId": ""
                            },
                            {
                                "dataId": "MOD0FYokSw",
                                "dataType": "科服超级管理员",
                                "field": "",
                                "sign": "",
                                "option": "",
                                "low": "",
                                "high": "",
                                "userId": ""
                            },
                            {
                                "dataId": "SiAy5R2zOt",
                                "dataType": "前端管理员",
                                "field": "",
                                "sign": "",
                                "option": "",
                                "low": "",
                                "high": "",
                                "userId": ""
                            }
                        ],
                        "pageId": page_rsa_encrpt("menu_4_10", "sam/queryMysamplesNew"),
                        "SPRAS": "1",
                        "info": {
                            "ZSAMPLE": [
                                original_zcatalo,
                                "EQ"
                            ],
                            "pageSize": [
                                "20",
                                "EQ"
                            ],
                            "pageIndex": [
                                "1",
                                "EQ"
                            ]
                        }
                    }
                    response = res.post_request("/mybgi/web/webintf.do?method=sam/queryMysamplesNew",
                                                data=urlencode(data)).json()
                    if response['code'] =="200" and response['msg']=="没有查询到数据！":
                        id_num = self.random_idcard()
                        name = self.random_name()
                        phone_num = self.random_phone_number()
                        sample_info = {
                            "yl": [
                                {
                                    "ZITEMYL": "1",
                                    "ZSAMPLENAME": name,
                                    "ZIDCARDTYPE": "01",
                                    "ZIDCARD": id_num,
                                    "ZBIRTHDATE": id_num[6:14],
                                    "ZSAMPLEAGE": str(int(time.strftime('%Y')) - int(id_num[6:10])),
                                    "ZPHONENUM": phone_num
                                }
                            ],
                            "yb": [
                                {
                                    "ZITEMYL": "1",
                                    "ZTUBETYPE": self.sample_excel_list[i]["ZTUBETYPE"],
                                    "ZYBLX_CODE": zyblx_code
                                }
                            ],
                            "produc": [
                                {
                                    "ZITEMYL": "1",
                                    "MATNR": matnr
                                }
                            ],
                            "info": {
                                "KUNNR": "1000027045",
                                "ZBLOODDATE": self.date_str,
                                "ZGESTATIONALWEEKS": "16w+4",
                                "ZADOSCULATION": "ZADOSCULATIONNO",
                                "ZADOSCULATIONNO": "X",
                                "ZRESEARCHUSE": "3",
                                "ZISNEEDOTHERCHROMREPORT": "ZISNEEDOTHERCHROMREPORTYES",
                                "ZISNEEDOTHERCHROMREPORTYES": "X",
                                "ZISNEEDREPORTOTHERRESULT": "ZISNEEDREPORTOTHERRESULTYES",
                                "ZISNEEDREPORTOTHERRESULTYES": "X",
                                "ZGRZSY": "否",
                                "ZALLOGENETICTRAN": "ZALLOGENETICTRANSFUSIONNO",
                                "ZALLOGENETICTRANSFUSIONNO": "X",
                                "ZSURGICALTRANSPLAN": "ZSURGICALTRANSPLANTATIONNO",
                                "ZSURGICALTRANSPLANTATIONNO": "X",
                                "ZCELLULARIMMUN": "ZCELLULARIMMUNITYNO",
                                "ZCELLULARIMMUNITYNO": "X",
                                "ZSTEMCELL": "ZSTEMCELLNO",
                                "ZSTEMCELLNO": "X",
                                "ZZQTY": "01",
                                # "ZECZF": "X"
                            },
                            "sampleSite": []
                        }
                        if self.sample_excel_list[i]["tireType"]=="双胎":
                            if self.sample_excel_list[i]["special_sample_id"]:
                                sample_info["yb"][0]["ZCATALO"] = self.sample_excel_list[i]["special_sample_id"].rstrip('R')
                            else:
                                sample_info["yb"][0]["ZCATALO"] = zcatalo + "D"
                            sample_info["info"]["ZNORMAL"] = "ZNORMALDOUBLEBIRTH"
                            sample_info["info"]["ZNORMALDOUBLEBIRTH"] = "X"
                        if self.sample_excel_list[i]["tireType"] == "双胎减胎":
                            if self.sample_excel_list[i]["special_sample_id"]:
                                sample_info["yb"][0]["ZCATALO"] = self.sample_excel_list[i]["special_sample_id"].rstrip('R')
                            else:
                                sample_info["yb"][0]["ZCATALO"] = zcatalo + "V"
                            sample_info["info"]["ZNORMAL"] = "ZABNORMALITYDOUBLEREDUCTION"
                            sample_info["info"]["ZABNORMALITYDOUBLEREDUCTION"] = "X"
                        if self.sample_excel_list[i]["tireType"] == "单胎":
                            if self.sample_excel_list[i]["special_sample_id"]:
                                sample_info["yb"][0]["ZCATALO"] = self.sample_excel_list[i]["special_sample_id"].rstrip('R')
                            else:
                                sample_info["yb"][0]["ZCATALO"] = zcatalo
                            sample_info["info"]["ZNORMAL"] = "ZNORMALSINGLEBIRTH"
                            sample_info["info"]["ZNORMALSINGLEBIRTH"] = "X"
                        sample_data = {
                            "userName": "huxiaofeng_A020",
                            "token": new_token,
                            "datas": [
                                {
                                    "dataId": "3fa86594d6ec420b8bcc8ac900eb8e46",
                                    "dataType": "MYBGI-医学页面数据大权限",
                                    "field": "",
                                    "sign": "",
                                    "option": "",
                                    "low": "",
                                    "high": "",
                                    "userId": ""
                                },
                                {
                                    "dataId": "4f4e9c2f20e14569a12c2d42d20792b5",
                                    "dataType": "深圳信息审核组",
                                    "field": "",
                                    "sign": "",
                                    "option": "",
                                    "low": "",
                                    "high": "",
                                    "userId": ""
                                },
                                {
                                    "dataId": "8fe91e19f13c43b2a68678406ddbae63",
                                    "dataType": "管理员",
                                    "field": "",
                                    "sign": "",
                                    "option": "",
                                    "low": "",
                                    "high": "",
                                    "userId": ""
                                },
                                {
                                    "dataId": "ea72c7ef32874cada50c6ac707d780af",
                                    "dataType": "鲸云管理员",
                                    "field": "",
                                    "sign": "",
                                    "option": "",
                                    "low": "",
                                    "high": "",
                                    "userId": ""
                                },
                                {
                                    "dataId": "MOD0FYokSw",
                                    "dataType": "科服超级管理员",
                                    "field": "",
                                    "sign": "",
                                    "option": "",
                                    "low": "",
                                    "high": "",
                                    "userId": ""
                                },
                                {
                                    "dataId": "SiAy5R2zOt",
                                    "dataType": "前端管理员",
                                    "field": "",
                                    "sign": "",
                                    "option": "",
                                    "low": "",
                                    "high": "",
                                    "userId": ""
                                }
                            ],
                            "pageId": page_rsa_encrpt("menu_2_1", "sam/addSamples"),
                            "SPRAS": "1",
                            "sampleInfo": sample_info,
                            "xinxi": "Nifty"
                        }
                        response = res.post_request("/mybgi/web/webintf.do?method=sam/addSamples", data=urlencode(sample_data)).json()
                        if response['code'] =="200":
                            logger.info("创建原始样本成功!")
                        else:
                            logger.error(f"创建原始样本失败，返回结果：{response}")
                            raise Exception
                    else:
                        id_num = self.get_sample_detail(new_token,self.sample_excel_list[i]["special_sample_id"].rstrip('R'))["id_num"]
                        name = self.get_sample_detail(new_token,self.sample_excel_list[i]["special_sample_id"].rstrip('R'))["name"]
                else:
                    id_num = self.random_idcard()
                    name = self.random_name()
                phone_num = self.random_phone_number()
                sample_info = {
                    "yl": [
                        {
                            "ZITEMYL": "1",
                            "ZSAMPLENAME": name,
                            "ZIDCARDTYPE": "01",
                            "ZIDCARD": id_num,
                            "ZBIRTHDATE": id_num[6:14],
                            "ZSAMPLEAGE": str(int(time.strftime('%Y')) - int(id_num[6:10])),
                            "ZPHONENUM": phone_num
                        }
                    ],
                    "yb": [
                        {
                            "ZITEMYL": "1",
                            "ZTUBETYPE": self.sample_excel_list[i]["ZTUBETYPE"],
                            "ZYBLX_CODE": zyblx_code
                        }
                    ],
                    "produc": [
                        {
                            "ZITEMYL": "1",
                            "MATNR": matnr
                        }
                    ],
                    "info": {
                        "KUNNR": "1000027045",
                        "ZBLOODDATE": self.date_str,
                        "ZGESTATIONALWEEKS": "16w+4",
                        "ZADOSCULATION": "ZADOSCULATIONNO",
                        "ZADOSCULATIONNO": "X",
                        "ZRESEARCHUSE": "3",
                        "ZISNEEDOTHERCHROMREPORT": "ZISNEEDOTHERCHROMREPORTYES",
                        "ZISNEEDOTHERCHROMREPORTYES": "X",
                        "ZISNEEDREPORTOTHERRESULT": "ZISNEEDREPORTOTHERRESULTYES",
                        "ZISNEEDREPORTOTHERRESULTYES": "X",
                        "ZGRZSY": "否",
                        "ZALLOGENETICTRAN": "ZALLOGENETICTRANSFUSIONNO",
                        "ZALLOGENETICTRANSFUSIONNO": "X",
                        "ZSURGICALTRANSPLAN": "ZSURGICALTRANSPLANTATIONNO",
                        "ZSURGICALTRANSPLANTATIONNO": "X",
                        "ZCELLULARIMMUN": "ZCELLULARIMMUNITYNO",
                        "ZCELLULARIMMUNITYNO": "X",
                        "ZSTEMCELL": "ZSTEMCELLNO",
                        "ZSTEMCELLNO": "X",
                        "ZZQTY": "01",
                        # "ZECZF": "X"
                    },
                    "sampleSite": []
                }
                if self.sample_excel_list[i]["tireType"]=="双胎":
                    if self.sample_excel_list[i]["special_sample_id"]:
                        sample_info["yb"][0]["ZCATALO"] = self.sample_excel_list[i]["special_sample_id"]
                    else:
                        sample_info["yb"][0]["ZCATALO"] = zcatalo + "D"
                    sample_info["info"]["ZNORMAL"] = "ZNORMALDOUBLEBIRTH"
                    sample_info["info"]["ZNORMALDOUBLEBIRTH"] = "X"
                if self.sample_excel_list[i]["tireType"] == "双胎减胎":
                    if self.sample_excel_list[i]["special_sample_id"]:
                        sample_info["yb"][0]["ZCATALO"] = self.sample_excel_list[i]["special_sample_id"]
                    else:
                        sample_info["yb"][0]["ZCATALO"] = zcatalo + "V"
                    sample_info["info"]["ZNORMAL"] = "ZABNORMALITYDOUBLEREDUCTION"
                    sample_info["info"]["ZABNORMALITYDOUBLEREDUCTION"] = "X"
                if self.sample_excel_list[i]["tireType"] == "单胎":
                    if self.sample_excel_list[i]["special_sample_id"]:
                        sample_info["yb"][0]["ZCATALO"] = self.sample_excel_list[i]["special_sample_id"]
                    else:
                        sample_info["yb"][0]["ZCATALO"] = zcatalo
                    sample_info["info"]["ZNORMAL"] = "ZNORMALSINGLEBIRTH"
                    sample_info["info"]["ZNORMALSINGLEBIRTH"] = "X"
                sample_data = {
                    "userName": "huxiaofeng_A020",
                    "token": new_token,
                    "datas": [
                        {
                            "dataId": "3fa86594d6ec420b8bcc8ac900eb8e46",
                            "dataType": "MYBGI-医学页面数据大权限",
                            "field": "",
                            "sign": "",
                            "option": "",
                            "low": "",
                            "high": "",
                            "userId": ""
                        },
                        {
                            "dataId": "4f4e9c2f20e14569a12c2d42d20792b5",
                            "dataType": "深圳信息审核组",
                            "field": "",
                            "sign": "",
                            "option": "",
                            "low": "",
                            "high": "",
                            "userId": ""
                        },
                        {
                            "dataId": "8fe91e19f13c43b2a68678406ddbae63",
                            "dataType": "管理员",
                            "field": "",
                            "sign": "",
                            "option": "",
                            "low": "",
                            "high": "",
                            "userId": ""
                        },
                        {
                            "dataId": "ea72c7ef32874cada50c6ac707d780af",
                            "dataType": "鲸云管理员",
                            "field": "",
                            "sign": "",
                            "option": "",
                            "low": "",
                            "high": "",
                            "userId": ""
                        },
                        {
                            "dataId": "MOD0FYokSw",
                            "dataType": "科服超级管理员",
                            "field": "",
                            "sign": "",
                            "option": "",
                            "low": "",
                            "high": "",
                            "userId": ""
                        },
                        {
                            "dataId": "SiAy5R2zOt",
                            "dataType": "前端管理员",
                            "field": "",
                            "sign": "",
                            "option": "",
                            "low": "",
                            "high": "",
                            "userId": ""
                        }
                    ],
                    "pageId": page_rsa_encrpt("menu_2_1", "sam/addSamples"),
                    "SPRAS": "1",
                    "sampleInfo": sample_info,
                    "xinxi": "Nifty"
                }
                response = res.post_request("/mybgi/web/webintf.do?method=sam/addSamples", data=urlencode(sample_data)).json()
                if response['code'] =="200":
                    text = response['msg']
                    pattern = r"样例号: (\w+)"
                    match = re.search(pattern, text)
                    sample = match.group(1)
                    logger.info(f"造数工具提交送检单成功，返回样本数据：{sample}")
                    self.sample.append(sample)
                else:
                    logger.error(f"造数工具提交送检单失败，返回结果：{response}")
                    raise Exception
            else:
                logger.info(f"送检单已存在，返回样本数据：{special_zcatalo}")
                self.sample.append(special_zcatalo)
        # exp_data = DataGenerate().send_package( ','.join(self.sample))
        login_data = login(user='sitest')
        DataGenerate(token=login_data["token"]).sample_batch_receive(self.sample)
        # DataGenerate(token=login_data["token"]).receive_package(exp_data)
        # DataGenerate(token=login_data["token"]).unpack(exp_data)
        # for i in range(0,len(self.sample),96):
        #     container_data = DataGenerate(token=login_data["token"]).create_container(self.container_prefix,
        #                                                                               self.specifications)
        #
        #     DataGenerate(token=login_data["token"]).locate_position(samples=self.sample[i:i + 96],
        #                                                             container_num=container_data[0],
        #                                                             container_id=container_data[1])
        DataGenerate(token=login_data["token"]).medical_info_audit_batch(self.sample)
    def technical_route_confirmation(self, sample=None):
        """
        造数工具确认技术路线
        :return:
        """
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        data = {
            "task":{
                "zsample": sampleid,
                "zsjd_type": "YX"
            },
            "pageNumber": "1",
            "pageSize": "1000",
            "token": self.token,
            "menuId": "MedicalScienceRouting_RoutingAgainConfirm"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=findSampleRoutingMS", data=urlencode(data))
        if response.json()["code"] == "200":
            logger.info("技术路线确认列表查询成功！")
            response_data = (response.json()["data"])
            datas = []
            for i in range(len(response_data)):
                data_dic = {k: "" if v is None else v for k, v in response_data[i].items()}
                data_dic["_key"] = i + 1
                data_dic["chkstatus"] = '未审核'
                data_dic["_id"] = i + 1
                data_dic["_zcatalo"] = {
                                "type": "a",
                                "key": "",
                                "ref": "",
                                "props": {
                                    "children": (sampleid.split(','))[i]
                                },
                                "_owner": ""
                            }
                datas.append(data_dic)
            data = {
                "task": datas,
                "token": self.token,
                "menuId": "MedicalScienceRouting_RoutingAgainConfirm"
            }
            confirm_response = self.nifty_res.post_request("/presap/webintf.do?method=updateSampleRoutingMS",
                                                           data=urlencode(data))
            if confirm_response.status_code == 200 and confirm_response.json()["code"] == "200" and \
                    "数据保存成功." in confirm_response.json()["msg"]:
                logger.info("技术路线确认成功！")
            else:
                logger.error(f"技术路线确认失败！response：{confirm_response.json()}")
                raise Exception(f"技术路线确认失败！response：{confirm_response.json()}")
        else:
            logger.error(f"技术路线确认列表查询失败！response：{response.json()}")
            raise Exception(f"技术路线确认列表查询失败！response：{response.json()}")

    def plasma_separation(self, sample=None):
        """
        造数工具血浆分离
        :return:
        """
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        data = {
            "task": {
                "zsampling_datum": "",
                "zsampling_datumend": "",
                "zreceiveddate": "",
                "zreceiveddateend": "",
                "zsample": sampleid
            },
            "pageNumber": "1",
            "pageSize": "1000",
            "zgxbh": "MAB",
            "token": self.token,
            "menuId": "MSTaskMaster_XueJiangFenLi_XJFLOrder"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=task_assign_samplems", data=urlencode(data))
        if response.json()["code"] == "200":
            logger.info("血浆分离新建任务列表查询成功！")
            response_data = response.json()["data"]
            now_date = datetime.now()
            date_str = now_date.strftime('%Y%m%d')
            time_str = now_date.strftime('%H%M%S')
            datatime_str = now_date.strftime('%Y-%m-%d %H:%M:%S')
            task = [
                {
                    "zplate_bcode": response_data[i]["zplate_bcode"],
                    "matnr": response_data[i]["matnr"],
                    "kunnr": response_data[i]["kunnr"],
                    "zdfpp_datum": response_data[i]["zdfpp_datum"],
                    "zitemyl": response_data[i]["zitemyl"],
                    "zitemyb": response_data[i]["zitemyb"],
                    "ztabname": "",
                    "zname_all": response_data[i]["zname_all"],
                    "maktx": response_data[i]["maktx"],
                    "zreceiveddate": response_data[i]["zreceiveddate"],
                    "zsjdid": response_data[i]["zsjdid"],
                    "zcatalo": response_data[i]["zcatalo"],
                    "posnr": response_data[i]["posnr"],
                    "zsampling_time": response_data[i]["zsampling_time"],
                    "zutime": response_data[i]["zutime"],
                    "zwkh": response_data[i]["zwkh"],
                    "ewbez": response_data[i]["ewbez"],
                    "werks": response_data[i]["werks"],
                    "zeile": response_data[i]["zeile"],
                    "ztempbs": response_data[i]["ztempbs"],
                    "zcdate": response_data[i]["zcdate"],
                    "zyblx_code": response_data[i]["zyblx_code"],
                    "zgxbh": response_data[i]["zgxbh"],
                    "zyxuzt": response_data[i]["zyxuzt"],
                    "zudate": response_data[i]["zudate"],
                    "zctime": response_data[i]["zctime"],
                    "zyjpcdate": response_data[i]["zyjpcdate"],
                    "zyxdat": response_data[i]["zyxdat"],
                    "zreceiveduzit": response_data[i]["zreceiveduzit"],
                    "zsampling_datum": response_data[i]["zsampling_datum"],
                    "zsample": response_data[i]["zsample"],
                    "zcxlx": response_data[i]["zcxlx"],
                    "zguid": response_data[i]["zguid"],
                    "ztubetype": response_data[i]["ztubetype"],
                    "zsjd_type": response_data[i]["zsjd_type"],
                    "zcreator": response_data[i]["zcreator"],
                    "znum": response_data[i]["znum"],
                    "zdfpp_uzeit": response_data[i]["zdfpp_uzeit"],
                    "_key": i+1,
                    "zgxbh_des": "血浆分离",
                    "_id": i+1,
                    "zsyyn": "冼育萍",
                    "pernr": "00000103",
                    "zmethod_des": "1755.01.01深圳_产前组_样本分离",
                    "zmethod": "1755.01.01",
                    "zshould_finish_datum": date_str,
                    "znote": "自动化测试执行",
                    "zshould_finish_uzeit": time_str
                }
                for i in range(len(response_data))
                ]
            confirm_data = {
                "task": task,
                "token": self.token,
                "menuId": "MSTaskMaster_XueJiangFenLi_XJFLOrder"
            }
            confirm_response = self.nifty_res.post_request("/presap/webintf.do?method=saveTaskAssignSampleMS",
                                                   data=urlencode(confirm_data))
            if confirm_response.status_code == 200 and confirm_response.json()["code"] == "200":
                logger.info("血浆分离新建任务成功！")
                confirm_response_data = confirm_response.json()["data"][0]
                data = {
                    "task": {
                        "zsfxd_datum": "",
                        "zsfxd_datumend": "",
                        "zsfks_datum": "",
                        "zsfks_datumend": "",
                        "zjob_code": confirm_response_data["zjob_code"],
                        "zsfwc": "",
                        "task": "query",
                        "zcxlx": "0"
                    },
                    "pageNumber": "1",
                    "zgxbh": "MAB",
                    "pageSize": "50",
                    "token": self.token,
                    "menuId": "MSTaskProduct_XueJiangFenLi"
                }
                rwd_response = self.nifty_res.post_request("/presap/webintf.do?method=findJobMS", data=urlencode(data))
                rwd_data = rwd_response.json()["data"][0]
                if rwd_response.status_code == 200 and rwd_response.json()["code"] == "200" and \
                        rwd_data["zjob_code"]:
                    logger.info("血浆分离任务单查询成功！")
                    self.zjob_code = rwd_data["zjob_code"]
                    if self.abnormal_input =='1':
                        # 判断是否异常录入 0 否 1 是
                        data = {
                            "zjob_code": rwd_data["zjob_code"],
                            "token": self.token,
                            "menuId": "MSTaskProduct_XueJiangFenLi"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=msStartTask",
                                                               data=urlencode(data))
                        if response.status_code == 200 and response.json()["code"] == "200" and response.json()[
                            "msg"] == "success":
                            logger.info("血浆分离任务开始成功！")
                            data = {
                                "zgxbh": "MAB",
                                "werks": "A020",
                                "arbpl": "",
                                "token": self.token,
                                "menuId": "MSTaskProduct_XueJiangFenLi"
                            }
                            exception_response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                                       data=urlencode(data))
                            if exception_response.json()["code"] == "200" and len(exception_response.json()["data"]) > 0:
                                logger.info("血浆分离异常原因查询成功！")
                                exception_data = exception_response.json()["data"][0]
                                data = {
                                    "task": {
                                        "zpzxlh_bcode": "",
                                        "zcjdat": "",
                                        "ybbhrq": "",
                                        "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                        "zsfwc_czr": "",
                                        "zsfks_czr": "",
                                        "zcdate": "",
                                        "chkdate": "",
                                        "matnr": rwd_data["matnr"],
                                        "zseqplatform": "",
                                        "zsfks_uzeit": rwd_data["zsfks_uzeit"],
                                        "zdlhwkh": "",
                                        "qdfkrq": "",
                                        "zsampling_datum": "",
                                        "zdlvdate": """ """,
                                        "zgxbh": "MAB",
                                        "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                        "zsigndate": "",
                                        "zsfks": "",
                                        "zsyyn": rwd_data["zsyyn"],
                                        "syncdate": "",
                                        "ddl21": "",
                                        "yctjrq": "",
                                        "zcxlx": rwd_data["zcxlx"],
                                        "zjob_begin_datum": "",
                                        "zsfwc_uzeit": "",
                                        "zsample_quantity": rwd_data["zsample_quantity"],
                                        "zmethod": rwd_data["zmethod"],
                                        "zsfxd_czr": rwd_data["zsfxd_czr"],
                                        "zudate": "",
                                        "addate": "",
                                        "zycdjsflr": rwd_data["zycdjsflr"],
                                        "zsfxd_datum": rwd_data["zsfxd_datum"],
                                        "werks": rwd_data["werks"],
                                        "zsfks_datum": rwd_data["zsfks_datum"],
                                        "ycwcrq": "",
                                        "pre_zsfwc_datum": "",
                                        "zsfwc": "",
                                        "pre_zjob_code": "",
                                        "maktx": rwd_data["maktx"],
                                        "zsfwc_datum": "",
                                        "zupdat": "",
                                        "zjob_code": rwd_data["zjob_code"],
                                        "gt_action": "",
                                        "zcbdate": "",
                                        "gt_sendtime": "",
                                        "zfinish_time": "",
                                        "_key": 1,
                                        "_id": 1
                                    },
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_XueJiangFenLi"
                                }
                                rwd_detail_response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS", data=urlencode(data))
                                if rwd_detail_response.json()["code"] == "200" and (rwd_detail_response.json()["data"][0])["zjob_code"]==rwd_data["zjob_code"]:
                                    rwd_detail_data = rwd_detail_response.json()["data"]
                                    logger.info("血浆分离任务单详情查询成功！")
                                    task = [
                                            {
                                                "zplate_bcode": rwd_detail_data[i]["zplate_bcode"],
                                                "zzjcwbs": "",
                                                "zjob_code_source": "",
                                                "group_zcatalo": "",
                                                "zhhwkh_type": "",
                                                "zzjcwpdlen": "",
                                                "zzzcwsflr": "",
                                                "z_jk_ybsyl": "",
                                                "zsldfsh_uzeit": rwd_detail_data[i]["zsldfsh_uzeit"],
                                                "zsbgx": "",
                                                "zjw_type": "",
                                                "mbt_zjob_code": "",
                                                "zspl_check": "",
                                                "zbgijgdh_item": rwd_detail_data[i]["zbgijgdh_item"],
                                                "zhandle_actcode": rwd_detail_data[i]["zhandle_actcode"],
                                                "zdata_type": "",
                                                "zcybs": "",
                                                "zsfjieg": "",
                                                "zlgfsflr": "",
                                                "z_fzh": "",
                                                "zsfxd_datum": rwd_detail_data[i]["zsfxd_datum"],
                                                "zname_all": rwd_detail_data[i]["zname_all"],
                                                "zsldfsh_datum": rwd_detail_data[i]["zsldfsh_datum"],
                                                "zplate_x": "",
                                                "maktx": rwd_detail_data[i]["maktx"],
                                                "yc_lx": "",
                                                "zplate_y": "",
                                                "zafter_excp": "",
                                                "zkyks": rwd_detail_data[i]["zkyks"],
                                                "zreceiveddate": rwd_detail_data[i]["zreceiveddate"],
                                                "zsjdid": rwd_detail_data[i]["zsjdid"],
                                                "zflowid_source": "",
                                                "z_lane_no": rwd_detail_data[i]["z_lane_no"],
                                                "zload_machn_bcode": "",
                                                "zsfxd_uzeit": rwd_detail_data[i]["zsfxd_uzeit"],
                                                "zsfwc_czr": "",
                                                "zwkpd_d": rwd_detail_data[i]["zwkpd_d"],
                                                "zbq": "",
                                                "zcheckusername": "",
                                                "zback_job_repid": "",
                                                "zindex_name": "",
                                                "matnr_before": "",
                                                "zplate": "",
                                                "zbgijgdh_val": "",
                                                "zdlhwkh": "",
                                                "zpbzbh": "",
                                                "zgxdm": rwd_detail_data[i]["zgxdm"],
                                                "ztype": "",
                                                "zwkpd_u": rwd_detail_data[i]["zwkpd_u"],
                                                "zshdat": rwd_detail_data[i]["zshdat"],
                                                "stpch": "",
                                                "zutime": rwd_detail_data[i]["zutime"],
                                                "pooling_sum": "",
                                                "zinfoid": "",
                                                "datapath": "",
                                                "zgene": "",
                                                "zstatus": "",
                                                "ztj": "",
                                                "zeile": rwd_detail_data[i]["zeile"],
                                                "vbeln_item": rwd_detail_data[i]["vbeln_item"],
                                                "aufnr": "",
                                                "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                                "ztx": "",
                                                "zkpwbbh": "",
                                                "z_jk_ml": "",
                                                "project": "",
                                                "zrun_uzeit": rwd_detail_data[i]["zrun_uzeit"],
                                                "zcwname": "",
                                                "zgxbh": rwd_detail_data[i]["zgxbh"],
                                                "zcheckdesc": "",
                                                "zsfks": "X",
                                                "zrun_datum": rwd_detail_data[i]["zrun_datum"],
                                                "ddl21": "",
                                                "zcwlx": "",
                                                "area": "",
                                                "water_volume": "",
                                                "zmethod": rwd_detail_data[i]["zmethod"],
                                                "zlane_id": "",
                                                "zctime": rwd_detail_data[i]["zctime"],
                                                "zback_reason": "",
                                                "zindex": "",
                                                "zfgsl": rwd_detail_data[i]["zfgsl"],
                                                "zycsfwc": "",
                                                "ztx_el": "",
                                                "zindextype_el": "",
                                                "zhhwkh": "",
                                                "time_beascall": "",
                                                "zxmnum": "",
                                                "zplwkh": "",
                                                "zsampling_datum": rwd_detail_data[i]["zsampling_datum"],
                                                "vbeln": "",
                                                "zindex_seq": "",
                                                "volume_sum": "",
                                                "zcxlx": rwd_detail_data[i]["zcxlx"],
                                                "zshnam": "",
                                                "zreportid": "",
                                                "zsfhg": "",
                                                "zchangerby": "",
                                                "zguid": rwd_detail_data[i]["zguid"],
                                                "ztubetype": rwd_detail_data[i]["ztubetype"],
                                                "zjob_code_up": "",
                                                "zbglx": "",
                                                "zcreator": rwd_detail_data[i]["zcreator"],
                                                "zplate_sample_num": "",
                                                "zybrwwczt": rwd_detail_data[i]["zybrwwczt"],
                                                "chipholder": "",
                                                "zindex_el": "",
                                                "redeal_reason": "",
                                                "zscheme": "",
                                                "matnr": rwd_detail_data[i]["matnr"],
                                                "zconcentration": "",
                                                "z_zk_bfb": "",
                                                "zdnb_uzeit": rwd_detail_data[i]["zdnb_uzeit"],
                                                "zclbs": "",
                                                "kunnr": rwd_detail_data[i]["kunnr"],
                                                "zsfxd": rwd_detail_data[i]["zsfxd"],
                                                "zprimername": "",
                                                "zrow": rwd_detail_data[i]["zrow"],
                                                "matnr_sc": "",
                                                "zclassify": "",
                                                "zsfwc_uzeit": rwd_detail_data[i]["zsfwc_uzeit"],
                                                "zyblx": "",
                                                "zdnb_datum": rwd_detail_data[i]["zdnb_datum"],
                                                "zax_vol_spl": "",
                                                "z_jk_cycle": "",
                                                "zjob_status": "",
                                                "zsfwc": "",
                                                "theory_con": "",
                                                "zsfwc_datum": rwd_detail_data[i]["zsfwc_datum"],
                                                "zprimer_method": "",
                                                "zcatalo": rwd_detail_data[i]["zcatalo"],
                                                "zindextype": "",
                                                "zy": "",
                                                "zsfzkp": "",
                                                "posnr": rwd_detail_data[i]["posnr"],
                                                "zgxlb": "",
                                                "zseqplatform": "",
                                                "z_lane_gs": "",
                                                "zdatasource": rwd_detail_data[i]["zdatasource"],
                                                "zindex_seq_el": "",
                                                "zsampling_time": "000000",
                                                "zsfchanw": "",
                                                "zzk_matnr": "",
                                                "z_lane_no_bcode": "",
                                                "zwkh": "",
                                                "ewbez": rwd_detail_data[i]["ewbez"],
                                                "zsfxd_czr": rwd_detail_data[i]["zsfxd_czr"],
                                                "zycdjsflr": "",
                                                "zindex_name_el": "",
                                                "werks": rwd_detail_data[i]["werks"],
                                                "zsqldate": rwd_detail_data[i]["zsqldate"],
                                                "zshuzt": rwd_detail_data[i]["zshuzt"],
                                                "zprocodenum": "",
                                                "radat": "",
                                                "ztestnum": rwd_detail_data[i]["ztestnum"],
                                                "project_item": rwd_detail_data[i]["project_item"],
                                                "ztempid": "",
                                                "zcross_od": rwd_detail_data[i]["zcross_od"],
                                                "znote": "自动化测试执行",
                                                "zsfks_czr": rwd_detail_data[i]["zsfks_czr"],
                                                "zcdate": rwd_detail_data[i]["zcdate"],
                                                "zplatename": "",
                                                "z_wk_lane": rwd_detail_data[i]["z_wk_lane"],
                                                "mae_zcwbh": "",
                                                "mcc_stat": "",
                                                "zxmbm": "",
                                                "zbgijgdh_fld": "",
                                                "mbt_zjob_code_item": rwd_detail_data[i]["mbt_zjob_code_item"],
                                                "zsyyn": rwd_detail_data[i]["zsyyn"],
                                                "zzkpbm_inner": "",
                                                "zdna_method": "",
                                                "mo": "",
                                                "zguid_old": rwd_detail_data[i]["zguid_old"],
                                                "merar": "",
                                                "zudate": rwd_detail_data[i]["zudate"],
                                                "zrtpz": rwd_detail_data[i]["zrtpz"],
                                                "zzkxh": "",
                                                "znote_prd": "",
                                                "zsfks_datum": rwd_detail_data[i]["zsfks_datum"],
                                                "zload_location": "",
                                                "znm": "",
                                                "zindex01": rwd_detail_data[i]["zindex01"],
                                                "yc_no": "",
                                                "zreportid_item": rwd_detail_data[i]["zreportid_item"],
                                                "zitem_no": rwd_detail_data[i]["zitem_no"],
                                                "zkplx": "",
                                                "zpzxlh_bcode": "",
                                                "zzjcwsflr": rwd_detail_data[i]["zzjcwsflr"],
                                                "zgxsxh": rwd_detail_data[i]["zgxsxh"],
                                                "zrwdjgsflr": rwd_detail_data[i]["zrwdjgsflr"],
                                                "zsfks_uzeit": rwd_detail_data[i]["zsfks_uzeit"],
                                                "pre_zcwlx": "",
                                                "zcwbh": "",
                                                "zreceiveduzit": rwd_detail_data[i]["zreceiveduzit"],
                                                "group_zindex": "",
                                                "zsample": rwd_detail_data[i]["zsample"],
                                                "zpoint": "",
                                                "zfghsl": "",
                                                "zcxlx_txt": rwd_detail_data[i]["zcxlx_txt"],
                                                "zshould_sj_datum": rwd_detail_data[i]["zshould_sj_datum"],
                                                "zpzxlh": "",
                                                "maplog": "",
                                                "zdelete_flg": "",
                                                "zpoolingzdnum": "",
                                                "zberaid": "",
                                                "zpzxlh_location": "",
                                                "pre_zsfwc_datum": rwd_detail_data[i]["pre_zsfwc_datum"],
                                                "zshould_sj_uzeit": rwd_detail_data[i]["zshould_sj_uzeit"],
                                                "zcheck_result": "",
                                                "zjob_code": rwd_detail_data[i]["zjob_code"],
                                                "zfmd": "",
                                                "zgxxh": rwd_detail_data[i]["zgxxh"],
                                                "_id": 1,
                                                "zreason": exception_data["zreason"],
                                                "name_xl": exception_data["name_xl"],
                                                "yc_level": "",
                                                "yc_dl": exception_data["yc_dl"],
                                                "yc_ms": exception_data["yc_ms"],
                                                "name_dl": exception_data["name_dl"],
                                                "yc_xl": exception_data["yc_xl"],
                                                "zcjnam": rwd_detail_data[i]["zcreator"],
                                                "zcjdat": date_str,
                                                "zcjuzt": time_str,
                                                "dateTime": datatime_str,
                                                "yc_status": "A"
                                            }
                                        for i in range(len(rwd_detail_data))
                                        ]
                                    sub_exce_data = {
                                        "task": task,
                                        "werks": "A020",
                                        "arbpl": "",
                                        "zjob_code": rwd_data["zjob_code"],
                                        "zgxbh": "MAB",
                                        "token": self.token,
                                        "menuId": "MSTaskProduct_XueJiangFenLi"
                                    }
                                    sub_exce_response = self.nifty_res.post_request("/presap/webintf.do?method=save_exception_ms", data=urlencode(sub_exce_data))
                                    if sub_exce_response.json()["code"] == "200" and sub_exce_response.json()["msg"] == "数据保存成功.":
                                        logger.info("血浆分离登记异常录入成功！")
                                        self.zjob_code = rwd_data["zjob_code"]
                                        return self.zjob_code
                                    else:
                                        logger.error(f"血浆分离登记异常录入失败！response：{sub_exce_response.json()}")
                                        raise Exception(f"血浆分离登记异常录入失败！response：{sub_exce_response.json()}")
                                else:
                                    logger.error(f"血浆分离任务单详情查询失败！response：{rwd_detail_response.json()}")
                                    raise Exception(f"血浆分离任务单详情查询失败！response：{rwd_detail_response.json()}")
                        else:
                            logger.error(f"血浆分离任务开始失败！response：{response.json()}")
                            raise Exception(f"血浆分离任务开始失败！response：{response.json()}")
                    else:
                        data = {
                            "zjob_code": (rwd_response.json()["data"][0])["zjob_code"],
                            "token": self.token,
                            "menuId": "MSTaskProduct_XueJiangFenLi"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=msStartTask",
                                                               data=urlencode(data))
                        if response.status_code == 200 and response.json()["code"] == "200" and response.json()[
                            "msg"] == "success":
                            logger.info("血浆分离任务开始成功！")
                            data = {
                                "task": {
                                    "zjob_code": rwd_data["zjob_code"],
                                    "zgxbh": "MAB",
                                    "zsfzjcw": "ALL",
                                    "zsfwc": ""
                                },
                                "token": self.token,
                                "menuId": "MSTaskProduct_XueJiangFenLi"
                            }
                            response = self.nifty_res.post_request("/presap/webintf.do?method=find_chanWuMS",
                                                                   data=urlencode(data))
                            if response.status_code == 200 and response.json()["code"] == "200":
                                logger.info("血浆分离产物查询成功！")
                                result_response = response.json()["data"]
                                task = [
                                        {
                                            f"{i+1:06d}": {
                                                "tab1": [
                                                    {
                                                        "zplate_bcode": result_response["Samples"][i]["zplate_bcode"],
                                                        "zzjcwbs": "",
                                                        "zjob_code_source": "",
                                                        "group_zcatalo": "",
                                                        "zhhwkh_type": "",
                                                        "zzjcwpdlen": "",
                                                        "zzzcwsflr": "",
                                                        "z_jk_ybsyl": "",
                                                        "zsldfsh_uzeit": "000000",
                                                        "zsbgx": "",
                                                        "zjw_type": "",
                                                        "mbt_zjob_code": "",
                                                        "zspl_check": "",
                                                        "zbgijgdh_item": "000000",
                                                        "zhandle_actcode": "000000",
                                                        "zdata_type": "",
                                                        "zcybs": "",
                                                        "zsfjieg": "",
                                                        "zlgfsflr": "",
                                                        "z_fzh": "",
                                                        "zsfxd_datum": result_response["Samples"][i]["zsfxd_datum"],
                                                        "zname_all": result_response["Samples"][i]["zname_all"],
                                                        "zsldfsh_datum": result_response["Samples"][i]["zsldfsh_datum"],
                                                        "zplate_x": "",
                                                        "maktx": result_response["Samples"][i]["maktx"],
                                                        "yc_lx": "",
                                                        "zplate_y": "",
                                                        "zafter_excp": "",
                                                        "zkyks": result_response["Samples"][i]["zkyks"],
                                                        "zreceiveddate": result_response["Samples"][i]["zreceiveddate"],
                                                        "zsjdid": result_response["Samples"][i]["zsjdid"],
                                                        "zflowid_source": "",
                                                        "z_lane_no": "00",
                                                        "zload_machn_bcode": "",
                                                        "zsfxd_uzeit": result_response["Samples"][i]["zsfxd_uzeit"],
                                                        "zsfwc_czr": "",
                                                        "zwkpd_d": "0+",
                                                        "zbq": "",
                                                        "zcheckusername": "",
                                                        "zback_job_repid": "",
                                                        "zindex_name": "",
                                                        "matnr_before": "",
                                                        "zplate": "",
                                                        "zbgijgdh_val": "",
                                                        "zdlhwkh": "",
                                                        "zpbzbh": "",
                                                        "zgxdm": "MAB",
                                                        "ztype": "",
                                                        "zwkpd_u": "0+",
                                                        "zshdat": "00000000",
                                                        "stpch": "",
                                                        "zutime": "000000",
                                                        "pooling_sum": "",
                                                        "zinfoid": "",
                                                        "datapath": "",
                                                        "zgene": "",
                                                        "zstatus": "",
                                                        "ztj": 300,
                                                        "zeile": "0000",
                                                        "vbeln_item": "000000",
                                                        "aufnr": "",
                                                        "zjob_code_item": "000001",
                                                        "ztx": "",
                                                        "zkpwbbh": "",
                                                        "z_jk_ml": "",
                                                        "project": "",
                                                        "zrun_uzeit": "000000",
                                                        "zcwname": "",
                                                        "zgxbh": "MAB",
                                                        "zcheckdesc": "",
                                                        "zsfks": "X",
                                                        "zrun_datum": "00000000",
                                                        "ddl21": "",
                                                        "zcwlx": "血浆",
                                                        "area": "",
                                                        "water_volume": "",
                                                        "zmethod": result_response["Samples"][i]["zmethod"],
                                                        "zlane_id": "",
                                                        "zctime": result_response["Samples"][i]["zctime"],
                                                        "zback_reason": "",
                                                        "zindex": "",
                                                        "zfgsl": "00",
                                                        "zycsfwc": "",
                                                        "ztx_el": "",
                                                        "zindextype_el": "",
                                                        "zhhwkh": "",
                                                        "time_beascall": "",
                                                        "zxmnum": "",
                                                        "zplwkh": "",
                                                        "zsampling_datum": result_response["Samples"][i]["zsampling_datum"],
                                                        "vbeln": "",
                                                        "zindex_seq": "",
                                                        "volume_sum": "",
                                                        "zcxlx": "0",
                                                        "zshnam": "",
                                                        "zreportid": "",
                                                        "zsfhg": "",
                                                        "zchangerby": "",
                                                        "zguid": result_response["Samples"][i]["zguid"],
                                                        "ztubetype": result_response["Samples"][i]["ztubetype"],
                                                        "zjob_code_up": "",
                                                        "zbglx": "",
                                                        "zcreator": result_response["Samples"][i]["zcreator"],
                                                        "zplate_sample_num": "",
                                                        "zybrwwczt": "0",
                                                        "chipholder": "",
                                                        "zindex_el": "",
                                                        "redeal_reason": "",
                                                        "zscheme": "",
                                                        "matnr": result_response["Samples"][i]["matnr"],
                                                        "zconcentration": "",
                                                        "z_zk_bfb": "",
                                                        "zdnb_uzeit": "000000",
                                                        "zclbs": "",
                                                        "kunnr": result_response["Samples"][i]["kunnr"],
                                                        "zsfxd": "X",
                                                        "zprimername": "",
                                                        "zrow": "0000000000",
                                                        "matnr_sc": "",
                                                        "zclassify": "",
                                                        "zsfwc_uzeit": "000000",
                                                        "zyblx": "",
                                                        "zdnb_datum": "00000000",
                                                        "zax_vol_spl": "",
                                                        "z_jk_cycle": "",
                                                        "zjob_status": "",
                                                        "zsfwc": "",
                                                        "theory_con": "",
                                                        "zsfwc_datum": "00000000",
                                                        "zprimer_method": "",
                                                        "zcatalo": result_response["Samples"][i]["zcatalo"],
                                                        "zindextype": "",
                                                        "zy": "",
                                                        "zsfzkp": "",
                                                        "posnr": result_response["Samples"][i]["posnr"],
                                                        "zgxlb": "",
                                                        "zseqplatform": "",
                                                        "z_lane_gs": "",
                                                        "zdatasource": result_response["Samples"][i]["zdatasource"],
                                                        "zindex_seq_el": "",
                                                        "zsampling_time": "000000",
                                                        "zsfchanw": "",
                                                        "zzk_matnr": "",
                                                        "z_lane_no_bcode": "",
                                                        "zwkh": "",
                                                        "ewbez": "产前",
                                                        "zsfxd_czr": result_response["Samples"][i]["zsfxd_czr"],
                                                        "zycdjsflr": "",
                                                        "zindex_name_el": "",
                                                        "werks": result_response["Samples"][i]["werks"],
                                                        "zsqldate": "00000000",
                                                        "zshuzt": "000000",
                                                        "zprocodenum": "",
                                                        "radat": "",
                                                        "ztestnum": "00",
                                                        "project_item": "000000",
                                                        "ztempid": "",
                                                        "zcross_od": "00000",
                                                        "znote": "自动化测试执行",
                                                        "zsfks_czr": result_response["Samples"][i]["zsfks_czr"],
                                                        "zcdate": result_response["Samples"][i]["zcdate"],
                                                        "zplatename": "",
                                                        "z_wk_lane": "000",
                                                        "mae_zcwbh": "",
                                                        "mcc_stat": "",
                                                        "zxmbm": "",
                                                        "zbgijgdh_fld": "",
                                                        "mbt_zjob_code_item": "000000",
                                                        "zsyyn": "冼育萍",
                                                        "zzkpbm_inner": "",
                                                        "zdna_method": "",
                                                        "mo": "",
                                                        "zguid_old": "00000000000000000000000000000000",
                                                        "merar": "",
                                                        "zudate": "00000000",
                                                        "zrtpz": result_response["Samples"][i]["zrtpz"],
                                                        "zzkxh": "",
                                                        "znote_prd": "",
                                                        "zsfks_datum": result_response["Samples"][i]["zsfks_datum"],
                                                        "zload_location": "",
                                                        "znm": "",
                                                        "zindex01": "0+",
                                                        "yc_no": "",
                                                        "zreportid_item": "000000",
                                                        "zitem_no": "0000",
                                                        "zkplx": "",
                                                        "zpzxlh_bcode": "",
                                                        "zzjcwsflr": "0",
                                                        "zgxsxh": result_response["Samples"][i]["zgxsxh"],
                                                        "zrwdjgsflr": "0",
                                                        "zsfks_uzeit": result_response["Samples"][i]["zsfks_uzeit"],
                                                        "pre_zcwlx": "",
                                                        "zcwbh": (result_response["Samples"][i]["zsample"]).replace("B", "P")+"-1",
                                                        "zreceiveduzit": result_response["Samples"][i]["zreceiveduzit"],
                                                        "group_zindex": "",
                                                        "zsample": result_response["Samples"][i]["zsample"],
                                                        "zpoint": "",
                                                        "zfghsl": "",
                                                        "zcxlx_txt": "手工产线",
                                                        "zshould_sj_datum": "00000000",
                                                        "zpzxlh": "",
                                                        "maplog": "",
                                                        "zdelete_flg": "",
                                                        "zpoolingzdnum": "",
                                                        "zberaid": "",
                                                        "zpzxlh_location": "",
                                                        "pre_zsfwc_datum": "00000000",
                                                        "zshould_sj_uzeit": "000000",
                                                        "zcheck_result": "",
                                                        "zjob_code": result_response["Samples"][i]["zjob_code"],
                                                        "zfmd": "",
                                                        "zgxxh": "00000",
                                                        "zms_sfhg": "合格",
                                                        "zsfzzcw": "X"
                                                    },
                                                    {
                                                        "zplate_bcode": result_response["Samples"][i]["zplate_bcode"],
                                                        "zzjcwbs": "",
                                                        "zjob_code_source": "",
                                                        "group_zcatalo": "",
                                                        "zhhwkh_type": "",
                                                        "zzjcwpdlen": "",
                                                        "zzzcwsflr": "",
                                                        "z_jk_ybsyl": "",
                                                        "zsldfsh_uzeit": "000000",
                                                        "zsbgx": "",
                                                        "zjw_type": "",
                                                        "mbt_zjob_code": "",
                                                        "zspl_check": "",
                                                        "zbgijgdh_item": "000000",
                                                        "zhandle_actcode": "000000",
                                                        "zdata_type": "",
                                                        "zcybs": "",
                                                        "zsfjieg": "",
                                                        "zlgfsflr": "",
                                                        "z_fzh": "",
                                                        "zsfxd_datum": result_response["Samples"][i]["zsfxd_datum"],
                                                        "zname_all": result_response["Samples"][i]["zname_all"],
                                                        "zsldfsh_datum": result_response["Samples"][i]["zsldfsh_datum"],
                                                        "zplate_x": "",
                                                        "maktx": result_response["Samples"][i]["maktx"],
                                                        "yc_lx": "",
                                                        "zplate_y": "",
                                                        "zafter_excp": "",
                                                        "zkyks": result_response["Samples"][i]["zkyks"],
                                                        "zreceiveddate": result_response["Samples"][i]["zreceiveddate"],
                                                        "zsjdid": result_response["Samples"][i]["zsjdid"],
                                                        "zflowid_source": "",
                                                        "z_lane_no": "00",
                                                        "zload_machn_bcode": "",
                                                        "zsfxd_uzeit": result_response["Samples"][i]["zsfxd_uzeit"],
                                                        "zsfwc_czr": "",
                                                        "zwkpd_d": "0+",
                                                        "zbq": "",
                                                        "zcheckusername": "",
                                                        "zback_job_repid": "",
                                                        "zindex_name": "",
                                                        "matnr_before": "",
                                                        "zplate": "",
                                                        "zbgijgdh_val": "",
                                                        "zdlhwkh": "",
                                                        "zpbzbh": "",
                                                        "zgxdm": "MAB",
                                                        "ztype": "",
                                                        "zwkpd_u": "0+",
                                                        "zshdat": "00000000",
                                                        "stpch": "",
                                                        "zutime": "000000",
                                                        "pooling_sum": "",
                                                        "zinfoid": "",
                                                        "datapath": "",
                                                        "zgene": "",
                                                        "zstatus": "",
                                                        "ztj": 700,
                                                        "zeile": "0000",
                                                        "vbeln_item": "000000",
                                                        "aufnr": "",
                                                        "zjob_code_item": "000001",
                                                        "ztx": "",
                                                        "zkpwbbh": "",
                                                        "z_jk_ml": "",
                                                        "project": "",
                                                        "zrun_uzeit": "000000",
                                                        "zcwname": "",
                                                        "zgxbh": "MAB",
                                                        "zcheckdesc": "",
                                                        "zsfks": "X",
                                                        "zrun_datum": "00000000",
                                                        "ddl21": "",
                                                        "zcwlx": "血浆",
                                                        "area": "",
                                                        "water_volume": "",
                                                        "zmethod": result_response["Samples"][i]["zmethod"],
                                                        "zlane_id": "",
                                                        "zctime": result_response["Samples"][i]["zctime"],
                                                        "zback_reason": "",
                                                        "zindex": "",
                                                        "zfgsl": "00",
                                                        "zycsfwc": "",
                                                        "ztx_el": "",
                                                        "zindextype_el": "",
                                                        "zhhwkh": "",
                                                        "time_beascall": "",
                                                        "zxmnum": "",
                                                        "zplwkh": "",
                                                        "zsampling_datum": result_response["Samples"][i][
                                                            "zsampling_datum"],
                                                        "vbeln": "",
                                                        "zindex_seq": "",
                                                        "volume_sum": "",
                                                        "zcxlx": "0",
                                                        "zshnam": "",
                                                        "zreportid": "",
                                                        "zsfhg": "",
                                                        "zchangerby": "",
                                                        "zguid": result_response["Samples"][i]["zguid"],
                                                        "ztubetype": result_response["Samples"][i]["ztubetype"],
                                                        "zjob_code_up": "",
                                                        "zbglx": "",
                                                        "zcreator": result_response["Samples"][i]["zcreator"],
                                                        "zplate_sample_num": "",
                                                        "zybrwwczt": "0",
                                                        "chipholder": "",
                                                        "zindex_el": "",
                                                        "redeal_reason": "",
                                                        "zscheme": "",
                                                        "matnr": result_response["Samples"][i]["matnr"],
                                                        "zconcentration": "",
                                                        "z_zk_bfb": "",
                                                        "zdnb_uzeit": "000000",
                                                        "zclbs": "",
                                                        "kunnr": result_response["Samples"][i]["kunnr"],
                                                        "zsfxd": "X",
                                                        "zprimername": "",
                                                        "zrow": "0000000000",
                                                        "matnr_sc": "",
                                                        "zclassify": "",
                                                        "zsfwc_uzeit": "000000",
                                                        "zyblx": "",
                                                        "zdnb_datum": "00000000",
                                                        "zax_vol_spl": "",
                                                        "z_jk_cycle": "",
                                                        "zjob_status": "",
                                                        "zsfwc": "",
                                                        "theory_con": "",
                                                        "zsfwc_datum": "00000000",
                                                        "zprimer_method": "",
                                                        "zcatalo": result_response["Samples"][i]["zcatalo"],
                                                        "zindextype": "",
                                                        "zy": "",
                                                        "zsfzkp": "",
                                                        "posnr": result_response["Samples"][i]["posnr"],
                                                        "zgxlb": "",
                                                        "zseqplatform": "",
                                                        "z_lane_gs": "",
                                                        "zdatasource": result_response["Samples"][i]["zdatasource"],
                                                        "zindex_seq_el": "",
                                                        "zsampling_time": "000000",
                                                        "zsfchanw": "",
                                                        "zzk_matnr": "",
                                                        "z_lane_no_bcode": "",
                                                        "zwkh": "",
                                                        "ewbez": "产前",
                                                        "zsfxd_czr": result_response["Samples"][i]["zsfxd_czr"],
                                                        "zycdjsflr": "",
                                                        "zindex_name_el": "",
                                                        "werks": result_response["Samples"][i]["werks"],
                                                        "zsqldate": "00000000",
                                                        "zshuzt": "000000",
                                                        "zprocodenum": "",
                                                        "radat": "",
                                                        "ztestnum": "00",
                                                        "project_item": "000000",
                                                        "ztempid": "",
                                                        "zcross_od": "00000",
                                                        "znote": "自动化测试执行",
                                                        "zsfks_czr": result_response["Samples"][i]["zsfks_czr"],
                                                        "zcdate": result_response["Samples"][i]["zcdate"],
                                                        "zplatename": "",
                                                        "z_wk_lane": "000",
                                                        "mae_zcwbh": "",
                                                        "mcc_stat": "",
                                                        "zxmbm": "",
                                                        "zbgijgdh_fld": "",
                                                        "mbt_zjob_code_item": "000000",
                                                        "zsyyn": "冼育萍",
                                                        "zzkpbm_inner": "",
                                                        "zdna_method": "",
                                                        "mo": "",
                                                        "zguid_old": "00000000000000000000000000000000",
                                                        "merar": "",
                                                        "zudate": "00000000",
                                                        "zrtpz": result_response["Samples"][i]["zrtpz"],
                                                        "zzkxh": "",
                                                        "znote_prd": "",
                                                        "zsfks_datum": result_response["Samples"][i]["zsfks_datum"],
                                                        "zload_location": "",
                                                        "znm": "",
                                                        "zindex01": "0+",
                                                        "yc_no": "",
                                                        "zreportid_item": "000000",
                                                        "zitem_no": "0000",
                                                        "zkplx": "",
                                                        "zpzxlh_bcode": "",
                                                        "zzjcwsflr": "0",
                                                        "zgxsxh": result_response["Samples"][i]["zgxsxh"],
                                                        "zrwdjgsflr": "0",
                                                        "zsfks_uzeit": result_response["Samples"][i]["zsfks_uzeit"],
                                                        "pre_zcwlx": "",
                                                        "zcwbh": (result_response["Samples"][i]["zsample"]).replace("B",
                                                                                                                    "P") + "-2",
                                                        "zreceiveduzit": result_response["Samples"][i]["zreceiveduzit"],
                                                        "group_zindex": "",
                                                        "zsample": result_response["Samples"][i]["zsample"],
                                                        "zpoint": "",
                                                        "zfghsl": "",
                                                        "zcxlx_txt": "手工产线",
                                                        "zshould_sj_datum": "00000000",
                                                        "zpzxlh": "",
                                                        "maplog": "",
                                                        "zdelete_flg": "",
                                                        "zpoolingzdnum": "",
                                                        "zberaid": "",
                                                        "zpzxlh_location": "",
                                                        "pre_zsfwc_datum": "00000000",
                                                        "zshould_sj_uzeit": "000000",
                                                        "zcheck_result": "",
                                                        "zjob_code": result_response["Samples"][i]["zjob_code"],
                                                        "zfmd": "",
                                                        "zgxxh": "00000",
                                                        "zms_sfhg": "合格",
                                                        "zsfzzcw": "X"
                                                    },
                                                    {
                                                        "zplate_bcode": result_response["Samples"][i]["zplate_bcode"],
                                                        "zzjcwbs": "",
                                                        "zjob_code_source": "",
                                                        "group_zcatalo": "",
                                                        "zhhwkh_type": "",
                                                        "zzjcwpdlen": "",
                                                        "zzzcwsflr": "",
                                                        "z_jk_ybsyl": "",
                                                        "zsldfsh_uzeit": "000000",
                                                        "zsbgx": "",
                                                        "zjw_type": "",
                                                        "mbt_zjob_code": "",
                                                        "zspl_check": "",
                                                        "zbgijgdh_item": "000000",
                                                        "zhandle_actcode": "000000",
                                                        "zdata_type": "",
                                                        "zcybs": "",
                                                        "zsfjieg": "",
                                                        "zlgfsflr": "",
                                                        "z_fzh": "",
                                                        "zsfxd_datum": result_response["Samples"][i]["zsfxd_datum"],
                                                        "zname_all": result_response["Samples"][i]["zname_all"],
                                                        "zsldfsh_datum": result_response["Samples"][i]["zsldfsh_datum"],
                                                        "zplate_x": "",
                                                        "maktx": result_response["Samples"][i]["maktx"],
                                                        "yc_lx": "",
                                                        "zplate_y": "",
                                                        "zafter_excp": "",
                                                        "zkyks": result_response["Samples"][i]["zkyks"],
                                                        "zreceiveddate": result_response["Samples"][i]["zreceiveddate"],
                                                        "zsjdid": result_response["Samples"][i]["zsjdid"],
                                                        "zflowid_source": "",
                                                        "z_lane_no": "00",
                                                        "zload_machn_bcode": "",
                                                        "zsfxd_uzeit": result_response["Samples"][i]["zsfxd_uzeit"],
                                                        "zsfwc_czr": "",
                                                        "zwkpd_d": "0+",
                                                        "zbq": "",
                                                        "zcheckusername": "",
                                                        "zback_job_repid": "",
                                                        "zindex_name": "",
                                                        "matnr_before": "",
                                                        "zplate": "",
                                                        "zbgijgdh_val": "",
                                                        "zdlhwkh": "",
                                                        "zpbzbh": "",
                                                        "zgxdm": "MAB",
                                                        "ztype": "",
                                                        "zwkpd_u": "0+",
                                                        "zshdat": "00000000",
                                                        "stpch": "",
                                                        "zutime": "000000",
                                                        "pooling_sum": "",
                                                        "zinfoid": "",
                                                        "datapath": "",
                                                        "zgene": "",
                                                        "zstatus": "",
                                                        "ztj": 500,
                                                        "zeile": "0000",
                                                        "vbeln_item": "000000",
                                                        "aufnr": "",
                                                        "zjob_code_item": "000001",
                                                        "ztx": "",
                                                        "zkpwbbh": "",
                                                        "z_jk_ml": "",
                                                        "project": "",
                                                        "zrun_uzeit": "000000",
                                                        "zcwname": "",
                                                        "zgxbh": "MAB",
                                                        "zcheckdesc": "",
                                                        "zsfks": "X",
                                                        "zrun_datum": "00000000",
                                                        "ddl21": "",
                                                        "zcwlx": "血浆",
                                                        "area": "",
                                                        "water_volume": "",
                                                        "zmethod": result_response["Samples"][i]["zmethod"],
                                                        "zlane_id": "",
                                                        "zctime": result_response["Samples"][i]["zctime"],
                                                        "zback_reason": "",
                                                        "zindex": "",
                                                        "zfgsl": "00",
                                                        "zycsfwc": "",
                                                        "ztx_el": "",
                                                        "zindextype_el": "",
                                                        "zhhwkh": "",
                                                        "time_beascall": "",
                                                        "zxmnum": "",
                                                        "zplwkh": "",
                                                        "zsampling_datum": result_response["Samples"][i][
                                                            "zsampling_datum"],
                                                        "vbeln": "",
                                                        "zindex_seq": "",
                                                        "volume_sum": "",
                                                        "zcxlx": "0",
                                                        "zshnam": "",
                                                        "zreportid": "",
                                                        "zsfhg": "",
                                                        "zchangerby": "",
                                                        "zguid": result_response["Samples"][i]["zguid"],
                                                        "ztubetype": result_response["Samples"][i]["ztubetype"],
                                                        "zjob_code_up": "",
                                                        "zbglx": "",
                                                        "zcreator": result_response["Samples"][i]["zcreator"],
                                                        "zplate_sample_num": "",
                                                        "zybrwwczt": "0",
                                                        "chipholder": "",
                                                        "zindex_el": "",
                                                        "redeal_reason": "",
                                                        "zscheme": "",
                                                        "matnr": result_response["Samples"][i]["matnr"],
                                                        "zconcentration": "",
                                                        "z_zk_bfb": "",
                                                        "zdnb_uzeit": "000000",
                                                        "zclbs": "",
                                                        "kunnr": result_response["Samples"][i]["kunnr"],
                                                        "zsfxd": "X",
                                                        "zprimername": "",
                                                        "zrow": "0000000000",
                                                        "matnr_sc": "",
                                                        "zclassify": "",
                                                        "zsfwc_uzeit": "000000",
                                                        "zyblx": "",
                                                        "zdnb_datum": "00000000",
                                                        "zax_vol_spl": "",
                                                        "z_jk_cycle": "",
                                                        "zjob_status": "",
                                                        "zsfwc": "",
                                                        "theory_con": "",
                                                        "zsfwc_datum": "00000000",
                                                        "zprimer_method": "",
                                                        "zcatalo": result_response["Samples"][i]["zcatalo"],
                                                        "zindextype": "",
                                                        "zy": "",
                                                        "zsfzkp": "",
                                                        "posnr": result_response["Samples"][i]["posnr"],
                                                        "zgxlb": "",
                                                        "zseqplatform": "",
                                                        "z_lane_gs": "",
                                                        "zdatasource": result_response["Samples"][i]["zdatasource"],
                                                        "zindex_seq_el": "",
                                                        "zsampling_time": "000000",
                                                        "zsfchanw": "",
                                                        "zzk_matnr": "",
                                                        "z_lane_no_bcode": "",
                                                        "zwkh": "",
                                                        "ewbez": "产前",
                                                        "zsfxd_czr": result_response["Samples"][i]["zsfxd_czr"],
                                                        "zycdjsflr": "",
                                                        "zindex_name_el": "",
                                                        "werks": result_response["Samples"][i]["werks"],
                                                        "zsqldate": "00000000",
                                                        "zshuzt": "000000",
                                                        "zprocodenum": "",
                                                        "radat": "",
                                                        "ztestnum": "00",
                                                        "project_item": "000000",
                                                        "ztempid": "",
                                                        "zcross_od": "00000",
                                                        "znote": "自动化测试执行",
                                                        "zsfks_czr": result_response["Samples"][i]["zsfks_czr"],
                                                        "zcdate": result_response["Samples"][i]["zcdate"],
                                                        "zplatename": "",
                                                        "z_wk_lane": "000",
                                                        "mae_zcwbh": "",
                                                        "mcc_stat": "",
                                                        "zxmbm": "",
                                                        "zbgijgdh_fld": "",
                                                        "mbt_zjob_code_item": "000000",
                                                        "zsyyn": "冼育萍",
                                                        "zzkpbm_inner": "",
                                                        "zdna_method": "",
                                                        "mo": "",
                                                        "zguid_old": "00000000000000000000000000000000",
                                                        "merar": "",
                                                        "zudate": "00000000",
                                                        "zrtpz": result_response["Samples"][i]["zrtpz"],
                                                        "zzkxh": "",
                                                        "znote_prd": "",
                                                        "zsfks_datum": result_response["Samples"][i]["zsfks_datum"],
                                                        "zload_location": "",
                                                        "znm": "",
                                                        "zindex01": "0+",
                                                        "yc_no": "",
                                                        "zreportid_item": "000000",
                                                        "zitem_no": "0000",
                                                        "zkplx": "",
                                                        "zpzxlh_bcode": "",
                                                        "zzjcwsflr": "0",
                                                        "zgxsxh": result_response["Samples"][i]["zgxsxh"],
                                                        "zrwdjgsflr": "0",
                                                        "zsfks_uzeit": result_response["Samples"][i]["zsfks_uzeit"],
                                                        "pre_zcwlx": "",
                                                        "zcwbh": (result_response["Samples"][i]["zsample"]).replace("B",
                                                                                                                    "P") + "-3",
                                                        "zreceiveduzit": result_response["Samples"][i]["zreceiveduzit"],
                                                        "group_zindex": "",
                                                        "zsample": result_response["Samples"][i]["zsample"],
                                                        "zpoint": "",
                                                        "zfghsl": "",
                                                        "zcxlx_txt": "手工产线",
                                                        "zshould_sj_datum": "00000000",
                                                        "zpzxlh": "",
                                                        "maplog": "",
                                                        "zdelete_flg": "",
                                                        "zpoolingzdnum": "",
                                                        "zberaid": "",
                                                        "zpzxlh_location": "",
                                                        "pre_zsfwc_datum": "00000000",
                                                        "zshould_sj_uzeit": "000000",
                                                        "zcheck_result": "",
                                                        "zjob_code": result_response["Samples"][i]["zjob_code"],
                                                        "zfmd": "",
                                                        "zgxxh": "00000",
                                                        "zms_sfhg": "合格",
                                                        "zsfzzcw": "X"
                                                    }
                                                ],
                                                "tab2": {
                                                    "zjob_code": result_response["Samples"][i]["zjob_code"],
                                                    "zjob_code_item": result_response["Samples"][i]["zjob_code_item"],
                                                    "zgxbh": "MAB",
                                                    "zguid": result_response["Samples"][i]["zguid"],
                                                    "zcatalo": result_response["Samples"][i]["zcatalo"],
                                                    "zsample": result_response["Samples"][i]["zsample"],
                                                    "zsjdid": result_response["Samples"][i]["zsjdid"],
                                                    "zms_sfhg": "合格",
                                                    "znote_prd": ""
                                                }
                                            }
                                        }
                                    for i in range(len(result_response["Samples"]))
                                    ]
                                data = {
                                    "task": task,
                                    "zgxbh": "MAB",
                                    "zjob_code": self.zjob_code,
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_XueJiangFenLi"
                                }
                                response = self.nifty_res.post_request("/presap/webintf.do?method=msUpdateTaskResult",
                                                                       data=urlencode(data))
                                if response.status_code == 200 and response.json()["code"] == "200":
                                    logger.info("血浆分离任务单结果录入成功！")
                                    data = {
                                        "task": {
                                            "zpzxlh_bcode": "",
                                            "zcjdat": "",
                                            "ybbhrq": "",
                                            "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                            "zsfwc_czr": "",
                                            "zsfks_czr": "",
                                            "zcdate": "",
                                            "chkdate": "",
                                            "matnr": rwd_data["matnr"],
                                            "zseqplatform": "",
                                            "zsfks_uzeit": rwd_data["zsfks_uzeit"],
                                            "zdlhwkh": "",
                                            "qdfkrq": "",
                                            "zsampling_datum": "",
                                            "zdlvdate": """ """,
                                            "zgxbh": "MAB",
                                            "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                            "zsigndate": "",
                                            "zsfks": "",
                                            "zsyyn": rwd_data["zsyyn"],
                                            "syncdate": "",
                                            "ddl21": "",
                                            "yctjrq": "",
                                            "zcxlx": rwd_data["zcxlx"],
                                            "zjob_begin_datum": "",
                                            "zsfwc_uzeit": "",
                                            "zsample_quantity": rwd_data["zsample_quantity"],
                                            "zmethod": rwd_data["zmethod"],
                                            "zsfxd_czr": rwd_data["zsfxd_czr"],
                                            "zudate": "",
                                            "addate": "",
                                            "zycdjsflr": rwd_data["zycdjsflr"],
                                            "zsfxd_datum": rwd_data["zsfxd_datum"],
                                            "werks": rwd_data["werks"],
                                            "zsfks_datum": rwd_data["zsfks_datum"],
                                            "ycwcrq": "",
                                            "pre_zsfwc_datum": "",
                                            "zsfwc": "",
                                            "pre_zjob_code": "",
                                            "maktx": rwd_data["maktx"],
                                            "zsfwc_datum": "",
                                            "zupdat": "",
                                            "zjob_code": rwd_data["zjob_code"],
                                            "gt_action": "",
                                            "zcbdate": "",
                                            "gt_sendtime": "",
                                            "zfinish_time": "",
                                            "_key": 1,
                                            "_id": 1
                                        },
                                        "token": self.token,
                                        "menuId": "MSTaskProduct_XueJiangFenLi"
                                    }
                                    rwd_detail_response = self.nifty_res.post_request(
                                        "/presap/webintf.do?method=findJobSampleMessageMS", data=urlencode(data))
                                    if rwd_detail_response.json()["code"] == "200" and \
                                            (rwd_detail_response.json()["data"][0])["zjob_code"] == rwd_data[
                                        "zjob_code"]:
                                        rwd_detail_data = rwd_detail_response.json()["data"]
                                        logger.info("血浆分离任务单详情查询成功！")
                                    #     data1 = [
                                    #     {
                                    #         "no": f"{i+1:06d}",
                                    #         "zpzxlh": "",
                                    #         "zmethod": "1755.01.01",
                                    #         "zproduct": rwd_detail_data[i]["matnr"],
                                    #         "zmethod_name": "产前_血浆分离",
                                    #         "zplate": "",
                                    #         "werks": "A020",
                                    #         "zindex": 0,
                                    #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                    #         "zgxdm": "MAB",
                                    #         "tab": "WL",
                                    #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                    #         "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                    #         "zcatalo": rwd_detail_data[i]["zcatalo"],
                                    #         "_key": i+1,
                                    #         "zwkh": "",
                                    #         "_id": i+2,
                                    #         "rfpnt": "",
                                    #         "matnr": "1000000731",
                                    #         "maktx": "AT01MT2-H-X#1000UL国产导电吸头/品牌&杭州金源/规格&96个/",
                                    #         "datuv": "99990606",
                                    #         "zshl": "0.00 ",
                                    #         "zcksjyl": 2.2,
                                    #         "zbzdw": "EA",
                                    #         "zbzyl": 2.2,
                                    #         "charg": "0000021055",
                                    #         "zsjyl": 2.2,
                                    #         "zcomment": "",
                                    #         "zybs": len(rwd_detail_data),
                                    #         "zcost_condition": "1"
                                    #     }
                                    #         for i in range(len(rwd_detail_data))
                                    # ]
                                    #     data2 = [
                                    #     {
                                    #         "no": f"{i+1:06d}",
                                    #         "zpzxlh": "",
                                    #         "zmethod": "1755.01.01",
                                    #         "zproduct": rwd_detail_data[i]["matnr"],
                                    #         "zmethod_name": "产前_血浆分离",
                                    #         "zplate": "",
                                    #         "werks": "A020",
                                    #         "zindex": 0,
                                    #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                    #         "zgxdm": "MAB",
                                    #         "tab": "YQ",
                                    #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                    #         "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                    #         "zcatalo": rwd_detail_data[i]["zcatalo"],
                                    #         "matnr": "54191",
                                    #         "maktx": "离心机（5427R）",
                                    #         "zbzyl": 0,
                                    #         "zbzdw": "MIN",
                                    #         "zsjyl": 0.007,
                                    #         "zcost_condition": "1"
                                    #     }
                                    #         for i in range(len(rwd_detail_data))
                                    #     ]
                                    #     data3 = [
                                    #     {
                                    #         "no": f"{i+1:06d}",
                                    #         "zpzxlh": "",
                                    #         "zmethod": "1755.01.01",
                                    #         "zproduct": rwd_detail_data[i]["matnr"],
                                    #         "zmethod_name": "产前_血浆分离",
                                    #         "zplate": "",
                                    #         "werks": "A020",
                                    #         "zindex": 0,
                                    #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                    #         "zgxdm": "MAB",
                                    #         "tab": "YQ",
                                    #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                    #         "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                    #         "zcatalo": rwd_detail_data[i]["zcatalo"],
                                    #         "matnr": "25395",
                                    #         "maktx": "离心机",
                                    #         "zbzyl": 0,
                                    #         "zbzdw": "MIN",
                                    #         "zsjyl": 0.016,
                                    #         "zcost_condition": "1"
                                    #     }
                                    #         for i in range(len(rwd_detail_data))
                                    #     ]
                                    #     data4 = [
                                    #     {
                                    #         "no": f"{i+1:06d}",
                                    #         "zpzxlh": "",
                                    #         "zmethod": "1755.01.01",
                                    #         "zproduct": rwd_detail_data[i]["matnr"],
                                    #         "zmethod_name": "产前_血浆分离",
                                    #         "zplate": "",
                                    #         "werks": "A020",
                                    #         "zindex": 0,
                                    #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                    #         "zgxdm": "MAB",
                                    #         "tab": "YQ",
                                    #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                    #         "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                    #         "zcatalo": rwd_detail_data[i]["zcatalo"],
                                    #         "matnr": "31027",
                                    #         "maktx": "离心机 TGL-16",
                                    #         "zbzyl": 0,
                                    #         "zbzdw": "MIN",
                                    #         "zsjyl": 0.014,
                                    #         "zcost_condition": "1"
                                    #     }
                                    #         for i in range(len(rwd_detail_data))
                                    #     ]
                                    #     data5 = [
                                    #     {
                                    #         "no": f"{i+1:06d}",
                                    #         "zpzxlh": "",
                                    #         "zmethod": "1755.01.01",
                                    #         "zproduct": rwd_detail_data[i]["matnr"],
                                    #         "zmethod_name": "产前_血浆分离",
                                    #         "zplate": "",
                                    #         "werks": "A020",
                                    #         "zindex": 0,
                                    #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                    #         "zgxdm": "MAB",
                                    #         "tab": "YQ",
                                    #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                    #         "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                    #         "zcatalo": rwd_detail_data[i]["zcatalo"],
                                    #         "matnr": "49030",
                                    #         "maktx": "血浆分离工作站",
                                    #         "zbzyl": 0,
                                    #         "zbzdw": "MIN",
                                    #         "zsjyl": 0.013,
                                    #         "zcost_condition": "1"
                                    #     }
                                    #         for i in range(len(rwd_detail_data))
                                    #     ]
                                    #     data6 = [
                                    #         {
                                    #             "no": f"{i+1:06d}",
                                    #             "zpzxlh": "",
                                    #             "zmethod": "1755.01.01",
                                    #             "zproduct": rwd_detail_data[i]["matnr"],
                                    #             "zmethod_name": "产前_血浆分离",
                                    #             "zplate": "",
                                    #             "werks": "A020",
                                    #             "zindex": 0,
                                    #             "zproduct_name": rwd_detail_data[i]["maktx"],
                                    #             "zgxdm": "MAB",
                                    #             "tab": "GS",
                                    #             "zjob_code": rwd_detail_data[i]["zjob_code"],
                                    #             "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                    #             "zcatalo": rwd_detail_data[i]["zcatalo"],
                                    #             "matnr": "A0201755.01.0101",
                                    #             "zbzyl": 0,
                                    #             "zbzdw": "MIN",
                                    #             "zsjyl": 2.824,
                                    #             "zcost_condition": "1"
                                    #         }
                                    #         for i in range(len(rwd_detail_data))
                                    #     ]
                                    #     total_data = [
                                    #         {
                                    #             "zbzdw": "EA",
                                    #             "zbzyl": str(2.2*len(rwd_detail_data))+"000000 ",
                                    #             "no": "0",
                                    #             "zshl": "0.00 ",
                                    #             "matnr": "1000000731",
                                    #             "zsjyl": str(2.2*len(rwd_detail_data))+"000000 ",
                                    #             "zsjdw": "EA",
                                    #             "zmethod_name": "产前_血浆分离",
                                    #             "zcksjyl": str(2.2*len(rwd_detail_data))+"000000 ",
                                    #             "zplate": "",
                                    #             "zcomment": "",
                                    #             "zgxdm": "MAB",
                                    #             "tab": "WL",
                                    #             "datuv": "99990606",
                                    #             "zpzxlh": "",
                                    #             "zmethod": "1755.01.01",
                                    #             "zproduct": "DX1331",
                                    #             "zybs": len(rwd_detail_data),
                                    #             "mtart": "Z002",
                                    #             "werks": "A020",
                                    #             "ddtext": "",
                                    #             "charg": "0000021055",
                                    #             "zindex": "1 ",
                                    #             "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                    #             "rfpnt": "",
                                    #             "maktx": "AT01MT2-H-X#1000UL国产导电吸头/品牌&杭州金源/规格&96个/",
                                    #             "zjob_code": self.zjob_code,
                                    #             "zjob_code_item": "000000",
                                    #             "zcatalo": "合计",
                                    #             "cksyjl": 2.2*len(rwd_detail_data),
                                    #             "_key": "",
                                    #             "_id": 1,
                                    #             "zcost_condition": "1"
                                    #         },
                                    #         {
                                    #             "zbzyl": "0.0000000 ",
                                    #             "zbzdw": "MIN",
                                    #             "no": "0",
                                    #             "zpzxlh": "",
                                    #             "zmethod": "1755.01.01",
                                    #             "zproduct": "DX1331",
                                    #             "zybs": str(len(rwd_detail_data))+" ",
                                    #             "matnr": "54191",
                                    #             "zsjyl": str(0.007*len(rwd_detail_data))+"0000 ",
                                    #             "zsjdw": "MIN",
                                    #             "zmethod_name": "产前_血浆分离",
                                    #             "zcksjyl": str(0.007*len(rwd_detail_data))+"0000 ",
                                    #             "zplate": "",
                                    #             "werks": "A020",
                                    #             "zcomment": "",
                                    #             "zindex": "1 ",
                                    #             "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                    #             "zgxdm": "MAB",
                                    #             "tab": "YQ",
                                    #             "maktx": "离心机（5427R）",
                                    #             "zjob_code": self.zjob_code,
                                    #             "zjob_code_item": "000000",
                                    #             "zcatalo": "合计",
                                    #             "zcost_condition": "1"
                                    #         },
                                    #         {
                                    #             "zbzyl": "0.0000000 ",
                                    #             "zbzdw": "MIN",
                                    #             "no": "0",
                                    #             "zpzxlh": "",
                                    #             "zmethod": "1755.01.01",
                                    #             "zproduct": "DX1331",
                                    #             "zybs": str(len(rwd_detail_data))+" ",
                                    #             "matnr": "25395",
                                    #             "zsjyl": str(0.016*len(rwd_detail_data))+"0000 ",
                                    #             "zsjdw": "MIN",
                                    #             "zmethod_name": "产前_血浆分离",
                                    #             "zcksjyl": str(0.016*len(rwd_detail_data))+"0000 ",
                                    #             "zplate": "",
                                    #             "werks": "A020",
                                    #             "zcomment": "",
                                    #             "zindex": "2 ",
                                    #             "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                    #             "zgxdm": "MAB",
                                    #             "tab": "YQ",
                                    #             "maktx": "离心机",
                                    #             "zjob_code": self.zjob_code,
                                    #             "zjob_code_item": "000000",
                                    #             "zcatalo": "合计",
                                    #             "zcost_condition": "1"
                                    #         },
                                    #         {
                                    #             "zbzyl": "0.0000000 ",
                                    #             "zbzdw": "MIN",
                                    #             "no": "0",
                                    #             "zpzxlh": "",
                                    #             "zmethod": "1755.01.01",
                                    #             "zproduct": "DX1331",
                                    #             "zybs": str(len(rwd_detail_data))+" ",
                                    #             "matnr": "31027",
                                    #             "zsjyl": str(0.014*len(rwd_detail_data))+"0000 ",
                                    #             "zsjdw": "MIN",
                                    #             "zmethod_name": "产前_血浆分离",
                                    #             "zcksjyl": str(0.014*len(rwd_detail_data))+"0000 ",
                                    #             "zplate": "",
                                    #             "werks": "A020",
                                    #             "zcomment": "",
                                    #             "zindex": "3 ",
                                    #             "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                    #             "zgxdm": "MAB",
                                    #             "tab": "YQ",
                                    #             "maktx": "离心机+TGL-16",
                                    #             "zjob_code": self.zjob_code,
                                    #             "zjob_code_item": "000000",
                                    #             "zcatalo": "合计",
                                    #             "zcost_condition": "1"
                                    #         },
                                    #         {
                                    #             "zbzyl": "0.0000000 ",
                                    #             "zbzdw": "MIN",
                                    #             "no": "0",
                                    #             "zpzxlh": "",
                                    #             "zmethod": "1755.01.01",
                                    #             "zproduct": "DX1331",
                                    #             "zybs": str(len(rwd_detail_data))+" ",
                                    #             "matnr": "49030",
                                    #             "zsjyl": str(0.013*len(rwd_detail_data))+"0000 ",
                                    #             "zsjdw": "MIN",
                                    #             "zmethod_name": "产前_血浆分离",
                                    #             "zcksjyl": str(0.013*len(rwd_detail_data))+"0000 ",
                                    #             "zplate": "",
                                    #             "werks": "A020",
                                    #             "zcomment": "",
                                    #             "zindex": "4 ",
                                    #             "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                    #             "zgxdm": "MAB",
                                    #             "tab": "YQ",
                                    #             "maktx": "血浆分离工作站",
                                    #             "zjob_code": self.zjob_code,
                                    #             "zjob_code_item": "000000",
                                    #             "zcatalo": "合计",
                                    #             "zcost_condition": "1"
                                    #         },
                                    #         {
                                    #             "zbzyl": "0.0000000 ",
                                    #             "zbzdw": "MIN",
                                    #             "no": "0",
                                    #             "zpzxlh": "",
                                    #             "zmethod": "1755.01.01",
                                    #             "zproduct": "DX1331",
                                    #             "zybs": str(len(rwd_detail_data))+" ",
                                    #             "matnr": "A0201755.01.0101",
                                    #             "zsjyl": str(2.824*len(rwd_detail_data))+"0000 ",
                                    #             "zsjdw": "MIN",
                                    #             "zmethod_name": "产前_血浆分离",
                                    #             "zcksjyl": str(2.824*len(rwd_detail_data))+"0000 ",
                                    #             "zplate": "",
                                    #             "werks": "A020",
                                    #             "zcomment": "",
                                    #             "zindex": "1 ",
                                    #             "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                    #             "zgxdm": "MAB",
                                    #             "tab": "GS",
                                    #             "maktx": "",
                                    #             "zjob_code": self.zjob_code,
                                    #             "zjob_code_item": "000000",
                                    #             "zcatalo": "合计",
                                    #             "zcost_condition": "1"
                                    #         }
                                    #     ]
                                    #     data = list(zip(data1, data2, data3, data4, data5, data6, total_data))
                                    #     wuliao_data = {
                                    #         "data": [data[0][0]],
                                    #         "token": self.token,
                                    #         "menuId": "MSTaskProduct_XueJiangFenLi"
                                    #     }
                                    #     wuliao_response =  self.nifty_res.post_request("/presap/webintf.do?method=save_liao_gong_fei_nifty",
                                    #                                        data=urlencode(wuliao_data))
                                        # if wuliao_response.json()["code"] == "200" and wuliao_response.json()["msg"] == "success":
                                        #     logger.info(f"血浆分离料工费结果录入成功！")
                                        data = [
                                            {
                                                "zplate_bcode": rwd_detail_data[i]["zplate_bcode"],
                                                "zzjcwbs": "",
                                                "zjob_code_source": "",
                                                "group_zcatalo": "",
                                                "zhhwkh_type": "",
                                                "zzjcwpdlen": "",
                                                "zzzcwsflr": "",
                                                "z_jk_ybsyl": "",
                                                "zsldfsh_uzeit": rwd_detail_data[i]["zsldfsh_uzeit"],
                                                "zsbgx": "",
                                                "zjw_type": "",
                                                "mbt_zjob_code": "",
                                                "zspl_check": "",
                                                "zbgijgdh_item": rwd_detail_data[i]["zbgijgdh_item"],
                                                "zhandle_actcode": rwd_detail_data[i]["zhandle_actcode"],
                                                "zdata_type": "",
                                                "zcybs": "",
                                                "zsfjieg": "",
                                                "zlgfsflr": "",
                                                "z_fzh": "",
                                                "zsfxd_datum": rwd_detail_data[i]["zsfxd_datum"],
                                                "zname_all": rwd_detail_data[i]["zname_all"],
                                                "zsldfsh_datum": rwd_detail_data[i]["zsldfsh_datum"],
                                                "zplate_x": "",
                                                "maktx": rwd_detail_data[i]["maktx"],
                                                "yc_lx": "",
                                                "zplate_y": "",
                                                "zafter_excp": "",
                                                "zkyks": "X",
                                                "zreceiveddate": rwd_detail_data[i]["zreceiveddate"],
                                                "zsjdid": rwd_detail_data[i]["zsjdid"],
                                                "zflowid_source": "",
                                                "z_lane_no": "00",
                                                "zload_machn_bcode": "",
                                                "zcatalo2": rwd_detail_data[i]["zcatalo2"],
                                                "zsfxd_uzeit": rwd_detail_data[i]["zsfxd_uzeit"],
                                                "zsfwc_czr": "",
                                                "zcatalo3": rwd_detail_data[i]["zcatalo3"],
                                                "zwkpd_d": rwd_detail_data[i]["zwkpd_d"],
                                                "zcatalo1": rwd_detail_data[i]["zcatalo1"],
                                                "zbq": "",
                                                "zcheckusername": "",
                                                "zback_job_repid": "",
                                                "zindex_name": "",
                                                "matnr_before": "",
                                                "zplate": "",
                                                "zbgijgdh_val": "",
                                                "zdlhwkh": "",
                                                "zpbzbh": "",
                                                "zgxdm": "MAB",
                                                "ztype": "",
                                                "zwkpd_u": "0 ",
                                                "zshdat": "00000000",
                                                "stpch": "",
                                                "zutime": "000000",
                                                "pooling_sum": "",
                                                "zinfoid": "",
                                                "datapath": "",
                                                "zgene": "",
                                                "zstatus": "",
                                                "ztj": "",
                                                "zeile": "0000",
                                                "vbeln_item": "000000",
                                                "aufnr": "",
                                                "zjob_code_item": rwd_detail_data[i]["zjob_code_item"],
                                                "ztx": "",
                                                "zkpwbbh": "",
                                                "z_jk_ml": "",
                                                "project": "",
                                                "zrun_uzeit": "000000",
                                                "zcwname": "",
                                                "zgxbh": "MAB",
                                                "zcheckdesc": "",
                                                "zsfks": "X",
                                                "zrun_datum": "00000000",
                                                "ddl21": "",
                                                "zcwlx": "",
                                                "area": "",
                                                "water_volume": "",
                                                "zmethod": "1755.01.01",
                                                "zlane_id": "",
                                                "zctime": rwd_detail_data[i]["zctime"],
                                                "zback_reason": "",
                                                "zindex": "",
                                                "zfgsl": "00",
                                                "zycsfwc": "",
                                                "ztx_el": "",
                                                "zindextype_el": "",
                                                "zhhwkh": "",
                                                "time_beascall": "",
                                                "zxmnum": "",
                                                "zplwkh": "",
                                                "zsampling_datum": rwd_detail_data[i]["zsampling_datum"],
                                                "zms_sfhg": rwd_detail_data[i]["zms_sfhg"],
                                                "vbeln": "",
                                                "zindex_seq": "",
                                                "volume_sum": "",
                                                "zcxlx": "0",
                                                "zshnam": "",
                                                "zreportid": "",
                                                "zsfhg": rwd_detail_data[i]["zsfhg"],
                                                "zchangerby": "",
                                                "zguid": rwd_detail_data[i]["zguid"],
                                                "ztubetype": rwd_detail_data[i]["ztubetype"],
                                                "zjob_code_up": "",
                                                "zbglx": "",
                                                "zcreator": rwd_detail_data[i]["zcreator"],
                                                "zplate_sample_num": "",
                                                "zybrwwczt": "0",
                                                "chipholder": "",
                                                "zindex_el": "",
                                                "redeal_reason": "",
                                                "zscheme": "",
                                                "matnr": rwd_detail_data[i]["matnr"],
                                                "zconcentration": "",
                                                "z_zk_bfb": "",
                                                "zdnb_uzeit": "000000",
                                                "zclbs": "",
                                                "kunnr": rwd_detail_data[i]["kunnr"],
                                                "zsfxd": "X",
                                                "zprimername": "",
                                                "zrow": "0000000000",
                                                "matnr_sc": "",
                                                "zclassify": "",
                                                "zsfwc_uzeit": "000000",
                                                "zyblx": "",
                                                "zdnb_datum": "00000000",
                                                "zax_vol_spl": "",
                                                "z_jk_cycle": "",
                                                "zjob_status": "",
                                                "zsfwc": "",
                                                "theory_con": "",
                                                "zsfwc_datum": "00000000",
                                                "zprimer_method": "",
                                                "zcatalo": rwd_detail_data[i]["zcatalo"],
                                                "zindextype": "",
                                                "zy": "",
                                                "zsfzkp": "",
                                                "posnr": rwd_detail_data[i]["posnr"],
                                                "zgxlb": "",
                                                "zseqplatform": "",
                                                "z_lane_gs": "",
                                                "zdatasource": rwd_detail_data[i]["zdatasource"],
                                                "zindex_seq_el": "",
                                                "zsampling_time": "000000",
                                                "zsfchanw": "X",
                                                "zzk_matnr": "",
                                                "z_lane_no_bcode": "",
                                                "zwkh": "",
                                                "ewbez": rwd_detail_data[i]["ewbez"],
                                                "zsfxd_czr": rwd_detail_data[i]["zsfxd_czr"],
                                                "zycdjsflr": "",
                                                "zindex_name_el": "",
                                                "werks": "A020",
                                                "zsqldate": "00000000",
                                                "zshuzt": "000000",
                                                "zprocodenum": "",
                                                "radat": "",
                                                "ztestnum": "00",
                                                "project_item": "000000",
                                                "ztempid": "",
                                                "zcross_od": "00000",
                                                "znote": "自动化测试执行",
                                                "zsfks_czr": rwd_detail_data[i]["zsfks_czr"],
                                                "zcdate": rwd_detail_data[i]["zcdate"],
                                                "zplatename": "",
                                                "z_wk_lane": "000",
                                                "mae_zcwbh": "",
                                                "mcc_stat": "",
                                                "zxmbm": "",
                                                "zbgijgdh_fld": "",
                                                "mbt_zjob_code_item": "000000",
                                                "zsyyn": rwd_detail_data[i]["zsyyn"],
                                                "zzkpbm_inner": "",
                                                "zdna_method": "",
                                                "mo": "",
                                                "zguid_old": "00000000000000000000000000000000",
                                                "merar": "",
                                                "zudate": "00000000",
                                                "zrtpz": rwd_detail_data[i]["zrtpz"],
                                                "zzkxh": "",
                                                "znote_prd": "",
                                                "zsfks_datum": rwd_detail_data[i]["zsfks_datum"],
                                                "zload_location": "",
                                                "znm": "",
                                                "zindex01": "0 ",
                                                "yc_no": "",
                                                "zreportid_item": "000000",
                                                "zitem_no": "0000",
                                                "zkplx": "",
                                                "zpzxlh_bcode": "",
                                                "zzjcwsflr": "0",
                                                "zgxsxh": "00010",
                                                "zrwdjgsflr": "0",
                                                "zsfks_uzeit": rwd_detail_data[i]["zsfks_uzeit"],
                                                "pre_zcwlx": "",
                                                "zcwbh": (rwd_detail_data[i]["zsample"]).replace("B", "P") + "-3",
                                                "zreceiveduzit": rwd_detail_data[i]["zreceiveduzit"],
                                                "group_zindex": "",
                                                "zsample": rwd_detail_data[i]["zsample"],
                                                "zpoint": "",
                                                "zfghsl": "",
                                                "zcxlx_txt": rwd_detail_data[i]["zcxlx_txt"],
                                                "zshould_sj_datum": "00000000",
                                                "zpzxlh": "",
                                                "maplog": "",
                                                "zdelete_flg": "",
                                                "zpoolingzdnum": "",
                                                "zberaid": "",
                                                "zpzxlh_location": "",
                                                "pre_zsfwc_datum": "00000000",
                                                "zshould_sj_uzeit": "000000",
                                                "zcheck_result": "",
                                                "zjob_code": rwd_detail_data[i]["zjob_code"],
                                                "ztj1": "300",
                                                "ztj3": "500",
                                                "zfmd": "",
                                                "ztj2": "700",
                                                "zgxxh": "00000",
                                                "_id": 1
                                            }
                                            for i in range(len(rwd_detail_data))
                                        ]
                                        complete_data = {
                                                "req": data,
                                                "token": self.token,
                                                "menuId": "MSTaskProduct_XueJiangFenLi"
                                        }
                                        complete_response = self.nifty_res.post_request(
                                            "/presap/webintf.do?method=task_sample_item_finishms", data=urlencode(complete_data))
                                        if complete_response.json()["code"] == "200" and complete_response.json()["msg"] == "SAP数据库更新成功!":
                                            logger.info("血浆分离任务完成成功！")
                                            return self.zjob_code
                                        else:
                                            logger.error(
                                                f"血浆分离任务完成失败！response：{complete_response.json()}")
                                            raise Exception(
                                                f"血浆分离任务完成失败！response：{complete_response.json()}")
                                        # else:
                                        #     logger.error(f"血浆分离料工费结果录入失败！response：{wuliao_response.json()}")
                                        #     raise Exception(f"血浆分离料工费结果录入失败！response：{wuliao_response.json()}")
                                else:
                                    logger.error(f"血浆分离任务单结果录入失败！response：{response.json()}")
                                    raise Exception(f"血浆分离任务单结果录入失败！response：{response.json()}")
                            else:
                                logger.error(f"血浆分离产物查询失败！response：{response.json()}")
                                raise Exception(f"血浆分离产物查询失败！response：{response.json()}")
                        else:
                            logger.error(f"血浆分离任务开始失败！response：{response.json()}")
                            raise Exception(f"血浆分离任务开始失败！response：{response.json()}")
                else:
                    logger.error(f"血浆分离任务单查询失败！response：{confirm_response.json()}")
                    raise Exception(f"血浆分离任务单查询失败！response：{confirm_response.json()}")
            else:
                logger.error(f"血浆分离新建任务失败！response：{confirm_response.json()}")
                raise Exception(f"血浆分离新建任务失败！response：{confirm_response.json()}")
        else:
            logger.error(f"血浆分离查询数据失败，返回结果：{response}")
            raise Exception

        # zjob_code = confirm_response_data["zjob_code"]
        # self.zjob_code = zjob_code
        # return zjob_code

    def build_the_library(self, sample_ids=None, zkp_need=None, lane_index_info=None):
        """
        医学建库
        @return:
        """
        if sample_ids:
            self.sample = sample_ids
        if lane_index_info:
            self.lane_index_info = lane_index_info
        # 定义医学建库技术路线阶段编号和名称
        stage_code = "MAE"
        stage_name = "jkfdealorder"
        # 设置实验信息
        task_detail = {
            # "num": 0,
            # "id": "a-1",
            "type": "indexA",
            # "indexA": "1",
            "noDrop": False,
            "isEmptyNode": False,
            "ztx": self.config_info["indexConfig"],  # index配置
            # "z_fzh": -1,
            "numsSample": 0,
            "z_jk_qyl": 0,
            "z_jk_qytj": 0,
            "z_jk_te": 0,
            "z_jk_ybsyl": 0,
            "z_jk_ml": 0,
            "z_jk_cycle": 0,
            "zseqplatform": self.config_info["sequencePlatform"],
            # "zpoint": "a-1",
            # "_key": 1,
            "_id": 1,
            # "index": 1,
            "ddl21": self.config_info["sequenceType"],
            "draged": False,
            "radat": self.config_info["data"],
            "x": 12,
            "y": 8,
            # "count": 2,
            # "zkpCount": 4,
            "zsyyn": "刘秋芳",
            "zcxlx_des": "手工产线",
            "zmethod": "1755.01.23",
            "zmethod_des": "1755.01.23深圳-产前组-建库-MGISEQ-2000产前_全流程-自动化",
            "zshould_finish_datum": time.strftime("%Y%m%d"),
            "zplate_x": 12,
            "zplate_y": 8,
            # "zcwbh": "22L09090000-1",
            # "zindex": "1",
            # "zindex_name": "1",
            "zindextype": "indexA",
            "zshould_finish_uzeit": time.strftime("%H%M%S")
        }

        # 第一步：根据样例编号查询待建库的任务单
        query_unassign_task_data = {
            "task": {"zsample":  ",".join(str(i) for i in self.sample)},
            "zgxbh": "MAE",
            "token": self.token,
            "menuId": "MSTaskMaster_JK_JKFDealOrder"
        }
        logger.info("医学建库:开始查询待处理数据列表！")
        logger.info("医学建库:查询待处理数据:{}", query_unassign_task_data)
        query_unassign_task_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=task_assign_samplems", data=urlencode(query_unassign_task_data)).json()
        if query_unassign_task_respone["code"] == "200" and query_unassign_task_respone["total"] > 0:
            logger.info("医学建库：查询待处理数据列表成功！")
            query_unassign_task_list = query_unassign_task_respone["data"]
            query_unassign_task_list.sort(key=lambda x: x["pre_zjob_code_item"])
            # 第二步：医学建库-任务下达
            logger.info("医学建库:当前测试平台为:{}，index配置为:{}", task_detail["zseqplatform"],
                        task_detail["ztx"])
            # 先获取质控品信息
            if zkp_need == "是":
                zkp_info, zkp_list = self.get_zkp_info(task_detail["zseqplatform"], task_detail["ztx"])
            else:
                zkp_info, zkp_list = list(), list()
            # 获取excel表中的lane信息
            # sample_input_lane_list = [i[1] for i in self.sample_excel_list]
            sample_input_lane_list = [i["lane"] for i in self.lane_index_info]
            # 提交孔位号信息，获取孔板号
            pointCode_info = []
            for i in range(len(query_unassign_task_list)):
                pointCode_info_copy = dict()
                pointCode_info.append(pointCode_info_copy)
                pointCode_info[i]["zcatalo"] = query_unassign_task_list[i]["zcatalo"]
                pointCode_info[i]["zcwbh"] = query_unassign_task_list[i]["zcatalo"]
                if task_detail["zseqplatform"] == "MGISEQ-2000" and task_detail["ztx"] == "NIFTY1-48":
                    if all(lane in sample_input_lane_list for lane in ["L01", "L02", "L03", "L04"]):
                        if sample_input_lane_list[i] in ("L01", "L02"):  # 根据输入信息lane号设置孔板编号
                            pointCode_info[i]["zplate_bcode"] = "01"
                        elif sample_input_lane_list[i] in ("L03", "L04"):
                            pointCode_info[i]["zplate_bcode"] = "02"
                    else:
                        logger.error("jkfdealorder：输入lane号不正确！需包含L01、L02、L03、L04，请修改后重新导入文件！")
                        raise Exception("jkfdealorder：输入lane号不正确！需包含L01、L02、L03、L04，请修改后重新导入文件！")
                elif task_detail["zseqplatform"] == "MGISEQ-2000" and task_detail["ztx"] == "NIFTY1-96":
                    if all(lane in sample_input_lane_list for lane in ["L01", "L02", "L03", "L04"]):
                        if sample_input_lane_list[i] == "L01":  # 根据输入信息lane号设置孔板编号
                            pointCode_info[i]["zplate_bcode"] = "01"
                        elif sample_input_lane_list[i] == "L02":
                            pointCode_info[i]["zplate_bcode"] = "02"
                        elif sample_input_lane_list[i] == "L03":
                            pointCode_info[i]["zplate_bcode"] = "03"
                        elif sample_input_lane_list[i] == "L04":
                            pointCode_info[i]["zplate_bcode"] = "04"
                    else:
                        logger.error("jkfdealorder：输入lane号不正确！需包含L01、L02、L03、L04，请修改后重新导入文件！")
                        raise Exception("jkfdealorder：输入lane号不正确！需包含L01、L02、L03、L04，请修改后重新导入文件！")
                elif task_detail["zseqplatform"] in ("DNBSEQ-T7", "MGISEQ-200"):
                    if "L02" in sample_input_lane_list or "L03" in sample_input_lane_list or "L04" in sample_input_lane_list:
                        logger.error("jkfdealorder：输入lane号不正确！不能包含L02、L03、L04，请修改后重新导入文件！")
                        raise Exception("jkfdealorder：输入lane号不正确！不能包含L02、L03、L04，请修改后重新导入文件！")
                    else:
                        if sample_input_lane_list[i] == "L01":  # 根据输入信息lane号设置孔板编号
                            pointCode_info[i]["zplate_bcode"] = "01"
                k, v = tuple(self.lane_index_info[i].values())  # 临时增加拆包参数 p 牛龙飞 20240826
                pointCode_info[i]["id"] = self.get_point_num(task_detail["ztx"], k, v)
            pointCode_info.extend(zkp_info)
            save_pointcode_data = {"task": json.dumps(pointCode_info, ensure_ascii=False),"token": self.token}
            # print(save_pointcode_data)
            logger.info("jkfdealorder：开始保存建库样本孔位号信息！")
            save_pointcode_respone = self.nifty_res.post_request("/presap/webintf.do?method=analysis_sample_for_jk",
                                                  data=urlencode(save_pointcode_data)).json()
            if save_pointcode_respone["code"] == "200":
                logger.info("jkfdealorder：保存建库样本孔位号信息成功！")
                # 获取孔板条码后续填充入参
                platecode_info = save_pointcode_respone["data"][:-1]
                platecode = list(set([j["zplate"]["fieldValue"] for j in platecode_info]))
                platecode.sort()
                logger.debug("jkfdealorder：生成孔板号：{}", platecode)
            else:
                logger.error("jkfdealorder：保存建库样本孔位号信息失败！")
                logger.debug("save_pointcode_respone：{}", save_pointcode_respone)
                raise Exception("jkfdealorder：保存建库样本孔位号信息失败！")
            # 组装样本信息
            for i in range(len(query_unassign_task_list)):
                task_detail_copy = copy.deepcopy(task_detail)
                query_unassign_task_list[i].update(task_detail_copy)
                k, v = tuple(self.lane_index_info[i].values())  # 临时增加拆包参数 p 牛龙飞 20240826
                query_unassign_task_list[i]["id"] = self.get_point_num(task_detail["ztx"], k, v)
                query_unassign_task_list[i]["indexA"] = self.lane_index_info[i]["index"]
                # query_unassign_task_list[i]["z_fzh"] = i-i*2-1
                query_unassign_task_list[i]["zpoint"] = self.get_point_num(task_detail["ztx"], k, v)
                query_unassign_task_list[i]["_key"] = i + 1
                query_unassign_task_list[i]["index"] = len(query_unassign_task_list)
                # query_unassign_task_list[i]["count"] = len(query_unassign_task_list)
                query_unassign_task_list[i]["zcwbh"] = str(query_unassign_task_list[i]["pre_zcwbh"]).replace("P", "L")
                query_unassign_task_list[i]["zindex"] = self.lane_index_info[i]["index"]
                query_unassign_task_list[i]["zindex_name"] = self.lane_index_info[i]["index"]
                query_unassign_task_list[i]["zplate"] = platecode[0]
                # query_unassign_task_list[i]["zreceiveddate"] =
                if task_detail["zseqplatform"] == "MGISEQ-2000" and task_detail["ztx"] == "NIFTY1-48":
                    if sample_input_lane_list[i] in ("L03", "L04"):
                        query_unassign_task_list[i]["_id"] = 2
                        query_unassign_task_list[i]["zplate"] = platecode[1]
                elif task_detail["zseqplatform"] == "MGISEQ-2000" and task_detail["ztx"] == "NIFTY1-96":
                    if sample_input_lane_list[i] == "L02":
                        query_unassign_task_list[i]["_id"] = 2
                        query_unassign_task_list[i]["zplate"] = platecode[1]
                    elif sample_input_lane_list[i] == "L03":
                        query_unassign_task_list[i]["_id"] = 3
                        query_unassign_task_list[i]["zplate"] = platecode[2]
                    elif sample_input_lane_list[i] == "L04":
                        query_unassign_task_list[i]["_id"] = 4
                        query_unassign_task_list[i]["zplate"] = platecode[3]
            # 组装质控品信息
            for i in range(len(zkp_list)):
                # zkp_list[i]["zplate"] = platecode[0]
                zkp_list[i]["ztx"] = task_detail["ztx"]
                zkp_list[i]["zseqplatform"] = task_detail["zseqplatform"]
                zkp_list[i]["ddl21"] = task_detail["ddl21"]
                # zkp_list[i]["count"] = len(query_unassign_task_list)
                zkp_list[i]["zsyyn"] = task_detail["zsyyn"]
                zkp_list[i]["zmethod"] = task_detail["zmethod"]
                zkp_list[i]["zmethod_des"] = task_detail["zmethod_des"]
                zkp_list[i]["zshould_finish_datum"] = task_detail["zshould_finish_datum"]
                zkp_list[i]["zshould_finish_uzeit"] = task_detail["zshould_finish_uzeit"]
                zkp_list[i]["radat"] = task_detail["radat"]
                if zkp_list[i]["_id"] == 1:
                    zkp_list[i]["zplate"] = platecode[0]
                elif zkp_list[i]["_id"] == 2:
                    zkp_list[i]["zplate"] = platecode[1]
                elif zkp_list[i]["_id"] == 3:
                    zkp_list[i]["zplate"] = platecode[2]
                elif zkp_list[i]["_id"] == 4:
                    zkp_list[i]["zplate"] = platecode[3]
            query_unassign_task_list.extend(zkp_list)
            logger.info("医学建库：任务开始下达！")
            assign_task_data = {"task": json.dumps(query_unassign_task_list, ensure_ascii=False),
                                "token": self.token}
            assign_task_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=saveTaskAssignSampleMS", data=urlencode(assign_task_data)).json()
            if assign_task_respone["code"] == "200":
                if assign_task_respone["msg"] is None:
                    self.jkfdealorder_task_code = assign_task_respone["data"][0]["zjob_code"]
                else:
                    self.jkfdealorder_task_code = assign_task_respone["msg"]
                logger.debug("assign_task_respone is: {}", assign_task_respone)
                logger.info("医学建库：任务下达成功！")
                logger.info("医学建库任务单号：{}", self.jkfdealorder_task_code)
                # return self.jkfdealorder_task_code  # 返回任务号
                # 第三步：医学建库-任务生产
                # 开始任务
                start_task_data = {"zjob_code": self.jkfdealorder_task_code, "token": self.token}
                logger.info("医学建库：{}开始生产任务！", self.jkfdealorder_task_code)
                task_start_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=msStartTask",
                                                                 data=urlencode(start_task_data)).json()
                if task_start_respone["code"] == "200" and task_start_respone["msg"] == "success":
                    logger.info("医学建库：{}开始生产任务成功！", self.jkfdealorder_task_code)
                    logger.debug("task_start_respone：{}", task_start_respone)
                    # 录入结果
                    result_info = self.query_chanwu_msg("MAE", "jkfdealorder", self.jkfdealorder_task_code)
                    data = {}
                    for i in range(len(result_info)):
                        data[str(i + 1).zfill(6)] = {"tab1": [{}], "tab2": {}}
                        data[str(i + 1).zfill(6)]["tab1"][0]["zcwbh"] = result_info[i]["zcwbh"]
                        data[str(i + 1).zfill(6)]["tab1"][0]["zcwlx"] = "文库"
                        data[str(i + 1).zfill(6)]["tab1"][0]["zguid"] = result_info[i]["zguid"]
                        data[str(i + 1).zfill(6)]["tab1"][0]["zindex"] = result_info[i]["zindex"]
                        data[str(i + 1).zfill(6)]["tab1"][0]["ztj"] = "30"
                        data[str(i + 1).zfill(6)]["tab2"] = result_info[i]
                        data[str(i + 1).zfill(6)]["tab2"]["z_wk_lane"] = 1
                        data[str(i + 1).zfill(6)]["tab2"]["_id"] = i + 1
                        data[str(i + 1).zfill(6)]["tab2"]["zms_sfhg"] = "合格"
                    task_result_data = {"task": json.dumps([data], ensure_ascii=False),
                                        "zgxbh": stage_code,
                                        "zjob_code": self.jkfdealorder_task_code,
                                        "token": self.token
                                        }
                    logger.info("医学建库：{}开始任务单结果录入！", self.jkfdealorder_task_code)
                    task_result_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=msUpdateTaskResult", data=urlencode(task_result_data)).json()
                    if task_result_respone["code"] == "200" and task_result_respone["msg"] == "success":
                        logger.info("医学建库：{}任务单结果录入成功！", self.jkfdealorder_task_code)
                        logger.debug("task_result_respone：{}", task_result_respone)
                        # 查询任务样本信息
                        completion_info = self.query_sample_msg(stage_code, stage_name, self.jkfdealorder_task_code)
                        for i in range(len(completion_info)):
                            completion_info[i]["zlgfsflr"] = "X"
                            completion_info[i]["_id"] = i + 1
                        task_completion_data = {"req": json.dumps(completion_info, ensure_ascii=False),
                                                "token": self.token}
                        # 提交任务生产
                        logger.info("医学建库：{}开始提交生产任务！", self.jkfdealorder_task_code)
                        task_completion_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=task_sample_item_finishms",
                                                               data=urlencode(task_completion_data)).json()
                        if task_completion_respone["code"] == "200" and "SAP数据库更新成功" in task_completion_respone[
                            "msg"]:
                            logger.info("医学建库：{}提交生产任务成功！",self.jkfdealorder_task_code)
                            logger.debug("task_completion_respone：{}", task_completion_respone)
                            # 医学建库阶段需获取产物信息
                            # 建库任务完成时提取建库孔板号
                            plate_code = list(set([j["zplate_bcode"] for j in completion_info]))
                            plate_code.sort()
                            self.jkfdealorder_plate_code = plate_code
                            return self.jkfdealorder_plate_code
                        else:
                            logger.error("医学建库：{}提交生产任务失败！", stage_name, self.jkfdealorder_task_code)
                            logger.debug("task_completion_respone：{}", task_completion_respone)
                            raise Exception(self.jkfdealorder_task_code + "提交生产任务失败！")

                    else:
                        logger.error("医学建库：{}任务单结果录入失败！", self.jkfdealorder_task_code)
                        logger.debug("task_result_respone：{}", task_result_respone)
                        raise Exception(self.jkfdealorder_task_code + "任务单结果录入失败！")
                else:
                    logger.error("医学建库：{}开始生产任务失败！", self.jkfdealorder_task_code)
                    logger.debug("task_start_respone：{}", task_start_respone)
                    raise Exception(self.jkfdealorder_task_code + "开始生产任务失败！")

            else:
                logger.error("医学建库：任务下达失败！")
                logger.debug("assign_task_respone：{}", assign_task_respone)
                raise Exception("医学建库" + "任务下达失败！")
        else:
            logger.error("查询失败或数据不存在！")
            logger.debug("query_unassign_task_respone：{}", query_unassign_task_respone)
            raise Exception("医学建库," + "查询失败或数据不存在！")

    def bmg(self, sample_ids=None, zkp_need=None, lane_index_info=None):
        """
        BMG/Qubit
        @param sample_ids: 样例编号列表
        @param zkp_need: 是否有质控品：是，否
        @param lane_index_info: lane和index列表，示例：[{"lane":"L01","index":"25"},{"lane":"L02","index":"25"}]
        @return: self.bmg_task_code
        """
        if sample_ids:
            self.sample = sample_ids
        if lane_index_info:
            self.lane_index_info = lane_index_info
        # 定义医学建库技术路线阶段编号和名称
        stage_code = "MBE"
        stage_name = "bmgindex"
        # 设置实验信息
        task_detail = {
            "ztx": "196067",
            "zplate_x": 12,
            "zplate_y": 8,
            "noDrop": False,
            "isEmptyNode": False,
            "z_jk_qyl": 0,
            "z_jk_qytj": 0,
            "z_jk_te": 0,
            "z_jk_ybsyl": 0,
            "z_jk_ml": 0,
            "z_jk_cycle": 0,
            "z_jk_note": "",
            "x": 12,
            "y": 8,
            "zsyyn": "刘梓兴",
            "zmethod": "1755.04.09",
            "zmethod_des": "1755.04.09深圳_文库质控与测序组_BMG定量（自动化）_BGISEQ-500&1000",
            "zshould_finish_datum": time.strftime("%Y%m%d"),
            "zshould_finish_uzeit": time.strftime("%H%M%S")
        }
        # 第一步：根据样例编号查询待BMG的任务单
        query_unassign_task_data = {
            "task": {"zsample": ",".join(str(i) for i in self.sample)},
            "zgxbh": "MBE",
            "token": self.token,
            "menuId": "MSTaskMaster_BMG_index"
        }
        logger.info("医学BMG:开始查询待处理数据列表！")
        logger.info("医学BMG:查询待处理数据:{}", query_unassign_task_data)
        query_unassign_task_respone = self.nifty_res.post_request("/presap/webintf.do?method=task_assign_samplems",
                                                                  data=urlencode(query_unassign_task_data)).json()
        # print(query_unassign_task_respone)
        if query_unassign_task_respone["code"] == "200" and query_unassign_task_respone["total"] > 0:
            logger.info("医学BMG：查询待处理数据列表成功！")
            query_unassign_task_list = query_unassign_task_respone["data"]
            # 第二步：任务下达
            for i in range(len(query_unassign_task_list)):
                task_detail_copy = copy.deepcopy(task_detail)
                query_unassign_task_list[i].update(task_detail_copy)
                query_unassign_task_list[i]["_key"] = i + 1
                query_unassign_task_list[i]["zindex"] = query_unassign_task_list[i]["pre_zindex"]
                query_unassign_task_list[i]["zindex_name"] = query_unassign_task_list[i]["pre_zindex_name"]
                query_unassign_task_list[i]["zindextype"] = query_unassign_task_list[i]["pre_zindextype"]
                query_unassign_task_list[i]["id"] = query_unassign_task_list[i]["zpoint"]
                query_unassign_task_list[i]["type"] = query_unassign_task_list[i]["pre_zindextype"]
                query_unassign_task_list[i]["indexA"] = query_unassign_task_list[i]["pre_zindex"]
                query_unassign_task_list[i]["indexB"] = query_unassign_task_list[i]["pre_zindex"]
                query_unassign_task_list[i]["zcwbh"] = query_unassign_task_list[i]["pre_zcwbh"]
                query_unassign_task_list[i]["radat"] = query_unassign_task_list[i]["pre_radat"]
                query_unassign_task_list[i]["zplate"] = query_unassign_task_list[i]["zplate_bcode"].split("-")[1]
                query_unassign_task_list[i]["pre_zcwlx"] = "血浆"
                query_unassign_task_list[i]["_id"] = int(query_unassign_task_list[i]["zplate_bcode"].split("-")[1])
                if zkp_need == "是":
                    if "平台阴阳质控" in query_unassign_task_list[i]["pre_maktx"]:
                        query_unassign_task_list[i]["zindextype"] = "controller"
                        query_unassign_task_list[i]["type"] = "controller"
                        query_unassign_task_list[i]["zzkxh"] = "X"
                        query_unassign_task_list[i]["zreceiveddate"] = query_unassign_task_list[i]["pre_zreceiveddate"]
                        query_unassign_task_list[i]["zsampling_datum"] = query_unassign_task_list[i][
                            "pre_zsampling_datum"]
                        query_unassign_task_list[i]["pre_zcwlx"] = "文库"
                    if query_unassign_task_list[i]["_key"] == 1:
                        query_unassign_task_list[i]["maktx"] = query_unassign_task_list[i][
                                                                   "pre_maktx"] + ",NIFTY500平台阴阳质控"  # ,NIFTY500平台空白对照"
                        query_unassign_task_list[i]["matnr"] = query_unassign_task_list[i][
                                                                   "pre_matnr"] + ",RC0036"  # ,RC0037"
            if zkp_need == "是":
                # sample_input_lane_list = [i[1] for i in self.sample_excel_list]
                sample_input_lane_list = [i["lane"] for i in self.lane_index_info]
                if "L03" or "L04" in sample_input_lane_list:
                    list_index = sample_input_lane_list.index("L03")
                    query_unassign_task_list[list_index]["_key"] = 2
                    query_unassign_task_list[list_index]["maktx"] = query_unassign_task_list[list_index][
                                                                        "pre_maktx"] + ",NIFTY500平台阴阳质控"  # ,NIFTY500平台空白对照"
                    query_unassign_task_list[list_index]["matnr"] = query_unassign_task_list[list_index][
                                                                        "pre_matnr"] + ",RC0036"  # ,RC0037"
            # 开始下达任务
            logger.info("BMG：任务开始下达！")
            assign_task_data = {"task": json.dumps(query_unassign_task_list, ensure_ascii=False),
                                "token": self.token}
            # print(assign_task_data)
            assign_task_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=saveTaskAssignSampleMS", data=urlencode(assign_task_data)).json()
            if assign_task_respone["code"] == "200":
                if assign_task_respone["msg"] is None:
                    self.bmg_task_code = assign_task_respone["data"][0]["zjob_code"]
                else:
                    self.bmg_task_code = assign_task_respone["msg"]
                logger.debug("assign_task_respone is: {}", assign_task_respone)
                logger.info("BMG：任务下达成功！")
                logger.info("BMG任务单号：{}", self.bmg_task_code)

                # 第三步：任务生产
                # 准备任务单结果录入数据
                task_info = {
                    "zbmgresult": "合格", #BMG结果是否合格
                    "bvalue": "6056.4", # B值
                    "kvalue": "196067", # K值
                    "zzlnd": 4.05, # 文库浓度
                    "sbxs": "0.95", # 设备系数
                    "xsbs": "60", # 稀释倍数
                    "zxgz": "20000" #吸光值
                }
                # 开始BMG生产任务
                task_start_data = {
                    "zjob_code": self.bmg_task_code,
                    "token": self.token}
                logger.info("医学BMG：{}开始生产任务！", self.bmg_task_code)
                task_start_respone = self.nifty_res.post_request("/presap/webintf.do?method=msStartTask", data=urlencode(task_start_data)).json()
                if task_start_respone["code"] == "200" and task_start_respone["msg"] == "success":
                    logger.info("医学BMG：{}开始生产任务成功！", self.bmg_task_code)
                    logger.debug("task_start_respone：{}", task_start_respone)
                    # 查询任务产物信息
                    result_info = self.query_chanwu_msg(stage_code, stage_name, self.bmg_task_code)
                    # 任务单结果录入
                    result_info.sort(key=lambda x: x["zjob_code_item"])
                    data = {}
                    for i in range(len(result_info)):
                        data[str(i + 1).zfill(6)] = {"tab1": [{}], "tab2": {}}
                        data[str(i + 1).zfill(6)]["tab1"][0]["zcwbh"] = result_info[i]["pre_zcwbh"]
                        data[str(i + 1).zfill(6)]["tab2"] = result_info[i]
                        task_info_copy = copy.deepcopy(task_info)
                        data[str(i + 1).zfill(6)]["tab2"].update(task_info_copy)
                        if "空白" in data[str(i + 1).zfill(6)]["tab2"]["maktx"]:
                            data[str(i + 1).zfill(6)]["tab2"]["zzlnd"] = 0.27
                            data[str(i + 1).zfill(6)]["tab2"]["zxgz"] = 7000
                    task_result_data = {"task": json.dumps([data], ensure_ascii=False),
                                        "zgxbh": stage_code,
                                        "zjob_code": self.bmg_task_code,
                                        "token": self.token
                                        }
                    # print(task_result_data)
                    logger.info("医学BMG：{}开始任务单结果录入！", self.bmg_task_code)
                    task_result_respone = self.nifty_res.post_request("/presap/webintf.do?method=msUpdateTaskResult", data=urlencode(task_result_data)).json()
                    if task_result_respone["code"] == "200" and task_result_respone["msg"] == "success":
                        logger.info("医学BMG：{}任务单结果录入成功！", self.bmg_task_code)
                        logger.debug("task_result_respone：{}", task_result_respone)
                    else:
                        logger.error("医学BMG：{}任务单结果录入失败！", self.bmg_task_code)
                        logger.debug("task_result_respone：{}", task_result_respone)
                        raise Exception(self.bmg_task_code + "任务单结果录入失败！")

                    # 查询任务样本信息
                    completion_info = self.query_sample_msg(stage_code, stage_name, self.bmg_task_code)
                    for i in range(len(completion_info)):
                        completion_info[i]["zlgfsflr"] = "X"
                        completion_info[i]["_id"] = i + 1
                    task_completion_data = {"req": json.dumps(completion_info, ensure_ascii=False),"token": self.token}

                    # 完成BMG生产任务
                    logger.info("医学BMG：{}开始任务单完成任务！", self.bmg_task_code)
                    task_completion_respone = self.nifty_res.post_request("/presap/webintf.do?method=task_sample_item_finishms",
                                                           data=urlencode(task_completion_data)).json()
                    if task_completion_respone["code"] == "200" and "SAP数据库更新成功" in task_completion_respone[
                        "msg"]:
                        logger.info("医学BMG：{}任务单完成任务成功！", self.bmg_task_code)
                        logger.debug("task_completion_respone：{}", task_completion_respone)
                        return self.bmg_task_code
                    else:
                        logger.error("医学BMG：{}任务单完成任务失败！", self.bmg_task_code)
                        logger.debug("task_completion_respone：{}", task_completion_respone)
                        raise Exception(self.bmg_task_code + "任务单完成任务失败！")
                else:
                    logger.error("医学BMG：{}开始生产任务失败！", self.bmg_task_code)
                    logger.debug("task_start_respone：{}", task_start_respone)
                    raise Exception(self.bmg_task_code + "开始任务失败！")
            else:
                logger.error("BMG：任务下达失败！")
                logger.debug("assign_task_respone：{}", assign_task_respone)
                raise Exception(stage_name + "任务下达失败！")
        else:
            logger.error("查询失败或数据不存在！")
            logger.debug("query_unassign_task_respone：{}", query_unassign_task_respone)
            raise Exception(stage_name + "查询失败或数据不存在！")

    def query_chanwu_msg(self, stage_code, stage_name, task_code):
        """
        查询任务产物信息
         @param task_code: 阶段编号
        @param stage_code: 任务号
        @param stage_name: 阶段名称
        @return: 产物
        """
        # 根据stage_code获取stage_name
        # 查询产物信息
        query_chanwu_data = {"task": json.dumps({"zjob_code": task_code, "zgxbh": stage_code, "zsfzjcw": "ALL"}, ensure_ascii=False),
                             "token": self.token}
        logger.info("{}：{}开始查询任务产物信息！", stage_name, task_code)
        query_chanwu_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=find_chanWuMS", data=urlencode(query_chanwu_data)).json()
        if query_chanwu_respone["code"] == "200" and query_chanwu_respone["total"] > 0:
            logger.info("{}：{}查询任务产物信息成功！", stage_name, task_code)
            return query_chanwu_respone["data"]["Samples"]
        else:
            logger.error("{}：{}查询任务产物信息失败！", stage_name, task_code)
            logger.debug("query_chanwu_respone：{}", query_chanwu_respone)
            raise Exception(task_code + "查询任务产物信息失败！")

    @staticmethod
    def get_point_num(index_type, lane, index):
        """
        根据Excel表格的lane和index生成孔位号
        :param index_type: 传入index配置类型
        :param lane: 传入lane号
        :param index: 传入index号
        :return:
        """
        # 根据96孔板设置孔位号
        list_point_1 = ["a-1", "b-1", "c-1", "d-1", "e-1", "f-1", "g-1", "h-1", "a-2", "b-2", "c-2", "d-2", "e-2",
                        "f-2",
                        "g-2", "h-2", "a-3", "b-3", "c-3", "d-3", "e-3", "f-3", "g-3", "h-3", "a-4", "b-4", "c-4",
                        "d-4",
                        "e-4", "f-4", "g-4", "h-4", "a-5", "b-5", "c-5", "d-5", "e-5", "f-5", "g-5", "h-5", "a-6",
                        "b-6",
                        "c-6", "d-6", "e-6", "f-6", "g-6", "h-6"]
        list_point_2 = ["a-7", "b-7", "c-7", "d-7", "e-7", "f-7", "g-7", "h-7", "a-8", "b-8", "c-8", "d-8", "e-8",
                        "f-8",
                        "g-8", "h-8", "a-9", "b-9", "c-9", "d-9", "e-9", "f-9", "g-9", "h-9", "a-10", "b-10", "c-10",
                        "d-10", "e-10", "f-10", "g-10", "h-10", "a-11", "b-11", "c-11", "d-11", "e-11", "f-11", "g-11",
                        "h-11", "a-12", "b-12", "c-12", "d-12", "e-12", "f-12", "g-12", "h-12"]
        list_point_3 = list_point_1 + list_point_2
        if index_type == "NIFTY1-48":
            if 0 < int(index) < 49:
                if lane in ("L01", "L03"):
                    return list_point_1[int(index) - 1]
                else:
                    return list_point_2[int(index) - 1]
            else:
                raise IndexError(f"输入index号超出index配置范围！index配置：{index_type}，index号：{index}")
        elif index_type == "NIFTY1-96":
            if 0 < int(index) < 97:
                return list_point_3[int(index) - 1]
            else:
                raise IndexError(f"输入index号超出index配置范围！index配置：{index_type}，index号：{index}")

    def get_zkp_info(self, seq_platform, index_type):
        """
        获取质控品信息
        :param seq_platform: 测序机型
        :param index_type: index配置
        :return:
        """
        create_zkp_type = ["KBN", "KBP"]  # 阴性，阳性，固定，随机 -- ["KBN", "KBP", "KBF", "KBR"]
        create_zkp_list = []
        logger.info("jkfdealorder：开始获取质控品编号！")
        for i in range(len(create_zkp_type)):
            create_zkp_data = {"datas": json.dumps([{"zkplx": create_zkp_type[i], "zkbxh": 1}], ensure_ascii=False),
                               "token": self.token}
            create_zkp_list.append(self.nifty_res.post_request(url="/presap/webintf.do?method=create_zkp_code", data=create_zkp_data).json()["data"][0]["ZKPNUM"]["fieldValue"])
        logger.info("jkfdealorder：质控品编号获取成功！")
        logger.debug("jkfdealorder：质控品编号：{}", create_zkp_list)
        # 组装质控品孔位号提交信息和质控品信息
        zkp_info = []
        zkp_list = []
        # 质控品类型
        zkp_info_yax = {
            "zcatalo": "9999999996",
            # "zcwbh":"ZKP0009982YAX-10101",
            # "zplate_bcode": "01",
            "id": "e-2"
        }
        zkp_info_yix = {
            "zcatalo": "9999999997",
            # "zcwbh":"ZKP0009981YIX-10101",
            # "zplate_bcode": "01",
            "id": "e-8"
        }
        if seq_platform == "MGISEQ-2000" and index_type == "NIFTY1-48":
            for i in range(2):
                zkp_info_yax = copy.deepcopy(nifty_zkp_template.zkp_info_yax)
                zkp_info_yix = copy.deepcopy(nifty_zkp_template.zkp_info_yix)
                zkp_info_yax["zcwbh"] = create_zkp_list[1]+"-10"+str(i+1)+"0"+str(i+1)
                zkp_info_yix["zcwbh"] = create_zkp_list[0]+"-10"+str(i+1)+"0"+str(i+1)
                zkp_info_yax["zplate_bcode"] = "0"+str(i+1)
                zkp_info_yix["zplate_bcode"] = "0"+str(i+1)
                zkp_info.append(zkp_info_yax)
                zkp_info.append(zkp_info_yix)
                zkp_list_yax = copy.deepcopy(nifty_zkp_template.zkp_template_yax_13)
                zkp_list_yix = copy.deepcopy(nifty_zkp_template.zkp_template_yix_13)
                zkp_list_yax["zcatalo"] = str(create_zkp_list[1])+"-10"+str(i+1)  # 阳性质控品
                zkp_list_yax["pre_zcwbh"] = str(create_zkp_list[1])+"-10"+str(i+1)
                zkp_list_yax["zzkpbm_inner"] = "04$"+str(create_zkp_list[1])+"-10"+str(i+1)
                zkp_list_yax["zcwbh"] = (str(create_zkp_list[1])+"-10"+str(i+1)).replace("P", "L")
                zkp_list_yax["_id"] = i+1
                zkp_list_yix["zcatalo"] = str(create_zkp_list[0]) + "-10" + str(i + 1)  # 阴性质控品
                zkp_list_yix["pre_zcwbh"] = str(create_zkp_list[0]) + "-10" + str(i + 1)
                zkp_list_yix["zzkpbm_inner"] = "03$" + str(create_zkp_list[0]) + "-10" + str(i + 1)
                zkp_list_yix["zcwbh"] = (str(create_zkp_list[0]) + "-10" + str(i + 1)).replace("P", "L")
                zkp_list_yix["_id"] = i+1
                zkp_list.append(zkp_list_yax)
                zkp_list.append(zkp_list_yix)
            return zkp_info, zkp_list
        elif seq_platform == "MGISEQ-2000" and index_type == "NIFTY1-96":
            for i in range(4):
                zkp_info_yax = copy.deepcopy(nifty_zkp_template.zkp_info_yax)
                zkp_info_yix = copy.deepcopy(nifty_zkp_template.zkp_info_yix)
                zkp_info_yax["zcwbh"] = create_zkp_list[1] + "-10" + str(i + 1) + "0" + str(i + 1)
                zkp_info_yix["zcwbh"] = create_zkp_list[0] + "-10" + str(i + 1) + "0" + str(i + 1)
                zkp_info_yax["zplate_bcode"] = "0" + str(i + 1)
                zkp_info_yix["zplate_bcode"] = "0" + str(i + 1)
                zkp_info.append(zkp_info_yax)
                zkp_info.append(zkp_info_yix)
                zkp_list_yax = copy.deepcopy(nifty_zkp_template.zkp_template_yax_13)
                zkp_list_yix = copy.deepcopy(nifty_zkp_template.zkp_template_yix_61)
                zkp_list_yax["zcatalo"] = str(create_zkp_list[1]) + "-10" + str(i + 1)  # 阳性质控品
                zkp_list_yax["pre_zcwbh"] = str(create_zkp_list[1]) + "-10" + str(i + 1)
                zkp_list_yax["zzkpbm_inner"] = "04$" + str(create_zkp_list[1]) + "-10" + str(i + 1)
                zkp_list_yax["zcwbh"] = (str(create_zkp_list[1]) + "-10" + str(i + 1)).replace("P", "L")
                zkp_list_yax["_id"] = i + 1
                zkp_list_yix["zcatalo"] = str(create_zkp_list[0]) + "-10" + str(i + 1)  # 阴性质控品
                zkp_list_yix["pre_zcwbh"] = str(create_zkp_list[0]) + "-10" + str(i + 1)
                zkp_list_yix["zzkpbm_inner"] = "03$" + str(create_zkp_list[0]) + "-10" + str(i + 1)
                zkp_list_yix["zcwbh"] = (str(create_zkp_list[0]) + "-10" + str(i + 1)).replace("P", "L")
                zkp_list_yix["_id"] = i + 1
                zkp_list.append(zkp_list_yax)
                zkp_list.append(zkp_list_yix)
            return zkp_info, zkp_list
        elif seq_platform in ("DNBSEQ-T7", "MGISEQ-200") and index_type == "NIFTY1-48":
            zkp_info_yax = copy.deepcopy(nifty_zkp_template.zkp_info_yax)
            zkp_info_yax["zcwbh"] = create_zkp_list[1] + "-10101"
            zkp_info_yax["zplate_bcode"] = "01"
            zkp_info.append(zkp_info_yax)
            zkp_list_yax = copy.deepcopy(nifty_zkp_template.zkp_template_yax_13)
            zkp_list_yax["zcatalo"] = str(create_zkp_list[1]) + "-101"  # 阳性质控品
            zkp_list_yax["pre_zcwbh"] = str(create_zkp_list[1]) + "-101"
            zkp_list_yax["zzkpbm_inner"] = "04$" + str(create_zkp_list[1]) + "-101"
            zkp_list_yax["zcwbh"] = (str(create_zkp_list[1]) + "-101").replace("P", "L")
            zkp_list_yax["_id"] = 1
            zkp_list.append(zkp_list_yax)
            return zkp_info, zkp_list
        elif seq_platform in ("DNBSEQ-T7", "MGISEQ-200") and index_type == "NIFTY1-96":
            zkp_info_yax = copy.deepcopy(nifty_zkp_template.zkp_info_yax)
            zkp_info_yix = copy.deepcopy(nifty_zkp_template.zkp_info_yix)
            zkp_info_yax["zcwbh"] = create_zkp_list[1] + "-10101"
            zkp_info_yix["zcwbh"] = create_zkp_list[0] + "-10101"
            zkp_info_yax["zplate_bcode"] = "01"
            zkp_info_yix["zplate_bcode"] = "01"
            zkp_info.append(zkp_info_yax)
            zkp_info.append(zkp_info_yix)
            zkp_list_yax = copy.deepcopy(nifty_zkp_template.zkp_template_yax_13)
            zkp_list_yix = copy.deepcopy(nifty_zkp_template.zkp_template_yix_61)
            zkp_list_yax["zcatalo"] = str(create_zkp_list[1]) + "-101"  # 阳性质控品
            zkp_list_yax["pre_zcwbh"] = str(create_zkp_list[1]) + "-101"
            zkp_list_yax["zzkpbm_inner"] = "04$" + str(create_zkp_list[1]) + "-101"
            zkp_list_yax["zcwbh"] = (str(create_zkp_list[1]) + "-101").replace("P", "L")
            zkp_list_yax["_id"] = 1
            zkp_list_yix["zcatalo"] = str(create_zkp_list[0]) + "-101"  # 阴性质控品
            zkp_list_yix["pre_zcwbh"] = str(create_zkp_list[0]) + "-101"
            zkp_list_yix["zzkpbm_inner"] = "03$" + str(create_zkp_list[0]) + "-101"
            zkp_list_yix["zcwbh"] = (str(create_zkp_list[0]) + "-101").replace("P", "L")
            zkp_list_yix["_id"] = 1
            zkp_list.append(zkp_list_yax)
            zkp_list.append(zkp_list_yix)
            return zkp_info, zkp_list

    def query_sample_msg(self, stage_code, stage_name, task_code):
        """
                任务生产-查询样本信息
                :param stage_code: 阶段编号
                :param stage_name: 阶段名称
                :param task_code: 任务号
                :return:
                """
        # 查询任务信息
        query_task_data = {"task": json.dumps({"zjob_code": task_code}, ensure_ascii=False),
                           "zgxbh": stage_code,
                           "token": self.token}
        # print(query_task_data)
        logger.info("{}：{}开始查询任务信息！", stage_name, task_code)
        query_task_respone = self.nifty_res.post_request("/presap/webintf.do?method=findJobMS", data=query_task_data).json()
        if query_task_respone["code"] == "200" and query_task_respone["total"] >= 1:
            logger.info("{}：{}查询任务信息成功！", stage_name, task_code)
        else:
            logger.error("{}：{}查询任务信息失败！", stage_name, task_code)
            logger.debug("query_task_respone：{}", query_task_respone)
            raise Exception(task_code + "查询任务信息失败！")
        query_task_list = query_task_respone["data"]
        # 查询任务内样本信息
        query_sample_value = query_task_list[0]
        query_sample_value["zjob_status"] = "已开始"
        query_sample_value["_key"] = 1
        query_sample_value["_id"] = 1
        query_sample_data = {"task": json.dumps(query_sample_value, ensure_ascii=False),
                             "token": self.token}
        # print(query_sample_data)
        logger.info("{}：{}开始查询任务样本信息！", stage_name, task_code)
        query_sample_respone = self.nifty_res.post_request(url="/presap/webintf.do?method=findJobSampleMessageMS", data=query_sample_data).json()
        if query_sample_respone["code"] == "200" and query_sample_respone["total"] > 0:
            logger.info("{}：{}查询任务样本信息成功！", stage_name, task_code)
            return query_sample_respone["data"]
        else:
            logger.error("{}：{}查询任务样本信息失败！", stage_name, task_code)
            logger.debug("query_sample_respone：{}", query_sample_respone)
            raise Exception(task_code + "查询任务样本信息失败！")

    def pooling(self, sample=None):
        """
        造数工具pooling
        :return:
        """
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        data = {
            "task": {
                "pre_zplate_bcode": "",
                "pre_zcwbh": "",
                "zsample": sampleid
            },
            "pageNumber": "1",
            "pageSize": "1000",
            "zgxbh": "MAX",
            "token": self.token,
            "menuId": "MSTaskMaster_Pooling_index"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=task_assign_samplems", data=urlencode(data))
        if response.json()["code"] == "200" :
            logger.info("pooling新建任务列表查询成功！")
            list_data = response.json()["data"]
            list_data.sort(key=lambda x: x["pre_zjob_code_item"])
            detail = {
                "zplate_x": 12,
                "zplate_y": 8,
                "type": "indexA",
                "noDrop": False,
                "isEmptyNode": False,
                "ztx": self.config_info["indexConfig"],
                "z_jk_qyl": 0,
                "z_jk_qytj": 0,
                "z_jk_te": 0,
                "z_jk_ybsyl": 0,
                "z_jk_ml": 0,
                "z_jk_cycle": 0,
                "zax_vol_spl": "15.80",
                "theory_con": "4.11",
                "water_volume": "14.19",
                "pooling_sum": "640",
                "x": 12,
                "y": 8,
                "zsyyn": "何华",
                "zmethod": "1755.04.38",
                "zmethod_des": "1755.04.38深圳-医学测序组-POOLING-MGISEQ-2000_POOLING&BMG定量(TECAN)",
                "zshould_finish_datum": self.date_str,
                "zindextype": "indexA",
                "zshould_finish_uzeit": self.time_str
            }
            datas = []
            for i in range(len(list_data)):
                data_dic = {k: "" if v is None else v for k, v in list_data[i].items()}
                data_dic.update(detail)
                data_dic["zplate"] = data_dic["zplate_bcode"].split("-")[1]
                data_dic["pre_zcwlx"] = "血浆"
                data_dic["_key"] = i + 1
                data_dic["zindex"] = data_dic["pre_zindex"]
                data_dic["radat"] = data_dic["pre_radat"]
                data_dic["_id"] = int(data_dic["zplate_bcode"].split("-")[1])
                data_dic["id"] = data_dic["zpoint"]
                data_dic["indexA"] = data_dic["pre_zindex"]
                data_dic["indexB"] = data_dic["pre_zindex"]
                data_dic["zcwbh"] = data_dic["pre_zcwbh"]
                data_dic["zindex_name"] = data_dic["pre_zindex_name"]
                data_dic["noDrop"] = ""
                data_dic["isEmptyNode"] = ""
                if data_dic["zseqplatform"] == "MGISEQ-2000" and data_dic[
                    "ztx"] == "NIFTY1-48":
                    if int(data_dic["zpoint"].split("-")[1]) <= 6:
                        data_dic["planNo"] = 1
                        data_dic["zscheme"] = str(
                            data_dic["zplate_bcode"].split("-")[1]) + "01"
                    else:
                        data_dic["planNo"] = 2
                        data_dic["zscheme"] = str(
                            data_dic["zplate_bcode"].split("-")[1]) + "02"
                elif data_dic["zseqplatform"] == "MGISEQ-2000" and data_dic[
                    "ztx"] == "NIFTY1-96":
                    data_dic["planNo"] = 1
                    data_dic["zscheme"] = str(
                        data_dic["zplate_bcode"].split("-")[1]) + "01"
                elif data_dic["zseqplatform"] in ("DNBSEQ-T7", "MGISEQ-200"):
                    data_dic["planNo"] = 1
                    data_dic["zscheme"] = str(
                        data_dic["zplate_bcode"].split("-")[1]) + "01"
                datas.append(data_dic)
            data = {
                "task": datas,
                "token": self.token,
                "menuId": "MSTaskMaster_Pooling_index"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=saveTaskAssignSampleMS",
                                                       data=urlencode(data))
            if response.json()["code"] =="200":
                logger.info("pooling下达任务成功！")
                self.zjob_code = response.json()["msg"]
                data = {
                    "task": {
                        "zsfxd_datum": "",
                        "zsfxd_datumend": "",
                        "zsfks_datum": "",
                        "zsfks_datumend": "",
                        "zjob_code": self.zjob_code,
                        "zsfwc": "",
                        "task": "query"
                    },
                    "pageNumber": "1",
                    "zgxbh": "MAX",
                    "pageSize": "1000",
                    "token": self.token,
                    "menuId": "MSTaskProduct_Pooling"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=findJobMS",data=urlencode(data))
                if response.json()["code"] =="200":
                    logger.info("pooling任务单列表查询成功！")
                    rwd_data = response.json()["data"][0]
                    data = {
                        "task": {
                            "zpzxlh_bcode": "",
                            "zcjdat": "",
                            "ybbhrq": "",
                            "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                            "zsfwc_czr": "",
                            "zsfks_czr": "",
                            "zcdate": "",
                            "chkdate": "",
                            "matnr": rwd_data["matnr"],
                            "zseqplatform": rwd_data["zseqplatform"],
                            "zsfks_uzeit": "",
                            "zdlhwkh": "",
                            "qdfkrq": "",
                            "zsampling_datum": "",
                            "zdlvdate": """ """,
                            "zgxbh": "MAX",
                            "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                            "zsigndate": "",
                            "zsfks": "",
                            "zsyyn": rwd_data["zsyyn"],
                            "syncdate": "",
                            "ddl21": rwd_data["ddl21"],
                            "yctjrq": "",
                            "zcxlx": "0",
                            "zjob_begin_datum": "",
                            "zsfwc_uzeit": "",
                            "zsample_quantity": rwd_data["zsample_quantity"],
                            "zmethod": rwd_data["zmethod"],
                            "zsfxd_czr": rwd_data["zsfxd_czr"],
                            "zudate": "",
                            "addate": "",
                            "zycdjsflr": "否",
                            "zjob_status": "未开始",
                            "zsfxd_datum": rwd_data["zsfxd_datum"],
                            "werks": "A020",
                            "zsfks_datum": "",
                            "ycwcrq": "",
                            "pre_zsfwc_datum": "",
                            "zsfwc": "",
                            "pre_zjob_code": rwd_data["zjob_code"],
                            "maktx": rwd_data["maktx"],
                            "zsfwc_datum": "",
                            "zupdat": "",
                            "zjob_code": rwd_data["zjob_code"],
                            "gt_action": "",
                            "zcbdate": "",
                            "gt_sendtime": "",
                            "zfinish_time": "",
                            "_key": 1,
                            "_id": 1
                        },
                        "token": self.token,
                        "menuId": "MSTaskProduct_Pooling"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS",data=urlencode(data))
                    if response.json()["code"] =="200":
                        logger.info("pooling任务单详情查询成功！")
                        rwd_detail_data = response.json()["data"]
                        data = {
                            "zgxbh": "MAX",
                            "zjob_code": self.zjob_code,
                            "token": self.token,
                            "menuId": "MSTaskProduct_Pooling"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=msStartTask",data=urlencode(data))
                        if response.json()["code"] =="200" and response.json()["msg"]=="success":
                            logger.info("pooling任务开始成功！")
                            data = {
                                "task": {
                                    "zpzxlh_bcode": "",
                                    "zcjdat": "",
                                    "ybbhrq": "",
                                    "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                    "zsfwc_czr": "",
                                    "zsfks_czr": "",
                                    "zcdate": "",
                                    "chkdate": "",
                                    "matnr": rwd_data["matnr"],
                                    "zseqplatform": rwd_data["zseqplatform"],
                                    "zsfks_uzeit": "",
                                    "zdlhwkh": "",
                                    "qdfkrq": "",
                                    "zsampling_datum": "",
                                    "zdlvdate": """ """,
                                    "zgxbh": "MAX",
                                    "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                    "zsigndate": "",
                                    "zsfks": "",
                                    "zsyyn": rwd_data["zsyyn"],
                                    "syncdate": "",
                                    "ddl21": rwd_data["ddl21"],
                                    "yctjrq": "",
                                    "zcxlx": "0",
                                    "zjob_begin_datum": "",
                                    "zsfwc_uzeit": "",
                                    "zsample_quantity": rwd_data["zsample_quantity"],
                                    "zmethod": rwd_data["zmethod"],
                                    "zsfxd_czr": rwd_data["zsfxd_czr"],
                                    "zudate": "",
                                    "addate": "",
                                    "zycdjsflr": "否",
                                    "zjob_status": "未开始",
                                    "zsfxd_datum": rwd_data["zsfxd_datum"],
                                    "werks": "A020",
                                    "zsfks_datum": "",
                                    "ycwcrq": "",
                                    "pre_zsfwc_datum": "",
                                    "zsfwc": "",
                                    "pre_zjob_code": rwd_data["zjob_code"],
                                    "maktx": rwd_data["maktx"],
                                    "zsfwc_datum": "",
                                    "zupdat": "",
                                    "zjob_code": rwd_data["zjob_code"],
                                    "gt_action": "",
                                    "zcbdate": "",
                                    "gt_sendtime": "",
                                    "zfinish_time": "",
                                    "_key": 1,
                                    "_id": 1
                                },
                                "token": self.token,
                                "menuId": "MSTaskProduct_Pooling"
                            }
                            response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS",
                                                                   data=urlencode(data))
                            rwd_detail_data = response.json()["data"]
                            if self.abnormal_input == '1':
                                # 录入异常
                                data = {
                                    "zgxbh": "MAX",
                                    "werks": "A020",
                                    "arbpl": "",
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_Pooling"
                                }
                                response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",data=urlencode(data))
                                if response.json()["code"] =="200" and len(response.json()["data"])>0:
                                    logger.info("pooling异常查询成功！")
                                    exception_data = response.json()["data"][0]
                                    datas =[]
                                    for i in range(len(rwd_detail_data)):
                                        data_dic = {k: "" if v is None else v for k, v in rwd_detail_data[i].items()}
                                        exception_data_dic = {k: "" if v is None else v for k, v in exception_data.items()}
                                        data_dic.update(exception_data_dic)
                                        data_dic["_id"] = i + 1
                                        data_dic["zcjnam"] = "auto_test"
                                        data_dic["zcjdat"] = self.date_str
                                        data_dic["zcjuzt"] = self.time_str
                                        data_dic["dateTime"] = self.datatime_str
                                        data_dic["yc_status"] = "A"
                                        datas.append(data_dic)
                                    data = {
                                        "task": datas,
                                        "werks": "A020",
                                        "arbpl": "",
                                        "zjob_code": rwd_data["zjob_code"],
                                        "zgxbh": "MAX",
                                        "token": self.token,
                                        "menuId": "MSTaskProduct_Pooling"
                                    }
                                    response = self.nifty_res.post_request("/presap/webintf.do?method=save_exception_ms", data=urlencode(data))
                                    if response.json()["code"] == "200" and response.json()["msg"] == "数据保存成功.":
                                        logger.info("pooling登记异常录入成功！")
                                    else:
                                        logger.error(f"pooling登记异常录入失败，返回结果：{response}")
                                        raise Exception(f"pooling登记异常录入失败，返回结果：{response}")
                                else:
                                    logger.error(f"pooling异常查询失败，返回结果：{response}")
                                    raise Exception(f"pooling异常查询失败，返回结果：{response}")
                            else:
                                data = {
                                    "task": {
                                        "zjob_code": rwd_data["zjob_code"],
                                        "zgxbh": "MAX",
                                        "zsfzjcw": "ALL",
                                        "zsfwc": ""
                                    },
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_Pooling"
                                }
                                response = self.nifty_res.post_request("/presap/webintf.do?method=find_chanWuMS",
                                                                       data=urlencode(data))
                                if response.json()["code"] == "200":
                                    logger.info("pooling产物查询成功！")
                                    chanwu_data = response.json()["data"]["Samples"]
                                    chanwu_data.sort(key=lambda x: x["zjob_code_item"])
                                    datas = []
                                    data0 = {}
                                    for i in range(len(chanwu_data)):
                                        data_dic = {k: "" if v is None else v for k, v in chanwu_data[i].items()}
                                        datas.append(data_dic["zscheme"])
                                        data0[str(i + 1).zfill(6)] = {"tab1": [{}], "tab2": {}}
                                        data0[str(i + 1).zfill(6)]["tab1"][0]["zcwbh"] = data_dic["zscheme"]
                                        data0[str(i + 1).zfill(6)]["tab1"][0]["zplwkh"] = data_dic["zscheme"]
                                        data0[str(i + 1).zfill(6)]["tab1"][0]["zguid"] = data_dic["zguid"]
                                        data0[str(i + 1).zfill(6)]["tab1"][0]["zcwlx"] = "Pooling文库"
                                        data0[str(i + 1).zfill(6)]["tab2"] = data_dic
                                    data = {
                                        "task": [data0],
                                        "zgxbh": "MAX",
                                        "zjob_code": self.zjob_code,
                                        "token": self.token,
                                        "menuId": "MSTaskProduct_Pooling"
                                    }
                                    response = self.nifty_res.post_request(
                                        "/presap/webintf.do?method=msUpdateTaskResult",
                                        data=urlencode(data))
                                    if response.json()["msg"] == "success" and response.json()["code"] == "200":
                                        logger.info("pooling任务单结果录入成功！")
                                        data = {
                                            "task": {
                                                "zpzxlh_bcode": "",
                                                "zcjdat": "",
                                                "ybbhrq": "",
                                                "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                                "zsfwc_czr": "",
                                                "zsfks_czr": "",
                                                "zcdate": "",
                                                "chkdate": "",
                                                "matnr": rwd_data["matnr"],
                                                "zseqplatform": rwd_data["zseqplatform"],
                                                "zsfks_uzeit": "",
                                                "zdlhwkh": "",
                                                "qdfkrq": "",
                                                "zsampling_datum": "",
                                                "zdlvdate": """ """,
                                                "zgxbh": "MAX",
                                                "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                                "zsigndate": "",
                                                "zsfks": "",
                                                "zsyyn": rwd_data["zsyyn"],
                                                "syncdate": "",
                                                "ddl21": rwd_data["ddl21"],
                                                "yctjrq": "",
                                                "zcxlx": "0",
                                                "zjob_begin_datum": "",
                                                "zsfwc_uzeit": "",
                                                "zsample_quantity": rwd_data["zsample_quantity"],
                                                "zmethod": rwd_data["zmethod"],
                                                "zsfxd_czr": rwd_data["zsfxd_czr"],
                                                "zudate": "",
                                                "addate": "",
                                                "zycdjsflr": "否",
                                                "zjob_status": "未开始",
                                                "zsfxd_datum": rwd_data["zsfxd_datum"],
                                                "werks": "A020",
                                                "zsfks_datum": "",
                                                "ycwcrq": "",
                                                "pre_zsfwc_datum": "",
                                                "zsfwc": "",
                                                "pre_zjob_code": rwd_data["zjob_code"],
                                                "maktx": rwd_data["maktx"],
                                                "zsfwc_datum": "",
                                                "zupdat": "",
                                                "zjob_code": rwd_data["zjob_code"],
                                                "gt_action": "",
                                                "zcbdate": "",
                                                "gt_sendtime": "",
                                                "zfinish_time": "",
                                                "_key": 1,
                                                "_id": 1
                                            },
                                            "token": self.token,
                                            "menuId": "MSTaskProduct_Pooling"
                                        }
                                        response = self.nifty_res.post_request(
                                            "/presap/webintf.do?method=findJobSampleMessageMS", data=urlencode(data))
                                        rwd_detail_data = response.json()["data"]
                                        # data1 = [
                                        #     {
                                        #         "z_plate_lane_no": "MGISEQ-2000PLA0202400388",
                                        #         "no": f"{i + 1:06d}",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[i]["matnr"],
                                        #         "zybs": "1 ",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zplate": rwd_detail_data[i]["zplate"],
                                        #         "werks": "A020",
                                        #         "zindex": 0,
                                        #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                        #         "zgxdm": "MAX",
                                        #         "tab": "WL",
                                        #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                        #         "zjob_code_item": f"{i + 1:06d}",
                                        #         "_key": i+1,
                                        #         "_id": i+2,
                                        #         "rfpnt": "",
                                        #         "matnr": "1000000813",
                                        #         "maktx": "PCOMAT-9-N#不干胶封口膜/品牌&杭州骏荣/规格&127*82，500片",
                                        #         "datuv": "99991231",
                                        #         "zshl": "0.00 ",
                                        #         "zcksjyl": 1,
                                        #         "zbzdw": "ZHA",
                                        #         "zbzyl": 1,
                                        #         "charg": "0000081954",
                                        #         "zsjyl": 1,
                                        #         "zcomment": "",
                                        #         "zcost_condition": "2"
                                        #     }
                                        #     for i in range(len(rwd_detail_data))
                                        # ]
                                        # data2 = [
                                        #     {
                                        #         "z_plate_lane_no": "MGISEQ-2000PLA0202400388",
                                        #         "no": f"{i + 1:06d}",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[i]["matnr"],
                                        #         "zybs": "1 ",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zplate": rwd_detail_data[i]["zplate"],
                                        #         "werks": "A020",
                                        #         "zindex": 0,
                                        #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                        #         "zjob_code_item": f"{i + 1:06d}",
                                        #         "matnr": "CZ10200010577",
                                        #         "maktx": "台式高速冷冻离心机",
                                        #         "zbzyl": 0,
                                        #         "zbzdw": "MIN",
                                        #         "zsjyl": 0.008,
                                        #         "zcost_condition": "2"
                                        #     }
                                        #     for i in range(len(rwd_detail_data))
                                        # ]
                                        # data3 = [
                                        #     {
                                        #         "z_plate_lane_no": "MGISEQ-2000PLA0202400388",
                                        #         "no": f"{i + 1:06d}",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[i]["matnr"],
                                        #         "zybs": "1 ",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zplate": rwd_detail_data[i]["zplate"],
                                        #         "werks": "A020",
                                        #         "zindex": 0,
                                        #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                        #         "zjob_code_item": f"{i + 1:06d}",
                                        #         "matnr": "415-1779",
                                        #         "maktx": "多功能酶标仪",
                                        #         "zbzyl": 0,
                                        #         "zbzdw": "MIN",
                                        #         "zsjyl": 0.021,
                                        #         "zcost_condition": "2"
                                        #     }
                                        #     for i in range(len(rwd_detail_data))
                                        # ]
                                        # data4 = [
                                        #     {
                                        #         "z_plate_lane_no": "MGISEQ-2000PLA0202400388",
                                        #         "no": f"{i + 1:06d}",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[i]["matnr"],
                                        #         "zybs": "1 ",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zplate": rwd_detail_data[i]["zplate"],
                                        #         "werks": "A020",
                                        #         "zindex": 0,
                                        #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                        #         "zjob_code_item": "000001",
                                        #         "matnr": "IES/953/2519",
                                        #         "maktx": "封膜机 S120499",
                                        #         "zbzyl": 0,
                                        #         "zbzdw": "MIN",
                                        #         "zsjyl": 0.025,
                                        #         "zcost_condition": "2"
                                        #     }
                                        #     for i in range(len(rwd_detail_data))
                                        # ]
                                        # data5 = [
                                        #     {
                                        #         "z_plate_lane_no": "MGISEQ-2000PLA0202400388",
                                        #         "no": f"{i + 1:06d}",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[i]["matnr"],
                                        #         "zybs": "1 ",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zplate": rwd_detail_data[i]["zplate"],
                                        #         "werks": "A020",
                                        #         "zindex": 0,
                                        #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                        #         "zjob_code_item": f"{i + 1:06d}",
                                        #         "matnr": "1505001386",
                                        #         "maktx": "自动化工作站",
                                        #         "zbzyl": 0,
                                        #         "zbzdw": "MIN",
                                        #         "zsjyl": 0.5,
                                        #         "zcost_condition": "2"
                                        #     }
                                        #     for i in range(len(rwd_detail_data))
                                        # ]
                                        # data6 = [
                                        #     {
                                        #         "z_plate_lane_no": "MGISEQ-2000PLA0202400388",
                                        #         "no": f"{i + 1:06d}",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[i]["matnr"],
                                        #         "zybs": "1 ",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zplate": rwd_detail_data[i]["zplate"],
                                        #         "werks": "A020",
                                        #         "zproduct_name": rwd_detail_data[i]["maktx"],
                                        #         "zgxdm": "MAX",
                                        #         "tab": "GS",
                                        #         "zjob_code": rwd_detail_data[i]["zjob_code"],
                                        #         "zjob_code_item": f"{i + 1:06d}",
                                        #         "matnr": "A0201755.04.3801",
                                        #         "zbzyl": 0,
                                        #         "zbzdw": "MIN",
                                        #         "zsjyl": 40,
                                        #         "zindex": 0,
                                        #         "zcost_condition": "2"
                                        #     }
                                        #     for i in range(len(rwd_detail_data))
                                        # ]
                                        # total_data = [
                                        #     {
                                        #         "zbzdw": "ZHA",
                                        #         "zbzyl": str(1 * len(rwd_detail_data)) + "000000 ",
                                        #         "no": "0",
                                        #         "zshl": "0.00 ",
                                        #         "matnr": "1000000813",
                                        #         "zsjyl": str(1 * len(rwd_detail_data)) + "000000 ",
                                        #         "zsjdw": "ZHA",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zcksjyl": str(1 * len(rwd_detail_data)) + "000000 ",
                                        #         "zplate": "合计",
                                        #         "zcomment": "",
                                        #         "zgxdm": "MAX",
                                        #         "tab": "WL",
                                        #         "datuv": "99991231",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[0]["matnr"],
                                        #         "zybs": str(len(rwd_detail_data))+" ",
                                        #         "mtart": "Z002",
                                        #         "werks": "A020",
                                        #         "ddtext": "",
                                        #         "charg": "0000081954",
                                        #         "zindex": "1 ",
                                        #         "zproduct_name": rwd_detail_data[0]["maktx"],
                                        #         "rfpnt": "",
                                        #         "maktx": "PCOMAT-9-N#不干胶封口膜/品牌&杭州骏荣/规格&127*82，500片",
                                        #         "zjob_code": self.zjob_code,
                                        #         "zjob_code_item": "000000",
                                        #         "zcatalo": "合计",
                                        #         "cksyjl": 1 * len(rwd_detail_data),
                                        #         "_key": "",
                                        #         "_id": 1,
                                        #         "zcost_condition": "2"
                                        #     },
                                        #     {
                                        #         "zbzyl": "0.0000000 ",
                                        #         "zbzdw": "MIN",
                                        #         "no": "0",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": rwd_detail_data[0]["matnr"],
                                        #         "zybs": str(len(rwd_detail_data)) + " ",
                                        #         "matnr": "CZ10200010577",
                                        #         "zsjyl": str(0.008 * len(rwd_detail_data)) + "0000 ",
                                        #         "zsjdw": "MIN",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zcksjyl": str(0.008 * len(rwd_detail_data)) + "0000 ",
                                        #         "zplate": "合计",
                                        #         "werks": "A020",
                                        #         "zcomment": "",
                                        #         "zindex": "1 ",
                                        #         "zproduct_name": rwd_detail_data[0]["maktx"],
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "maktx": "台式高速冷冻离心机",
                                        #         "zjob_code": self.zjob_code,
                                        #         "zjob_code_item": "000000",
                                        #         "zcatalo": "1 ",
                                        #         "zcost_condition": "2"
                                        #     },
                                        #     {
                                        #         "zbzyl": "0.0000000 ",
                                        #         "zbzdw": "MIN",
                                        #         "no": "0",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": "DX1331",
                                        #         "zybs": str(len(rwd_detail_data)) + " ",
                                        #         "matnr": "415-1779",
                                        #         "zsjyl": str(0.021 * len(rwd_detail_data)) + "0000 ",
                                        #         "zsjdw": "MIN",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zcksjyl": str(0.021 * len(rwd_detail_data)) + "0000 ",
                                        #         "zplate": "合计",
                                        #         "werks": "A020",
                                        #         "zcomment": "",
                                        #         "zindex": "2 ",
                                        #         "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "maktx": "多功能酶标仪",
                                        #         "zjob_code": self.zjob_code,
                                        #         "zjob_code_item": "000000",
                                        #         "zcatalo": "1 ",
                                        #         "zcost_condition": "2"
                                        #     },
                                        #     {
                                        #         "zbzyl": "0.0000000 ",
                                        #         "zbzdw": "MIN",
                                        #         "no": "0",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": "DX1331",
                                        #         "zybs": str(len(rwd_detail_data)) + " ",
                                        #         "matnr": "IES/953/2519",
                                        #         "zsjyl": str(0.025 * len(rwd_detail_data)) + "0000 ",
                                        #         "zsjdw": "MIN",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zcksjyl": str(0.025 * len(rwd_detail_data)) + "0000 ",
                                        #         "zplate": "合计",
                                        #         "werks": "A020",
                                        #         "zcomment": "",
                                        #         "zindex": "3 ",
                                        #         "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "maktx": "封膜机 S120499",
                                        #         "zjob_code": self.zjob_code,
                                        #         "zjob_code_item": "000000",
                                        #         "zcatalo": "1 ",
                                        #         "zcost_condition": "2"
                                        #     },
                                        #     {
                                        #         "zbzyl": "0.0000000 ",
                                        #         "zbzdw": "MIN",
                                        #         "no": "0",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": "DX1331",
                                        #         "zybs": str(len(rwd_detail_data)) + " ",
                                        #         "matnr": "1505001386",
                                        #         "zsjyl": str(0.5 * len(rwd_detail_data)) + "00000 ",
                                        #         "zsjdw": "MIN",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zcksjyl": str(0.5 * len(rwd_detail_data)) + "00000 ",
                                        #         "zplate": "合计",
                                        #         "werks": "A020",
                                        #         "zcomment": "",
                                        #         "zindex": "4 ",
                                        #         "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                        #         "zgxdm": "MAX",
                                        #         "tab": "YQ",
                                        #         "maktx": "自动化工作站",
                                        #         "zjob_code": self.zjob_code,
                                        #         "zjob_code_item": "000000",
                                        #         "zcatalo": "1 ",
                                        #         "zcost_condition": "2"
                                        #     },
                                        #     {
                                        #         "zbzyl": "0.0000000 ",
                                        #         "zbzdw": "MIN",
                                        #         "no": "0",
                                        #         "zpzxlh": "",
                                        #         "zmethod": "1755.04.38",
                                        #         "zproduct": "DX1331",
                                        #         "zybs": str(len(rwd_detail_data)) + " ",
                                        #         "matnr": "A0201755.04.3801",
                                        #         "zsjyl": str(40 * len(rwd_detail_data)) + "0000000 ",
                                        #         "zsjdw": "MIN",
                                        #         "zmethod_name": "MGISEQ-2000_Pooling&BMG定量(Tecan)",
                                        #         "zcksjyl": str(40 * len(rwd_detail_data)) + "0000000 ",
                                        #         "zplate": "合计",
                                        #         "werks": "A020",
                                        #         "zcomment": "",
                                        #         "zindex": "1 ",
                                        #         "zproduct_name": "政府项目-胎儿染色体非整倍体（T21、T18、T13）检测",
                                        #         "zgxdm": "MAX",
                                        #         "tab": "GS",
                                        #         "maktx": "",
                                        #         "zjob_code": self.zjob_code,
                                        #         "zjob_code_item": "000000",
                                        #         "zcatalo": "1 ",
                                        #         "zcost_condition": "2"
                                        #     }
                                        # ]
                                        # data = list(zip(data1, data2, data3, data4, data5, data6, total_data))
                                        # wuliao_data = {
                                        #     "data": [data[0][0]],
                                        #     "token": self.token,
                                        #     "menuId": "MSTaskProduct_Pooling"
                                        # }
                                        # response = self.nifty_res.post_request(
                                        #     "/presap/webintf.do?method=save_liao_gong_fei_nifty",
                                        #     data=urlencode(wuliao_data))
                                        # if response.json()["code"] == "200" and response.json()["msg"] == "success":
                                        #     logger.info(f"pooling料工费结果录入成功！")
                                        data = {
                                            "task": {
                                                "zpzxlh_bcode": "",
                                                "zcjdat": "",
                                                "ybbhrq": "",
                                                "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                                "zsfwc_czr": "",
                                                "zsfks_czr": "",
                                                "zcdate": "",
                                                "chkdate": "",
                                                "matnr": rwd_data["matnr"],
                                                "zseqplatform": rwd_data["zseqplatform"],
                                                "zsfks_uzeit": "",
                                                "zdlhwkh": "",
                                                "qdfkrq": "",
                                                "zsampling_datum": "",
                                                "zdlvdate": """ """,
                                                "zgxbh": "MAX",
                                                "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                                "zsigndate": "",
                                                "zsfks": "",
                                                "zsyyn": rwd_data["zsyyn"],
                                                "syncdate": "",
                                                "ddl21": rwd_data["ddl21"],
                                                "yctjrq": "",
                                                "zcxlx": "0",
                                                "zjob_begin_datum": "",
                                                "zsfwc_uzeit": "",
                                                "zsample_quantity": rwd_data["zsample_quantity"],
                                                "zmethod": rwd_data["zmethod"],
                                                "zsfxd_czr": rwd_data["zsfxd_czr"],
                                                "zudate": "",
                                                "addate": "",
                                                "zycdjsflr": "否",
                                                "zjob_status": "未开始",
                                                "zsfxd_datum": rwd_data["zsfxd_datum"],
                                                "werks": "A020",
                                                "zsfks_datum": "",
                                                "ycwcrq": "",
                                                "pre_zsfwc_datum": "",
                                                "zsfwc": "",
                                                "pre_zjob_code": rwd_data["zjob_code"],
                                                "maktx": rwd_data["maktx"],
                                                "zsfwc_datum": "",
                                                "zupdat": "",
                                                "zjob_code": rwd_data["zjob_code"],
                                                "gt_action": "",
                                                "zcbdate": "",
                                                "gt_sendtime": "",
                                                "zfinish_time": "",
                                                "_key": 1,
                                                "_id": 1
                                            },
                                            "token": self.token,
                                            "menuId": "MSTaskProduct_Pooling"
                                        }
                                        response = self.nifty_res.post_request(
                                            "/presap/webintf.do?method=findJobSampleMessageMS",
                                            data=urlencode(data))
                                        rwd_detail_data = response.json()["data"]
                                        datas = [
                                            {
                                                "isSetData": "X",
                                                "zplate_bcode": rwd_detail_data[i]["zplate_bcode"],
                                                "_pre_zfieldnr_": "",
                                                "_id": 4,
                                                "zpzxlh_bcode": "",
                                                "zcjdat": "",
                                                "ybbhrq": "",
                                                "zsfxd_uzeit": rwd_detail_data[i]["zsfxd_uzeit"],
                                                "zsfwc_czr": "",
                                                "zsfks_czr": rwd_detail_data[i]["zsfks_czr"],
                                                "zcdate": "",
                                                "chkdate": "",
                                                "matnr": rwd_detail_data[i]["matnr"],
                                                "zseqplatform": rwd_detail_data[i]["zseqplatform"],
                                                "zsfks_uzeit": rwd_detail_data[i]["zsfxd_uzeit"],
                                                "zdlhwkh": "",
                                                "qdfkrq": "",
                                                "zsampling_datum": "",
                                                "zdlvdate": """ """,
                                                "zgxbh": "MAX",
                                                "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                                "zsigndate": "",
                                                "zsfks": "X",
                                                "zsyyn": rwd_detail_data[i]["zsyyn"],
                                                "syncdate": "",
                                                "ddl21": rwd_detail_data[i]["ddl21"],
                                                "yctjrq": "",
                                                "zcxlx": "0",
                                                "zjob_begin_datum": "",
                                                "zsfwc_uzeit": "",
                                                "zsample_quantity": str(len(rwd_detail_data)),
                                                "zmethod": "1755.04.38",
                                                "zsfxd_czr": rwd_detail_data[i]["zsfxd_czr"],
                                                "zudate": "",
                                                "addate": "",
                                                "zycdjsflr": "否",
                                                "zjob_status": "已开始",
                                                "zsfxd_datum": rwd_detail_data[i]["zsfxd_datum"],
                                                "werks": "A020",
                                                "zsfks_datum": rwd_detail_data[i]["zsfks_datum"],
                                                "ycwcrq": "",
                                                "pre_zsfwc_datum": "",
                                                "zsfwc": "",
                                                "pre_zjob_code": rwd_data["zjob_code"],
                                                "maktx": rwd_detail_data[i]["maktx"],
                                                "zsfwc_datum": "",
                                                "zupdat": "",
                                                "zjob_code": rwd_detail_data[i]["zjob_code"],
                                                "gt_action": "",
                                                "zcbdate": "",
                                                "gt_sendtime": "",
                                                "zfinish_time": "",
                                                "_key": 4,
                                                "zfieldnr": [
                                                    {
                                                        "zhcxh": "00003",
                                                        "zcdate": "20211111",
                                                        "zsrlx": "STXT",
                                                        "zhcmc": "自动化移液工作站编号",
                                                        "zctime": "180000",
                                                        "zhcxl": "自动化移液工作站",
                                                        "zcreator": "zhongyingying",
                                                        "werks": "A020",
                                                        "zqdxh": "00001",
                                                        "zsfzh": "",
                                                        "zgxbh": "MAX",
                                                        "zsfqy": "X",
                                                        "zhcdl": "仪器选择",
                                                        "zshow": "X",
                                                        "ztitle": "仪器维护",
                                                        "zcxlx": "0",
                                                        "zhcnr": "A54012005431N00001"
                                                    }
                                                ],
                                                "zqbnr": {
                                                    "仪器选择": {
                                                        "自动化移液工作站编号": [
                                                            {
                                                                "zhcxh": "00001",
                                                                "zcdate": "20211111",
                                                                "zsrlx": "STXT",
                                                                "zhcmc": "自动化移液工作站编号",
                                                                "zctime": "180000",
                                                                "zhcxl": "自动化移液工作站",
                                                                "zcreator": "zhongyingying",
                                                                "werks": "A020",
                                                                "zqdxh": "00001",
                                                                "zsfzh": "",
                                                                "zgxbh": "MAX",
                                                                "zsfqy": "X",
                                                                "zhcdl": "仪器选择",
                                                                "ztitle": "仪器维护",
                                                                "zcxlx": "0",
                                                                "zhcnr": "119005499000096"
                                                            },
                                                            {
                                                                "zhcxh": "00002",
                                                                "zcdate": "20211111",
                                                                "zsrlx": "STXT",
                                                                "zhcmc": "自动化移液工作站编号",
                                                                "zctime": "180000",
                                                                "zhcxl": "自动化移液工作站",
                                                                "zcreator": "zhongyingying",
                                                                "werks": "A020",
                                                                "zqdxh": "00001",
                                                                "zsfzh": "",
                                                                "zgxbh": "MAX",
                                                                "zsfqy": "X",
                                                                "zhcdl": "仪器选择",
                                                                "ztitle": "仪器维护",
                                                                "zcxlx": "0",
                                                                "zhcnr": "12000540430019"
                                                            },
                                                            {
                                                                "zhcxh": "00003",
                                                                "zcdate": "20211111",
                                                                "zsrlx": "STXT",
                                                                "zhcmc": "自动化移液工作站编号",
                                                                "zctime": "180000",
                                                                "zhcxl": "自动化移液工作站",
                                                                "zcreator": "zhongyingying",
                                                                "werks": "A020",
                                                                "zqdxh": "00001",
                                                                "zsfzh": "",
                                                                "zgxbh": "MAX",
                                                                "zsfqy": "X",
                                                                "zhcdl": "仪器选择",
                                                                "zshow": "X",
                                                                "ztitle": "仪器维护",
                                                                "zcxlx": "0",
                                                                "zhcnr": "A54012005431N00001"
                                                            }
                                                        ]
                                                    }
                                                }
                                            }
                                            for i in range(len(rwd_detail_data))
                                        ]
                                        data = {
                                            "datas": datas,
                                            "token": self.token,
                                            "menuId": "MSTaskProduct_Pooling"
                                        }
                                        response = self.nifty_res.post_request(
                                        "/presap/webintf.do?method=saveTaskReagentInfo",
                                        data=urlencode(data))
                                        if response.json()["code"] == "200" and response.json()["msg"] == "success":
                                            logger.info(f"pooling试剂耗材录入成功！")
                                            data = {
                                                "task": {
                                                    "zpzxlh_bcode": "",
                                                    "zcjdat": "",
                                                    "ybbhrq": "",
                                                    "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                                    "zsfwc_czr": "",
                                                    "zsfks_czr": "",
                                                    "zcdate": "",
                                                    "chkdate": "",
                                                    "matnr": rwd_data["matnr"],
                                                    "zseqplatform": rwd_data["zseqplatform"],
                                                    "zsfks_uzeit": "",
                                                    "zdlhwkh": "",
                                                    "qdfkrq": "",
                                                    "zsampling_datum": "",
                                                    "zdlvdate": """ """,
                                                    "zgxbh": "MAX",
                                                    "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                                    "zsigndate": "",
                                                    "zsfks": "",
                                                    "zsyyn": rwd_data["zsyyn"],
                                                    "syncdate": "",
                                                    "ddl21": rwd_data["ddl21"],
                                                    "yctjrq": "",
                                                    "zcxlx": "0",
                                                    "zjob_begin_datum": "",
                                                    "zsfwc_uzeit": "",
                                                    "zsample_quantity": rwd_data["zsample_quantity"],
                                                    "zmethod": rwd_data["zmethod"],
                                                    "zsfxd_czr": rwd_data["zsfxd_czr"],
                                                    "zudate": "",
                                                    "addate": "",
                                                    "zycdjsflr": "否",
                                                    "zjob_status": "未开始",
                                                    "zsfxd_datum": rwd_data["zsfxd_datum"],
                                                    "werks": "A020",
                                                    "zsfks_datum": "",
                                                    "ycwcrq": "",
                                                    "pre_zsfwc_datum": "",
                                                    "zsfwc": "",
                                                    "pre_zjob_code": rwd_data["zjob_code"],
                                                    "maktx": rwd_data["maktx"],
                                                    "zsfwc_datum": "",
                                                    "zupdat": "",
                                                    "zjob_code": rwd_data["zjob_code"],
                                                    "gt_action": "",
                                                    "zcbdate": "",
                                                    "gt_sendtime": "",
                                                    "zfinish_time": "",
                                                    "_key": 1,
                                                    "_id": 1
                                                },
                                                "token": self.token,
                                                "menuId": "MSTaskProduct_Pooling"
                                            }
                                            response = self.nifty_res.post_request(
                                                "/presap/webintf.do?method=findJobSampleMessageMS",
                                                data=urlencode(data))
                                            rwd_detail_data = response.json()["data"]
                                            datas = []
                                            for i in range(len(rwd_detail_data)):
                                                data_dic = {k: "" if v is None else v for k, v in
                                                            rwd_detail_data[i].items()}
                                                data_dic["_id"] = i + 1
                                                data_dic["zlgfsflr"] = "X"
                                                datas.append(data_dic)
                                            complete_data = {
                                                "req": datas,
                                                "token": self.token,
                                                "menuId": "MSTaskProduct_Pooling"
                                            }
                                            complete_response = self.nifty_res.post_request(
                                                "/presap/webintf.do?method=task_sample_item_finishms",
                                                data=urlencode(complete_data))
                                            if complete_response.json()["code"] == "200" and \
                                                    complete_response.json()[
                                                        "msg"] == "SAP数据库更新成功!":
                                                pooling_scheme = []
                                                for i in range(len(rwd_detail_data)):
                                                    pooling_scheme.append(rwd_detail_data[i]["zcwbh"])
                                                logger.info("pooling任务完成成功！")
                                                unique_list = [x for i, x in enumerate(pooling_scheme) if
                                                               pooling_scheme.index(x) == i]
                                                self.pooling_scheme = unique_list
                                            else:
                                                logger.error(f"pooling任务完成失败，返回结果：{response}")
                                                raise Exception(f"pooling任务完成失败，返回结果：{response}")
                                        else:
                                            logger.error(f"pooling试剂耗材录入失败，返回结果：{response}")
                                            raise Exception(f"pooling试剂耗材录入失败，返回结果：{response}")
                                        # else:
                                        #     logger.error(f"ppooling料工费结果录入失败，返回结果：{response}")
                                        #     raise Exception(f"pooling料工费结果录入失败，返回结果：{response}")
                                    else:
                                        logger.error(f"pooling任务单结果录入失败，返回结果：{response}")
                                        raise Exception(f"pooling任务单结果录入失败，返回结果：{response}")
                                else:
                                    logger.error(f"pooling产物查询失败，返回结果：{response}")
                                    raise Exception(f"pooling产物查询失败，返回结果：{response}")
                        else:
                            logger.error(f"pooling任务开始失败，返回结果：{response}")
                            raise Exception(f"pooling任务开始失败，返回结果：{response}")
                    else:
                        logger.error(f"pooling任务单详情查询失败，返回结果：{response}")
                        raise Exception(f"pooling任务单详情查询失败，返回结果：{response}")
                else:
                    logger.error(f"pooling任务单列表查询失败，返回结果：{response}")
                    raise Exception(f"pooling任务单列表查询失败，返回结果：{response}")
            else:
                logger.error(f"pooling下达任务失败，返回结果：{response}")
                raise Exception(f"pooling下达任务失败，返回结果：{response}")
        else:
            logger.error(f"pooling新建任务列表查询失败，返回结果：{response}")
            raise Exception(f"pooling新建任务列表查询失败，返回结果：{response}")

    def sequencing(self, makednb_task_code=None):
        """
        上机测序
        @param makednb_task_code: makednb任务单号
        @return: self.sequencing_task_code, self.chip_num 任务号和芯片号
        """
        zmethod = ""
        zmethod_des = ""
        stage_code = "MBT"
        stage_name = "machine500order"
        if makednb_task_code:
            self.makednb_task_code = makednb_task_code
            # 设置上机信息
        if self.config_info["sequencePlatform"] == "DNBSEQ-T7":
            zmethod = "1755.04.51"
            zmethod_des ="1755.04.51深圳-医学测序组-上机-DNBSEQ-T7(SE50)"
        elif self.config_info["sequencePlatform"] == "MGISEQ-2000":
            zmethod = "1755.04.43"
            zmethod_des = "1755.04.43深圳-医学测序组-上机-MGISEQ-2000(SE35)"
        # 设置实验信息
        lane_count = None
        if self.config_info["sequencePlatform"] == "DNBSEQ-T7":
            lane_count = list()
        elif self.config_info["sequencePlatform"] == "MGISEQ-2000":
            lane_count = ["01", "02", "03", "04"]
        task_detail = {
            # "_key": 2,
            "lane_count": lane_count,
            "_lane_btn": False,
            # "zindex": "27",
            # "zindex_name": "27",
            # "radat": "6M",
            "ztx": self.config_info["indexConfig"],
            # "zzkxh": None,  # zkp
            "zsyyn": "刘梓兴",
            # "pernr": "00000218",
            "zmethod": zmethod,
            "zmethod_des": zmethod_des,
            "zshould_sj_datum": time.strftime("%Y-%m-%d %H:%M:%S")
        }
        # 第一步： 根据makeDNB任务单号查询待上机测序的任务单
        query_unassign_task_data = {
            "task": {"pre_zjob_code": self.makednb_task_code},
            "zgxbh": stage_code,
            "token": self.token
        }
        logger.info("上机测序:开始查询待处理数据列表！")
        logger.info("上机测序:查询待处理数据:{}", query_unassign_task_data)
        query_unassign_task_respone = self.nifty_res.post_request("/presap/webintf.do?method=task_assign_samplems", data=urlencode(query_unassign_task_data)).json()
        if query_unassign_task_respone["code"] == "200" and query_unassign_task_respone["total"] > 0:
            logger.info("{}：查询待处理数据列表成功！", stage_name)
            query_unassign_task_list = query_unassign_task_respone["data"]
            # 第二步：上机测序任务下达
            for i in range(len(query_unassign_task_list)):
                task_detail_copy = copy.deepcopy(task_detail)
                query_unassign_task_list[i].update(task_detail_copy)
                query_unassign_task_list[i]["_key"] = i + 1
                query_unassign_task_list[i]["zindex"] = query_unassign_task_list[i]["pre_zindex"]
                query_unassign_task_list[i]["zindex_name"] = query_unassign_task_list[i]["pre_zindex_name"]
                query_unassign_task_list[i]["radat"] = query_unassign_task_list[i]["pre_radat"]
                if "平台阴阳质控" in query_unassign_task_list[i]["maktx"]:
                    query_unassign_task_list[i]["zzkxh"] = "X"
                    query_unassign_task_list[i]["zsampling_datum"] = "00000000"
                    query_unassign_task_list[i]["zreceiveddate"] = "00000000"
            logger.info("{}：任务开始下达！", stage_name)
            assign_task_data = {"task": json.dumps(query_unassign_task_list, ensure_ascii=False),
                                "token": self.token}
            assign_task_respone = self.nifty_res.post_request("/presap/webintf.do?method=saveTaskAssignSampleMS", data=urlencode(assign_task_data)).json()

            if assign_task_respone["code"] == "200":
                if assign_task_respone["msg"] is None:
                    self.sequencing_task_code = assign_task_respone["data"][0]["zjob_code"]
                else:
                    self.sequencing_task_code = assign_task_respone["msg"]
                logger.debug("assign_task_respone is: {}", assign_task_respone)
                logger.info("上机测序：任务下达成功！", stage_name)
                logger.info("上机测序任务单号：{}", self.sequencing_task_code)
                if self.config_info["isChipCode"] == "是":  # 判断是否自定义新芯片号
                    chip_id = self.config_info["chipCode"]  # 获取表格芯片号
                else: # 根据测序平台类型自动生成芯片号
                    current_time = time.time()
                    if self.config_info["sequencePlatform"] == "DNBSEQ-T7":
                        chip_id = "E"+ str(int(current_time))[:10]
                    elif self.config_info["sequencePlatform"] == "PMSEQ-4500":
                        chip_id = "P"+ str(int(current_time))[:10]
                    else:
                        chip_id = "V"+ str(int(current_time))[:10]
                start_time = time.strftime("%Y-%m-%d %H:%M:%S")  # 上机开始时间
                # 香港的在千寻上，要插入片区为A450的
                if self.area_code == "A440":
                    self.area_code = "A450"
                    library = "香港实验室"
                elif self.area_code == "A230":
                    library = "天津实验室"
                else:
                    library = "深圳实验室"
                chihiro_db_config = GetConfig(configname=self.configname).get_chihiro_db_config()
                update_chihiro_db = UpdateDB(handler_db=HandleDB(chihiro_db_config['host'], chihiro_db_config['port'], chihiro_db_config['user'], chihiro_db_config['password'],chihiro_db_config['database']))
                # 千寻测试环境数据库插入芯片信息
                is_chip_exist, is_chip_finish = update_chihiro_db.update_chihiro_db_chip_info(chip_id, self.config_info["machineNum"], start_time,
                                                                      self.config_info["sequencePlatform"].replace("-", ""),
                                                                      self.config_info["sequenceType"], self.area_code,
                                                                      library)
                self.chip_num = chip_id
                # 第三步：上机测序任务生产
                # 准备任务生产需要的数据
                task_info = {
                    "zload_machn_bcode": self.config_info["machineNum"],
                    "zload_location": "A",
                    "zpzxlh_bcode": self.chip_num,
                    "zindextype_cn": "",
                    "check": False,
                    "_operate": {
                        "key": None,
                        "ref": None,
                        "props": {
                            "size": "small",
                            "children": "条码录入",
                            "loading": False,
                            "ghost": False,
                            "block": False,
                            "htmlType": "button"
                        },
                        "_owner": None
                    },
                }
                # 开始生产任务
                # 查询任务样本信息
                result_info = self.query_sample_msg(stage_code, stage_name, self.sequencing_task_code)
                zpoint_list = []
                task_start_list = []
                for i in range(len(result_info)):
                    if result_info[i]["zpoint"] not in zpoint_list:
                        zpoint_list.append(result_info[i]["zpoint"])
                        task_start_list.append(result_info[i])
                    else:
                        continue
                for i in range(len(task_start_list)):
                    task_info_copy = copy.deepcopy(task_info)
                    task_start_list[i].update(task_info_copy)
                    task_start_list[i]["zcatalo_before"] = task_start_list[i]["zcatalo"].split("-")[0]
                    task_start_list[i]["_id"] = i + 1
                    del task_start_list[i]["zjob_code_item"]
                    if task_start_list[i]["zpoint"] == "a-1":
                        task_start_list[i]["pz_lane_no"] = str(task_start_list[i]["zpzxlh"]) + "-01"
                    elif task_start_list[i]["zpoint"] == "b-1":
                        task_start_list[i]["pz_lane_no"] = str(task_start_list[i]["zpzxlh"]) + "-02"
                    elif task_start_list[i]["zpoint"] == "c-1":
                        task_start_list[i]["pz_lane_no"] = str(task_start_list[i]["zpzxlh"]) + "-03"
                    elif task_start_list[i]["zpoint"] == "d-1":
                        task_start_list[i]["pz_lane_no"] = str(task_start_list[i]["zpzxlh"]) + "-04"
                task_start_data = {
                    "task": json.dumps(task_start_list, ensure_ascii=False),
                    "zjob_code": self.sequencing_task_code,
                    "zgxbh": stage_code,
                    "token": self.token
                }
                logger.info("上机测序：{}开始上机测序任务！", self.sequencing_task_code)
                task_start_respone = self.nifty_res.post_request("/presap/webintf.do?method=msoperatastart_task", data=urlencode(task_start_data)).json()
                if task_start_respone["code"] == "200" and task_start_respone["msg"] == "success":
                    logger.info("上机测序：{}开始上机测序任务成功！", stage_name, self.sequencing_task_code)
                    logger.debug("task_start_respone：{}", task_start_respone)
                    # 上机测序任务单结果录入
                    # 查询任务产物信息
                    chanwu_result_info = self.query_chanwu_msg(stage_code, stage_name, self.sequencing_task_code)

                    data = {}
                    for i in range(len(chanwu_result_info)):
                        data[str(i + 1).zfill(6)] = {"tab1": [], "tab2": {}}
                        data[str(i + 1).zfill(6)]["tab2"] = chanwu_result_info[i]
                    task_result_data = {"task": json.dumps([data], ensure_ascii=False),
                                    "zgxbh": stage_code,
                                    "zjob_code": self.sequencing_task_code,
                                    "token": self.token
                                    }
                    logger.info("上机测序：{}开始任务单结果录入！", self.sequencing_task_code)
                    task_result_respone = self.nifty_res.post_request("/presap/webintf.do?method=msUpdateTaskResult", data=urlencode(task_result_data)).json()
                    if task_result_respone["code"] == "200" and task_result_respone["msg"] == "success":
                        logger.info("上机测序：{}任务单结果录入成功！", self.sequencing_task_code)
                        logger.debug("task_result_respone：{}", task_result_respone)
                        # 上机测序任务单完成，等待nifty定时器自动查询千寻芯片状态
                        if is_chip_exist and is_chip_finish:
                            logger.info("上机测序：{}上机测序任务单完成成功！", self.sequencing_task_code)
                            return self.sequencing_task_code, self.chip_num  # 返回任务号和芯片号
                    else:
                        logger.error("上机测序：{}任务单结果录入失败！", self.sequencing_task_code)
                        logger.debug("task_result_respone：{}", task_result_respone)
                        raise Exception(self.sequencing_task_code + "任务单结果录入失败！")
                else:
                    logger.error("上机测序：{}开始上机测序任务失败！", self.sequencing_task_code)
                    logger.debug("task_start_respone：{}", task_start_respone)
                    raise Exception(self.sequencing_task_code + "开始上机测序任务失败！")

            else:
                logger.error("{}：任务下达失败！", stage_name)
                logger.debug("assign_task_respone：{}", assign_task_respone)
                raise Exception(stage_name + "任务下达失败！")
        else:
            logger.error("查询失败或数据不存在！")
            logger.debug("query_unassign_task_respone：{}", query_unassign_task_respone)
            raise Exception(stage_name + "查询失败或数据不存在！")

    def quality_inspection_products(self,zjob_code=None):
        """
        质检产物
        :return:
        """
        if zjob_code:
            zjob_code = zjob_code
        else:
            zjob_code = self.zjob_code
        data = {
            "task": {
                "zfinish_date": "",
                "zfinish_dateend": "",
                "zsfxd_datum": "",
                "zsfxd_datumend": "",
                "zsfks_datum": "",
                "zsfks_datumend": "",
                "zjob_code": zjob_code,
                "zsfwc": "X",
                "task": "query"
            },
            "pageNumber": "1",
            "zgxbh": "MAX",
            "pageSize": "1000",
            "token": self.token,
            "menuId": "MSTaskProduct_Pooling"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=findJobMS", data=urlencode(data))
        if response.json()["code"] == "200" and (response.json()["data"])[0]["zjob_code"] == zjob_code:
            logger.info("pooling任务单已完成列表数据查询成功！")
            rwd_data = (response.json()["data"])[0]
            data = {
                "task": {
                    "zpzxlh_bcode": "",
                    "zcjdat": "",
                    "ybbhrq": "",
                    "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                    "zsfwc_czr": rwd_data["zsfwc_czr"],
                    "zsfks_czr": rwd_data["zsfks_czr"],
                    "zcdate": "",
                    "chkdate": "",
                    "matnr": rwd_data["matnr"],
                    "zseqplatform": rwd_data["zseqplatform"],
                    "zsfks_uzeit": rwd_data["zsfks_uzeit"],
                    "zdlhwkh": "",
                    "qdfkrq": "",
                    "zsampling_datum": "",
                    "zdlvdate": """ """,
                    "zgxbh": "MAX",
                    "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                    "zsigndate": "",
                    "zsfks": "X",
                    "zsyyn": rwd_data["zsyyn"],
                    "syncdate": "",
                    "ddl21": "",
                    "yctjrq": "",
                    "zcxlx": "0",
                    "zjob_begin_datum": "",
                    "zsfwc_uzeit": rwd_data["zsfwc_uzeit"],
                    "zsample_quantity": rwd_data["zsample_quantity"],
                    "zmethod": rwd_data["zmethod"],
                    "zsfxd_czr": rwd_data["zsfxd_czr"],
                    "zudate": "",
                    "addate": "",
                    "zycdjsflr": "否",
                    "zjob_status": "已完成",
                    "zsfxd_datum": rwd_data["zsfxd_datum"],
                    "werks": "A020",
                    "zsfks_datum": rwd_data["zsfks_datum"],
                    "ycwcrq": "",
                    "pre_zsfwc_datum": "",
                    "zsfwc": "X",
                    "pre_zjob_code": rwd_data["pre_zjob_code"],
                    "maktx": rwd_data["maktx"],
                    "zsfwc_datum": rwd_data["zsfwc_datum"],
                    "zupdat": "",
                    "zjob_code": rwd_data["zjob_code"],
                    "gt_action": "",
                    "zcbdate": "",
                    "gt_sendtime": "",
                    "zfinish_time": "",
                    "_key": 1,
                    "_id": 1
                },
                "token": self.token,
                "menuId": "MSTaskProduct_Pooling"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS", data=urlencode(data))
            if response.json()["code"] == "200":
                logger.info("pooling任务单已完成列表详情数据查询成功！")
                rwd_detail_data = response.json()["data"]
                datas = []
                for i in range(len(rwd_detail_data)):
                    dic_data = {k: "" if v is None else v for k, v in rwd_detail_data[i].items()}
                    dic_data.update(dic_data)
                    dic_data["matnr"] = "RC0090"
                    dic_data["zxmbm"] = "F15ZF1I1SJ2420"
                    dic_data["zplwkh"] = dic_data["zplwkh"]+"Z"
                    dic_data["zzjcwbs"] = "X"
                    datas.append(dic_data)
                data = {
                    "datas": datas,
                    "token": self.token,
                    "menuId": "MSTaskProduct_Pooling"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=saveQcResult", data=urlencode(data))
                if response.json()["code"] == "200" and response.json()["msg"] == "success":
                    logger.info("pooling任务单已完成列表质检产物成功！")
                    self.zjcwbh = rwd_detail_data[0]["zplwkh"] + "Z"
                    return self.zjcwbh
                else:
                    logger.error(f"pooling任务单已完成列表质检产物失败，返回结果：{response}")
                    raise Exception(f"pooling任务单已完成列表质检产物失败，返回结果：{response}")
            else:
                logger.error(f"pooling任务单已完成列表详情数据查询失败，返回结果：{response}")
                raise Exception(f"pooling任务单已完成列表详情数据查询失败，返回结果：{response}")
        else:
            logger.error(f"pooling任务单已完成列表数据查询失败，返回结果：{response}")
            raise Exception(f"pooling任务单已完成列表数据查询失败，返回结果：{response}")

    def dlhhorder(self,pooling_scheme=None):
        """
        单链环化
        :return:
        """
        if pooling_scheme:
            pooling_scheme = pooling_scheme
        else:
            pooling_scheme = self.pooling_scheme
        pooling_scheme = ','.join(pooling_scheme)
        detail = {
            "zplate_x": 12,
            "zplate_y": 8,
            "zindextype": "controller",
            "zsyyn": "何华",
            "zmethod": "1755.04.12",
            "zmethod_des": "1755.04.12深圳_文库质控与测序组_单链&环化&定量（手工）_BGISEQ-500",
            "zshould_finish_datum": self.date_str,
            "ztx": self.config_info["indexConfig"],
            "zshould_finish_uzeit": self.time_str
        }
        data = {
            "task": {
                "zplwkh": pooling_scheme
            },
            "pageNumber": "1",
            "pageSize": "1000",
            "zgxbh": "MBL",
            "token": self.token,
            "menuId": "MSTaskMaster_DangLianHuangHua_DLHHOrder"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=task_assign_samplems", data=urlencode(data))
        if response.json()["code"] == "200" :
            logger.info("单链环化新建任务列表查询成功！")
            res_data = response.json()["data"]
            pooling_scheme_list = list(dict.fromkeys(pooling_scheme.split(",")))
            datas = []
            for i in range(len(res_data)):
                dic_data = {k: "" if v is None else v for k, v in res_data[i].items()}
                dic_data.update(detail)
                dic_data["_key"] = i + 1
                dic_data["zcwbh"] = dic_data["pre_zcwbh"]
                dic_data["zplwkh"] = dic_data["pre_zcwbh"]
                dic_data["radat"] = dic_data["pre_radat"]
                dic_data["zindex"] = dic_data["pre_zindex"]
                dic_data["zindex_name"] = dic_data["pre_zindex_name"]
                dic_data["zscheme"] = dic_data["pre_zscheme"]
                dic_data["zplate"] = "01"
                dic_data["zreceiveddate"] = "00000000"
                dic_data["ztubetype"] = ""
                if dic_data["pre_zscheme"] == pooling_scheme_list[0]:
                    dic_data["zpoint"] = "a-1"
                elif dic_data["pre_zscheme"] == pooling_scheme_list[1]:
                    dic_data["zpoint"] = "b-1"
                elif dic_data["pre_zscheme"] == pooling_scheme_list[2]:
                    dic_data["zpoint"] = "c-1"
                elif dic_data["pre_zscheme"] == pooling_scheme_list[3]:
                    dic_data["zpoint"] = "d-1"
                datas.append(dic_data)
            data = {
                "task": datas,
                "token": self.token,
                "menuId": "MSTaskMaster_DangLianHuangHua_DLHHOrder"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=saveTaskAssignSampleMS",
                                                   data=urlencode(data))
            if response.json()["code"] == "200":
                logger.info("单链环化下达任务成功！")
                self.zjob_code = response.json()["msg"]
                data = {
                    "task": {
                        "zsfxd_datum": "",
                        "zsfxd_datumend": "",
                        "zsfks_datum": "",
                        "zsfks_datumend": "",
                        "zjob_code": self.zjob_code,
                        "zsfwc": "",
                        "task": "query"
                    },
                    "token": self.token,
                    "pageNumber": "1",
                    "zgxbh": "MBL",
                    "pageSize": "1000",
                    "menuId": "MSTaskProduct_DangLianHuangHua"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=findJobMS",data=urlencode(data))
                if response.json()["code"] == "200":
                    logger.info("单链环化任务单列表查询成功！")
                    rwd_data = response.json()["data"][0]
                    data = {
                        "task": {
                            "zpzxlh_bcode": "",
                            "zcjdat": "",
                            "ybbhrq": "",
                            "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                            "zsfwc_czr": "",
                            "zsfks_czr": "",
                            "zcdate": "",
                            "chkdate": "",
                            "matnr": rwd_data["matnr"],
                            "zseqplatform": rwd_data["zseqplatform"],
                            "zsfks_uzeit": "",
                            "zdlhwkh": "",
                            "qdfkrq": "",
                            "zsampling_datum": "",
                            "zdlvdate": """ """,
                            "zgxbh": "MAX",
                            "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                            "zsigndate": "",
                            "zsfks": "",
                            "zsyyn": rwd_data["zsyyn"],
                            "syncdate": "",
                            "ddl21": rwd_data["ddl21"],
                            "yctjrq": "",
                            "zcxlx": "0",
                            "zjob_begin_datum": "",
                            "zsfwc_uzeit": "",
                            "zsample_quantity": rwd_data["zsample_quantity"],
                            "zmethod": rwd_data["zmethod"],
                            "zsfxd_czr": rwd_data["zsfxd_czr"],
                            "zudate": "",
                            "addate": "",
                            "zycdjsflr": "否",
                            "zjob_status": "未开始",
                            "zsfxd_datum": rwd_data["zsfxd_datum"],
                            "werks": "A020",
                            "zsfks_datum": "",
                            "ycwcrq": "",
                            "pre_zsfwc_datum": "",
                            "zsfwc": "",
                            "pre_zjob_code": rwd_data["zjob_code"],
                            "maktx": rwd_data["maktx"],
                            "zsfwc_datum": "",
                            "zupdat": "",
                            "zjob_code": rwd_data["zjob_code"],
                            "gt_action": "",
                            "zcbdate": "",
                            "gt_sendtime": "",
                            "zfinish_time": "",
                            "_key": 1,
                            "_id": 1
                        },
                        "token": self.token,
                        "menuId": "MSTaskProduct_DangLianHuangHua"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS",
                                                           data=urlencode(data))
                    rwd_detail_data = response.json()["data"]
                    data = {
                        "zjob_code": self.zjob_code,
                        "token": self.token,
                        "menuId": "MSTaskProduct_DangLianHuangHua"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=msStartTask", data=urlencode(data))
                    if response.json()["code"] == "200":
                        data = {
                            "task": {
                                "zpzxlh_bcode": "",
                                "zcjdat": "",
                                "ybbhrq": "",
                                "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                "zsfwc_czr": "",
                                "zsfks_czr": "",
                                "zcdate": "",
                                "chkdate": "",
                                "matnr": rwd_data["matnr"],
                                "zseqplatform": rwd_data["zseqplatform"],
                                "zsfks_uzeit": "",
                                "zdlhwkh": "",
                                "qdfkrq": "",
                                "zsampling_datum": "",
                                "zdlvdate": """ """,
                                "zgxbh": "MAX",
                                "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                "zsigndate": "",
                                "zsfks": "",
                                "zsyyn": rwd_data["zsyyn"],
                                "syncdate": "",
                                "ddl21": rwd_data["ddl21"],
                                "yctjrq": "",
                                "zcxlx": "0",
                                "zjob_begin_datum": "",
                                "zsfwc_uzeit": "",
                                "zsample_quantity": rwd_data["zsample_quantity"],
                                "zmethod": rwd_data["zmethod"],
                                "zsfxd_czr": rwd_data["zsfxd_czr"],
                                "zudate": "",
                                "addate": "",
                                "zycdjsflr": "否",
                                "zjob_status": "未开始",
                                "zsfxd_datum": rwd_data["zsfxd_datum"],
                                "werks": "A020",
                                "zsfks_datum": "",
                                "ycwcrq": "",
                                "pre_zsfwc_datum": "",
                                "zsfwc": "",
                                "pre_zjob_code": rwd_data["zjob_code"],
                                "maktx": rwd_data["maktx"],
                                "zsfwc_datum": "",
                                "zupdat": "",
                                "zjob_code": rwd_data["zjob_code"],
                                "gt_action": "",
                                "zcbdate": "",
                                "gt_sendtime": "",
                                "zfinish_time": "",
                                "_key": 1,
                                "_id": 1
                            },
                            "token": self.token,
                            "menuId": "MSTaskProduct_DangLianHuangHua"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS",
                                                               data=urlencode(data))
                        rwd_detail_data = response.json()["data"]
                        logger.info("单链环化任务开始成功！")
                        if self.abnormal_input == '1':
                            data = {
                                "zgxbh": "MBL",
                                "werks": "A020",
                                "arbpl": "",
                                "token": self.token,
                                "menuId": "MSTaskProduct_DangLianHuangHua"
                            }
                            response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                                   data=urlencode(data))
                            if response.json()["code"] == "200" and len(response.json()["data"]) > 0:
                                logger.info("异常查询成功！")
                                exception_data = response.json()["data"][0]
                                datas = []
                                for i in range(len(rwd_detail_data)):
                                    data_dic = {k: "" if v is None else v for k, v in rwd_detail_data[i].items()}
                                    exception_data_dic = {k: "" if v is None else v for k, v in exception_data.items()}
                                    data_dic.update(exception_data_dic)
                                    data_dic["_id"] = i + 1
                                    data_dic["zcjnam"] = "auto_test"
                                    data_dic["zcjdat"] = self.date_str
                                    data_dic["zcjuzt"] = self.time_str
                                    data_dic["dateTime"] = self.datatime_str
                                    data_dic["yc_status"] = "A"
                                    datas.append(data_dic)
                                data = {
                                    "task": datas,
                                    "werks": "A020",
                                    "arbpl": "",
                                    "zjob_code": rwd_data["zjob_code"],
                                    "zgxbh": "MBL",
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_DangLianHuangHua"
                                }
                                response = self.nifty_res.post_request("/presap/webintf.do?method=save_exception_ms",
                                                                       data=urlencode(data))
                                if response.json()["code"] == "200" and response.json()["msg"] == "数据保存成功.":
                                    logger.info("登记异常录入成功！")
                                else:
                                    logger.error(f"登记异常录入失败，返回结果：{response}")
                                    raise Exception(f"登记异常录入失败，返回结果：{response}")
                            else:
                                logger.error(f"异常查询失败，返回结果：{response}")
                                raise Exception(f"异常查询失败，返回结果：{response}")
                        else:
                            datas = {}
                            for i in range(len(rwd_detail_data)):
                                dic_data = {k: "" if v is None else v for k, v in rwd_detail_data[i].items()}
                                datas[str(i + 1).zfill(6)] = {"tab1": [{}], "tab2": {}}
                                datas[str(i + 1).zfill(6)]["tab1"][0] = dic_data
                                datas[str(i + 1).zfill(6)]["tab1"][0]["zdlhwkh"] = dic_data["pre_zcwbh"].replace(
                                    "P", "B")
                                datas[str(i + 1).zfill(6)]["tab1"][0]["zcwlx"] = "环化后混合文库"
                                datas[str(i + 1).zfill(6)]["tab1"][0]["zcwbh"] = dic_data["pre_zcwbh"].replace("P",
                                                                                                                    "B")
                                datas[str(i + 1).zfill(6)]["tab1"][0]["zms_sfhg"] = "合格"
                                datas[str(i + 1).zfill(6)]["tab1"][0]["z_pmol_number"] = 1
                                datas[str(i + 1).zfill(6)]["tab2"] = dic_data
                                datas[str(i + 1).zfill(6)]["tab2"]["zdlhwkh"] = dic_data["pre_zcwbh"].replace("P",
                                                                                                                   "B")
                                datas[str(i + 1).zfill(6)]["tab2"]["zcwlx"] = "环化后混合文库"
                                datas[str(i + 1).zfill(6)]["tab2"]["zcwbh"] = dic_data["pre_zcwbh"].replace("P",
                                                                                                                 "B")
                                datas[str(i + 1).zfill(6)]["tab2"]["zms_sfhg"] = "合格"
                                datas[str(i + 1).zfill(6)]["tab2"]["z_pmol_number"] = 1
                            data = {
                                "task": [datas],
                                "zgxbh": "MBL",
                                "zjob_code": self.zjob_code,
                                "token": self.token,
                                "menuId": "MSTaskProduct_DangLianHuangHua"
                            }
                            body_data = urlencode(data)
                            response = self.nifty_res.post_request("/presap/webintf.do?method=msUpdateTaskResult",data=body_data)
                            if response.json()["code"] == "200" and response.json()["msg"] == "success":
                                logger.info("单链环化任务单结果录入成功！")
                                data = {
                                    "task": {
                                        "zpzxlh_bcode": "",
                                        "zcjdat": "",
                                        "ybbhrq": "",
                                        "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                        "zsfwc_czr": "",
                                        "zsfks_czr": "",
                                        "zcdate": "",
                                        "chkdate": "",
                                        "matnr": rwd_data["matnr"],
                                        "zseqplatform": rwd_data["zseqplatform"],
                                        "zsfks_uzeit": "",
                                        "zdlhwkh": "",
                                        "qdfkrq": "",
                                        "zsampling_datum": "",
                                        "zdlvdate": """ """,
                                        "zgxbh": "MAX",
                                        "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                        "zsigndate": "",
                                        "zsfks": "",
                                        "zsyyn": rwd_data["zsyyn"],
                                        "syncdate": "",
                                        "ddl21": rwd_data["ddl21"],
                                        "yctjrq": "",
                                        "zcxlx": "0",
                                        "zjob_begin_datum": "",
                                        "zsfwc_uzeit": "",
                                        "zsample_quantity": rwd_data["zsample_quantity"],
                                        "zmethod": rwd_data["zmethod"],
                                        "zsfxd_czr": rwd_data["zsfxd_czr"],
                                        "zudate": "",
                                        "addate": "",
                                        "zycdjsflr": "否",
                                        "zjob_status": "未开始",
                                        "zsfxd_datum": rwd_data["zsfxd_datum"],
                                        "werks": "A020",
                                        "zsfks_datum": "",
                                        "ycwcrq": "",
                                        "pre_zsfwc_datum": "",
                                        "zsfwc": "",
                                        "pre_zjob_code": rwd_data["zjob_code"],
                                        "maktx": rwd_data["maktx"],
                                        "zsfwc_datum": "",
                                        "zupdat": "",
                                        "zjob_code": rwd_data["zjob_code"],
                                        "gt_action": "",
                                        "zcbdate": "",
                                        "gt_sendtime": "",
                                        "zfinish_time": "",
                                        "_key": 1,
                                        "_id": 1
                                    },
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_DangLianHuangHua"
                                }
                                response = self.nifty_res.post_request(
                                    "/presap/webintf.do?method=findJobSampleMessageMS",
                                    data=urlencode(data))
                                completion_info = response.json()["data"]
                                list_bak = []
                                for i in range(len(completion_info)):
                                    completion_info[i]["zlgfsflr"] = "X"
                                    if completion_info[i]["zpoint"] == "a-1" and "a-1" not in list_bak:
                                        completion_info[i]["_id"] = 1
                                        list_bak.append("a-1")
                                    elif completion_info[i]["zpoint"] == "b-1" and "b-1" not in list_bak:
                                        completion_info[i]["_id"] = 2
                                        list_bak.append("b-1")
                                    elif completion_info[i]["zpoint"] == "c-1" and "c-1" not in list_bak:
                                        completion_info[i]["_id"] = 3
                                        list_bak.append("c-1")
                                    elif completion_info[i]["zpoint"] == "d-1" and "d-1" not in list_bak:
                                        completion_info[i]["_id"] = 4
                                        list_bak.append("d-1")
                                data = {
                                    "req": completion_info,
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_DangLianHuangHua"
                                }
                                response = self.nifty_res.post_request(
                                    "/presap/webintf.do?method=task_sample_item_finishms",
                                    data=urlencode(data))
                                if response.json()["code"] == "200" and response.json()["msg"] == "SAP数据库更新成功!":
                                    logger.info("单链环化任务完成成功！")
                                    dlhhorder_scheme = list(set([j["zcwbh"] for j in completion_info]))
                                    dlhhorder_scheme.sort()
                                    self.dlhhorder_scheme = dlhhorder_scheme
                                    return dlhhorder_scheme
                                else:
                                    logger.error(f"单链环化任务完成失败，返回结果：{response}")
                                    raise Exception(f"单链环化任务完成失败，返回结果：{response}")
                            else:
                                logger.error(f"单链环化任务单结果录入失败，返回结果：{response}")
                                raise Exception(f"单链环化任务单结果录入失败，返回结果：{response}")
                    else:
                        logger.error(f"单链环化任务开始失败，返回结果：{response}")
                        raise Exception(f"单链环化任务开始失败，返回结果：{response}")
                else:
                    logger.error(f"单链环化任务单列表查询失败，返回结果：{response}")
                    raise Exception(f"单链环化任务单列表查询失败，返回结果：{response}")
            else:
                logger.error(f"单链环化下达任务失败，返回结果：{response}")
                raise Exception(f"单链环化下达任务失败，返回结果：{response}")
        else:
            logger.error(f"单链环化新建任务列表查询失败，返回结果：{response}")
            raise Exception(f"单链环化新建任务列表查询失败，返回结果：{response}")

    def makednb500order(self,dlhhorder_scheme=None):
        """
        造数工具MakeDNB
        :return:
        """
        if dlhhorder_scheme:
            dlhhorder_scheme = dlhhorder_scheme
        else:
            dlhhorder_scheme = self.dlhhorder_scheme
        dlhhorder_scheme = ','.join(dlhhorder_scheme)
        data = {
            "task": {
                "zdlhwkh": dlhhorder_scheme
            },
            "pageNumber": "1",
            "zgxbh": "MAK",
            "token": self.token,
            "menuId": "MSTaskMaster_MakeDNB500_MakeDNB500Order"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=task_assign_samplems", data=urlencode(data))
        if response.json()["code"] == "200":
            query_unassign_task_list = response.json()["data"]
            logger.info("MakeDNB新建任务列表查询成功！")
            data = {
                "token": self.token,
                "menud": "MSTaskMaster_MakeDNB500_MakeDNB500Order"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=get_zpzxlh",data=urlencode(data))
            zpzxlh_num = response.json()["data"][0]["zpzxlh"]
            detail = {
                "scmat": "常规",
                "ztx": self.config_info["indexConfig"],
                "z_lane_gs": 4,
                "zsyyn": "何华",
                "zmethod": "1755.04.42",
                "zmethod_des": "1755.04.42MGISEQ-2000(SE35)_MAKE DNB(NIFTY全流程)",
                "zshould_sj_datum": self.datatime_str,
                "zplate_x": "12",
                "zplate_y": "8"
            }
            datas = []
            for i in range(len(query_unassign_task_list)):
                data_dic = {k: "" if v is None else v for k, v in query_unassign_task_list[i].items()}
                data_dic.update(detail)
                data_dic["radat"] = data_dic["pre_radat"]
                data_dic["zindex"] = data_dic["pre_zindex"]
                data_dic["zindex_name"] = data_dic["pre_zindex_name"]
                data_dic["zpzxlh"] = zpzxlh_num
                if data_dic["zpoint"] == "a-1":
                    data_dic["z_lane_no"] = 1
                elif data_dic["zpoint"] == "b-1":
                    data_dic["z_lane_no"] = 2
                elif data_dic["zpoint"] == "c-1":
                    data_dic["z_lane_no"] = 3
                elif data_dic["zpoint"] == "d-1":
                    data_dic["z_lane_no"] = 4
                datas.append(data_dic)
            data = {
                "task": datas,
                "token": self.token,
                "menuId": "MSTaskMaster_MakeDNB500_MakeDNB500Order"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=saveTaskAssignSampleMS",
                                                   data=urlencode(data))
            if response.json()["code"] == "200":
                logger.info("MakeDNB下达任务成功！")
                self.zjob_code = response.json()["data"][0]["zjob_code"]
                # 获取makednb任务单号
                self.makednb_task_code = self.zjob_code
                data = {
                    "task": {
                        "zsfxd_datum": "",
                        "zsfxd_datumend": "",
                        "zsfks_datum": "",
                        "zsfks_datumend": "",
                        "zjob_code": self.zjob_code,
                        "zsfwc": "",
                        "task": "query"
                    },
                    "pageNumber": "1",
                    "zgxbh": "MAK",
                    "pageSize": "1000",
                    "token": self.token,
                    "menuId": "MSTaskProduct_MakeDNB"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=findJobMS", data=urlencode(data))
                if response.json()["code"] == "200":
                    logger.info("MakeDNB任务单列表查询成功！")
                    rwd_data = response.json()["data"][0]
                    data = {
                        "task": {
                            "zpzxlh_bcode": "",
                            "zcjdat": "",
                            "ybbhrq": "",
                            "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                            "zsfwc_czr": "",
                            "zsfks_czr": "",
                            "zcdate": "",
                            "chkdate": "",
                            "matnr": rwd_data["matnr"],
                            "zseqplatform": rwd_data["zseqplatform"],
                            "zsfks_uzeit": "",
                            "zdlhwkh": "",
                            "qdfkrq": "",
                            "zsampling_datum": "",
                            "zdlvdate": """ """,
                            "zgxbh": "MAK",
                            "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                            "zsigndate": "",
                            "zsfks": "",
                            "zsyyn": rwd_data["zsyyn"],
                            "syncdate": "",
                            "ddl21": rwd_data["ddl21"],
                            "yctjrq": "",
                            "zcxlx": "0",
                            "zjob_begin_datum": "",
                            "zsfwc_uzeit": "",
                            "zsample_quantity": rwd_data["zsample_quantity"],
                            "zmethod": rwd_data["zmethod"],
                            "zsfxd_czr": rwd_data["zsfxd_czr"],
                            "zudate": "",
                            "addate": "",
                            "zycdjsflr": "否",
                            "zjob_status": "未开始",
                            "zsfxd_datum": rwd_data["zsfxd_datum"],
                            "werks": "A020",
                            "zsfks_datum": "",
                            "ycwcrq": "",
                            "pre_zsfwc_datum": "",
                            "zsfwc": "",
                            "pre_zjob_code": rwd_data["zjob_code"],
                            "maktx": rwd_data["maktx"],
                            "zsfwc_datum": "",
                            "zupdat": "",
                            "zjob_code": rwd_data["zjob_code"],
                            "gt_action": "",
                            "zcbdate": "",
                            "gt_sendtime": "",
                            "zfinish_time": "",
                            "_key": 1,
                            "_id": 1
                        },
                        "token": self.token,
                        "menuId": "MSTaskProduct_MakeDNB"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS",
                                                           data=urlencode(data))
                    if response.json()["code"] == "200":
                        logger.info("MakeDNB任务单详情查询成功！")
                        rwd_detail_data = response.json()["data"]
                        data = {
                            "zgxbh": "MAK",
                            "zjob_code": self.zjob_code,
                            "token": self.token,
                            "menuId": "MSTaskProduct_MakeDNB"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=msStartTask",
                                                               data=urlencode(data))
                        if response.json()["code"] == "200" and response.json()["msg"] == "success":
                            logger.info("MakeDNB任务开始成功！")
                            data = {
                                "task": {
                                    "zpzxlh_bcode": "",
                                    "zcjdat": "",
                                    "ybbhrq": "",
                                    "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                    "zsfwc_czr": "",
                                    "zsfks_czr": "",
                                    "zcdate": "",
                                    "chkdate": "",
                                    "matnr": rwd_data["matnr"],
                                    "zseqplatform": rwd_data["zseqplatform"],
                                    "zsfks_uzeit": "",
                                    "zdlhwkh": "",
                                    "qdfkrq": "",
                                    "zsampling_datum": "",
                                    "zdlvdate": """ """,
                                    "zgxbh": "MAK",
                                    "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                    "zsigndate": "",
                                    "zsfks": "",
                                    "zsyyn": rwd_data["zsyyn"],
                                    "syncdate": "",
                                    "ddl21": rwd_data["ddl21"],
                                    "yctjrq": "",
                                    "zcxlx": "0",
                                    "zjob_begin_datum": "",
                                    "zsfwc_uzeit": "",
                                    "zsample_quantity": rwd_data["zsample_quantity"],
                                    "zmethod": rwd_data["zmethod"],
                                    "zsfxd_czr": rwd_data["zsfxd_czr"],
                                    "zudate": "",
                                    "addate": "",
                                    "zycdjsflr": "否",
                                    "zjob_status": "未开始",
                                    "zsfxd_datum": rwd_data["zsfxd_datum"],
                                    "werks": "A020",
                                    "zsfks_datum": "",
                                    "ycwcrq": "",
                                    "pre_zsfwc_datum": "",
                                    "zsfwc": "",
                                    "pre_zjob_code": rwd_data["zjob_code"],
                                    "maktx": rwd_data["maktx"],
                                    "zsfwc_datum": "",
                                    "zupdat": "",
                                    "zjob_code": rwd_data["zjob_code"],
                                    "gt_action": "",
                                    "zcbdate": "",
                                    "gt_sendtime": "",
                                    "zfinish_time": "",
                                    "_key": 1,
                                    "_id": 1
                                },
                                "token": self.token,
                                "menuId": "MSTaskProduct_MakeDNB"
                            }
                            response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS",
                                                                   data=urlencode(data))
                            rwd_detail_data = response.json()["data"]
                            if self.abnormal_input == '1':
                                data = {
                                    "zgxbh": "MAK",
                                    "werks": "A020",
                                    "arbpl": "",
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_MakeDNB"
                                }
                                response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                                       data=urlencode(data))
                                if response.json()["code"] == "200" and len(response.json()["data"]) > 0:
                                    logger.info("MakeDNB异常查询成功！")
                                    exception_data = response.json()["data"][0]
                                    task = []
                                    for i in range(len(rwd_detail_data)):
                                        data_dic = {k: "" if v is None else v for k, v in rwd_detail_data[i].items()}
                                        exception_data_dic = {k: "" if v is None else v for k, v in
                                                              exception_data.items()}
                                        data_dic.update(exception_data_dic)
                                        data_dic["_id"] = i + 1
                                        data_dic["zcjnam"] = "auto_test"
                                        data_dic["zcjdat"] = self.date_str
                                        data_dic["zcjuzt"] = self.time_str
                                        data_dic["dateTime"] = self.datatime_str
                                        data_dic["yc_status"] = "A"
                                        task.append(data_dic)
                                    data = {
                                        "task": task,
                                        "werks": "A020",
                                        "arbpl": "",
                                        "zjob_code": rwd_data["zjob_code"],
                                        "zgxbh": "MAK",
                                        "token": self.token,
                                        "menuId": "MSTaskProduct_MakeDNB"
                                    }
                                    response = self.nifty_res.post_request(
                                        "/presap/webintf.do?method=save_exception_ms",
                                        data=urlencode(data))
                                    if response.json()["code"] == "200" and response.json()["msg"] == "数据保存成功.":
                                        logger.info("MakeDNB登记异常录入成功！")
                                    else:
                                        logger.error(f"MakeDNB登记异常录入失败，返回结果：{response}")
                                        raise Exception(f"MakeDNB登记异常录入失败，返回结果：{response}")
                                else:
                                    logger.error(f"MakeDNB异常查询失败，返回结果：{response}")
                                    raise Exception(f"MakeDNB异常查询失败，返回结果：{response}")
                            else:
                                data = {
                                    "task": {
                                        "zjob_code": rwd_data["zjob_code"],
                                        "zgxbh": "MAK",
                                        "zsfzjcw": "ALL",
                                        "zsfwc": ""
                                    },
                                    "token": self.token,
                                    "menuId": "MSTaskProduct_MakeDNB"
                                }
                                response = self.nifty_res.post_request("/presap/webintf.do?method=find_chanWuMS",
                                                                       data=urlencode(data))
                                if response.json()["code"] == "200":
                                    logger.info("MakeDNB产物查询成功！")
                                    chanwu_data = response.json()["data"]["Samples"]
                                    chanwu_data.sort(key=lambda x: x["zjob_code_item"])
                                    info = {
                                        "zms_sfhg": "合格",
                                        "bvalue": "4950.9",
                                        "kvalue": "193900",
                                        "xsbs": "53",
                                        "zdnbrd": 25.98,
                                        "zxgz": "99999"
                                    }
                                    data0 = {}
                                    for i in range(len(chanwu_data)):
                                        data_dic = {k: "" if v is None else v for k, v in chanwu_data[i].items()}
                                        data0[str(i + 1).zfill(6)] = {"tab1": [], "tab2": {}}
                                        data0[str(i + 1).zfill(6)]["tab2"] = data_dic
                                        task_info_copy = copy.deepcopy(info)
                                        data0[str(i + 1).zfill(6)]["tab2"].update(task_info_copy)
                                        if data_dic["zpoint"] == "a-1":
                                            data0[str(i + 1).zfill(6)]["tab2"]["pz_lane"] = str(
                                                data_dic["zpzxlh"]) + "-01"
                                        elif data_dic["zpoint"] == "b-1":
                                            data0[str(i + 1).zfill(6)]["tab2"]["pz_lane"] = str(
                                                data_dic["zpzxlh"]) + "-02"
                                        elif data_dic["zpoint"] == "c-1":
                                            data0[str(i + 1).zfill(6)]["tab2"]["pz_lane"] = str(
                                                data_dic["zpzxlh"]) + "-03"
                                        elif data_dic["zpoint"] == "d-1":
                                            data0[str(i + 1).zfill(6)]["tab2"]["pz_lane"] = str(
                                                data_dic["zpzxlh"]) + "-04"
                                    data = {
                                        "task": [data0],
                                        "zgxbh": "MAK",
                                        "zjob_code": self.zjob_code,
                                        "token": self.token,
                                        "menuId": "MSTaskProduct_MakeDNB"
                                    }
                                    response = self.nifty_res.post_request(
                                        "/presap/webintf.do?method=msUpdateTaskResult",
                                        data=urlencode(data))
                                    if response.json()["msg"] == "success" and response.json()["code"] == "200":
                                        logger.info("MakeDNB任务单结果录入成功！")
                                        data = {
                                            "task": {
                                                "zpzxlh_bcode": "",
                                                "zcjdat": "",
                                                "ybbhrq": "",
                                                "zsfxd_uzeit": rwd_data["zsfxd_uzeit"],
                                                "zsfwc_czr": "",
                                                "zsfks_czr": "",
                                                "zcdate": "",
                                                "chkdate": "",
                                                "matnr": rwd_data["matnr"],
                                                "zseqplatform": rwd_data["zseqplatform"],
                                                "zsfks_uzeit": "",
                                                "zdlhwkh": "",
                                                "qdfkrq": "",
                                                "zsampling_datum": "",
                                                "zdlvdate": """ """,
                                                "zgxbh": "MAK",
                                                "zshould_finish_datum": rwd_data["zshould_finish_datum"],
                                                "zsigndate": "",
                                                "zsfks": "",
                                                "zsyyn": rwd_data["zsyyn"],
                                                "syncdate": "",
                                                "ddl21": rwd_data["ddl21"],
                                                "yctjrq": "",
                                                "zcxlx": "0",
                                                "zjob_begin_datum": "",
                                                "zsfwc_uzeit": "",
                                                "zsample_quantity": rwd_data["zsample_quantity"],
                                                "zmethod": rwd_data["zmethod"],
                                                "zsfxd_czr": rwd_data["zsfxd_czr"],
                                                "zudate": "",
                                                "addate": "",
                                                "zycdjsflr": "否",
                                                "zjob_status": "未开始",
                                                "zsfxd_datum": rwd_data["zsfxd_datum"],
                                                "werks": "A020",
                                                "zsfks_datum": "",
                                                "ycwcrq": "",
                                                "pre_zsfwc_datum": "",
                                                "zsfwc": "",
                                                "pre_zjob_code": rwd_data["zjob_code"],
                                                "maktx": rwd_data["maktx"],
                                                "zsfwc_datum": "",
                                                "zupdat": "",
                                                "zjob_code": rwd_data["zjob_code"],
                                                "gt_action": "",
                                                "zcbdate": "",
                                                "gt_sendtime": "",
                                                "zfinish_time": "",
                                                "_key": 1,
                                                "_id": 1
                                            },
                                            "token": self.token,
                                            "menuId": "MSTaskProduct_MakeDNB"
                                        }
                                        response = self.nifty_res.post_request(
                                            "/presap/webintf.do?method=findJobSampleMessageMS",
                                            data=urlencode(data))
                                        completion_info = response.json()["data"]
                                        list_bak = []
                                        for i in range(len(completion_info)):
                                            completion_info[i]["zlgfsflr"] = "X"
                                            if completion_info[i]["zpoint"] == "a-1":
                                                completion_info[i]["pz_lane"] = str(
                                                    completion_info[i]["zpzxlh"]) + "-01"
                                            elif completion_info[i]["zpoint"] == "b-1":
                                                completion_info[i]["pz_lane"] = str(
                                                    completion_info[i]["zpzxlh"]) + "-02"
                                            elif completion_info[i]["zpoint"] == "c-1":
                                                completion_info[i]["pz_lane"] = str(
                                                    completion_info[i]["zpzxlh"]) + "-03"
                                            elif completion_info[i]["zpoint"] == "d-1":
                                                completion_info[i]["pz_lane"] = str(
                                                    completion_info[i]["zpzxlh"]) + "-04"
                                            if completion_info[i]["zpoint"] == "a-1" and "a-1" not in list_bak:
                                                completion_info[i]["_id"] = 1
                                                list_bak.append("a-1")
                                            elif completion_info[i]["zpoint"] == "b-1" and "b-1" not in list_bak:
                                                completion_info[i]["_id"] = 2
                                                list_bak.append("b-1")
                                            elif completion_info[i]["zpoint"] == "c-1" and "c-1" not in list_bak:
                                                completion_info[i]["_id"] = 3
                                                list_bak.append("c-1")
                                            elif completion_info[i]["zpoint"] == "d-1" and "d-1" not in list_bak:
                                                completion_info[i]["_id"] = 4
                                                list_bak.append("d-1")
                                        data = {
                                            "req": completion_info,
                                            "token": self.token,
                                            "menuId": "MSTaskProduct_MakeDNB"
                                        }
                                        complete_response = self.nifty_res.post_request(
                                            "/presap/webintf.do?method=task_sample_item_finishms",
                                            data=urlencode(data))
                                        if complete_response.json()["code"] == "200" and \
                                                complete_response.json()[
                                                    "msg"] == "SAP数据库更新成功!":
                                            logger.info("MakeDNB任务完成成功！")
                                        else:
                                            logger.error(f"MakeDNB任务完成失败，返回结果：{response}")
                                            raise Exception(f"MakeDNB任务完成失败，返回结果：{response}")
                                    else:
                                        logger.error(f"MakeDNB任务单结果录入失败，返回结果：{response}")
                                        raise Exception(f"MakeDNB任务单结果录入失败，返回结果：{response}")
                                else:
                                    logger.error(f"MakeDNB产物查询失败，返回结果：{response}")
                                    raise Exception(f"MakeDNB产物查询失败，返回结果：{response}")
                        else:
                            logger.error(f"MakeDNB任务开始失败，返回结果：{response}")
                            raise Exception(f"MakeDNB任务开始失败，返回结果：{response}")
                    else:
                        logger.error(f"MakeDNB任务单详情查询失败，返回结果：{response}")
                        raise Exception(f"MakeDNB任务单详情查询失败，返回结果：{response}")
                else:
                    logger.error(f"MakeDNB任务单列表查询失败，返回结果：{response}")
                    raise Exception(f"MakeDNB任务单列表查询失败，返回结果：{response}")
            else:
                logger.error(f"MakeDNB下达任务失败，返回结果：{response}")
                raise Exception(f"MakeDNB下达任务失败，返回结果：{response}")

        else:
            logger.error(f"MakeDNB新建任务列表查询失败，返回结果：{response}")
            raise Exception(f"MakeDNB新建任务列表查询失败，返回结果：{response}")

    def data_review(self,sample=None):
        """
        造数工具数据审核
        :return:
        """
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        data = {
            "task": {
                "zsample": sampleid,
                "zsfwc": ""
            },
            "pageNumber": "1",
            "zgxbh": "MCA",
            "pageSize": "1500",
            "token": self.token,
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=query_data_audit",data=urlencode(data))
        if response.json()["code"]=="200" and len(response.json()["data"])>0:
            logger.info("数据审核待处理样本列表查询成功")
            data_list = response.json()["data"]
            if self.abnormal_input == '1':
                data = {
                    "zgxbh": "MCA",
                    "werks": "A020",
                    "arbpl": "",
                    "token": self.token,
                    "menuId": "MSTaskProduct_DataAudit"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                       data=urlencode(data))
                if response.json()["code"] == "200" and len(response.json()["data"]) > 0:
                    logger.info("数据审核异常查询成功！")
                    exception_data = response.json()["data"][0]
                    task = []
                    for i in range(len(data_list)):
                        data_dic = {k: "" if v is None else v for k, v in data_list[i].items()}
                        exception_data_dic = {k: "" if v is None else v for k, v in
                                              exception_data.items()}
                        data_dic.update(exception_data_dic)
                        data_dic["_id"] = i + 1
                        data_dic["zcjnam"] = "auto_test"
                        data_dic["zcjdat"] = self.date_str
                        data_dic["zcjuzt"] = self.time_str
                        data_dic["dateTime"] = self.datatime_str
                        data_dic["yc_status"] = "A"
                        task.append(data_dic)
                    data = {
                        "task": task,
                        "werks": "A020",
                        "arbpl": "",
                        "zjob_code": data_list[0]["zjob_code"],
                        "zgxbh": "MCA",
                        "token": self.token,
                        "menuId": "MSTaskProduct_DataAudit"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=save_exception_ms",
                                                           data=urlencode(data))
                    if response.json()["code"] == "200" and response.json()["msg"] == "数据保存成功.":
                        logger.info("数据审核登记异常录入成功！")
                    else:
                        logger.error(f"数据审核登记异常录入失败，返回结果：{response}")
                        raise Exception(f"数据审核登记异常录入失败，返回结果：{response}")
                else:
                    logger.error(f"数据审核异常查询失败，返回结果：{response}")
                    raise Exception(f"数据审核异常查询失败，返回结果：{response}")
            else:
                task = [
                    {
                        "zguid": data_list[i]["zguid"],
                        "zeile": data_list[i]["zeile"],
                        "zbglx": "REGULAR",
                        "zbgijgdh_val": data_list[i]["jg_zbgijgdh_val"],
                        "zjob_code": data_list[i]["zjob_code"],
                        "zjob_code_item": data_list[i]["zjob_code_item"],
                        "note3": "",
                        "zcatalo": data_list[i]["zcatalo"],
                        "zsample": data_list[i]["zsample"],
                        "zsjdid": data_list[i]["zsjdid"]
                    }
                    for i in range(len(data_list))
                ]
                data = {
                    "task": task,
                    "token": self.token,
                    "menuId": "MSTaskProduct_DataAudit"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=save_audit_sample_datas",
                                                       data=urlencode(data))
                if response.json()["code"] == "200" and response.json()["msg"] == "success":
                    logger.info("数据审核结束成功！")
                else:
                    logger.error(f"数据审核结束失败，返回结果：{response}")
                    raise Exception(f"数据审核结束失败，返回结果：{response}")
        else:
            logger.error(f"数据审核待处理样本列表查询失败，返回结果：{response}")
            raise Exception(f"数据审核待处理样本列表查询失败，返回结果：{response}")

    @staticmethod
    def query_sample_info(sequencing_task_code):
        """根据上机任务单号查询梧桐测试环境数据库中样本对应样本信息和实验信息"""
        sql = "select lps.lims_id,lps.lane_id,lps.bar_code,lps.pooling_id,lps.sub_no,lps.product_code,ls.sample_no,ls.sample_id, " \
               "lsi.seq_no, lsi.seq_start_time, lsi.seq_platform, lp.pooling_no, lsl.slide_no from lab_pooling_sub lps " \
                "join lab_sample ls on lps.sample_id=ls.sample_id join lab_seq_info lsi on lps.seq_id=lsi.seq_id join lab_pooling lp on lps.pooling_id=lp.pooling_id " \
        "join lab_slide lsl on lps.seq_id = lsl.seq_id where lsi.seq_no ='" + sequencing_task_code + "' ORDER BY lps.lims_id ASC "

        sycamore_db_config = GetConfig(configname="sycamore_config.yaml").get_mysql_config()
        handler_db = HandleDB(sycamore_db_config['host'], sycamore_db_config['port'], sycamore_db_config['user'],
                              sycamore_db_config['password'], sycamore_db_config['database'])
        while True:
            query_result = list(handler_db.select(sql))
            if query_result:
                break
            time.sleep(30)
        # 过滤掉阴阳质控品
        sample_info = []
        for i in query_result:
            if "YAX" in i["sample_no"] or "YIX" in i["sample_no"]:
                continue
            else:
                sample_info.append(i)
        logger.debug(sample_info)
        handler_db.close_database()
        return sample_info

    def information_analysis(self, sequencing_task_code=None):
        """
        信息分析
        @return:
        """
        if sequencing_task_code:
            self.sequencing_task_code = sequencing_task_code
        logger.debug("当前流程阶段为：信息分析")
        # 梧桐推送分析结果
        data_template = {
            "chrList": [
                {
                    "chr": "7",
                    "chrTest": "T",
                    "filterFlag": "0",
                    "fra": "1335",
                    "risk": "5.336367e-10",
                    "t": "178400",
                    "zScore": "81500"
                },
                {
                    "chr": "8",
                    "chrTest": "S",
                    "filterFlag": "0",
                    "fra": "1335",
                    "risk": "5.336367e-10",
                    "t": "178400",
                    "zScore": "81500"
                },
                {
                    "chr": "13",
                    "chrTest": "Negative",
                    "filterFlag": "0",
                    "fra": "327",
                    "risk": "5.333773e-10",
                    "t": "42500",
                    "zScore": "11700"
                },
                {
                    "chr": "18",
                    "chrTest": "SH",
                    "filterFlag": "0",
                    "fra": "1335",
                    "risk": "5.336367e-10",
                    "t": "178400",
                    "zScore": "81500"
                },
                {
                    "chr": "21",
                    "chrTest": "Negative",
                    "filterFlag": "0",
                    "fra": "-2100",
                    "risk": "2.786081e-10",
                    "t": "-210400",
                    "zScore": "-126600"
                },
                {
                    "chr": "23",
                    "chrTest": "Negative",
                    "filterFlag": "1",
                    "fra": "22808",
                    "risk": "3.885314e-10",
                    "t": "-35200",
                    "zScore": "-35200"
                }
            ],
            "cnvList": [],
            "disease": "",
            "dupDel": "",
            "fra": "0.22613",
            "gc": "0.40352",
            "gender": "Male",
            "idaZ13": "0.117",
            "idaZ18": "0.815",
            "idaZ21": "-1.266",
            "intransList": [],
            "jobCode": "TSK24000000567",
            "jobCodeItem": "000161",
            "ldaZ13": "0.117",
            "ldaZ18": "0.815",
            "ldaZ21": "-1.266",
            "limsId": "000161",
            "note3": "",
            "position": "",
            "qc": "通过",
            "qualified": "是",
            "recommend": "",
            "recommended": "",
            "riskIndex13": "1/1874845346",
            "riskIndex18": "1/1873934059",
            "riskIndex21": "1/3589270615",
            "sampleId": "24D05280561-1_48_TSK24000000567_L04_MGISEQ-2000PLA0202400126_24L05280561-1-48_20240527_V350231139_SZ_MGISEQ-2000_DX0558",
            "sendStatus": "发送中",
            "size": "",
            "supplement": "",
            "tScore13": "0.425",
            "tScore18": "1.784",
            "tScore21": "-2.104",
            "test13": "未检出T13",
            "test18": "未检出T18",
            "test21": "未检出T21",
            "testAuto": "",
            "testPos": "",
            "testSex": "未检出异常",
            "testZone": "",
            "updateTime": 0,
            "ur": "5495295",
            "yPer": "0.02717",
            "zScore13": "0.117",
            "zScore18": "0.815",
            "zScore21": "-1.266"
        }
        # 查询上机测序任务单的样本信息和实验信息
        sample_info = self.query_sample_info(self.sequencing_task_code)
        # 获取导入的excl中的所有样本的cnv信息
        cnv_result = handle_cnv_list(self.qc_info)

        for sample, cnv_list, qc, item in zip(sample_info, cnv_result, self.qc_info, self.chr_info):
            copy_data = deepcopy(data_template)
            copy_data["jobCode"] = self.sequencing_task_code
            copy_data["jobCodeItem"] = sample["lims_id"]
            copy_data["limsId"] = sample["lims_id"]
            copy_data["sample_no"] = sample["sample_no"]
            copy_data["cnvList"] = cnv_list
            new_chr = []
            found = False
            for chr_item in copy_data["chrList"]:
                if chr_item["chr"] == item["chr"]:
                    new_chr.append(item)
                    found = True
                else:
                    new_chr.append(chr_item)
            if not found:
                new_chr.append(item)
            copy_data["chrList"] = new_chr
            key_mapping = {
                "cnv_band": "testZone",
                "cnv": "testPos",
                "test13": "test13",
                "test18": "test18",
                "test21": "test21",
                "test_auto": "testAuto",
                "test_sex": "testSex",
                "note3": "note3",
                "note2": "supplement",
                "disease": "disease",
                "report_tag": "report_tag",
                "product_no": "product_no",
                "fetus_type": "fetus_type"
            }
            for qc_key, copy_key in key_mapping.items():
                if qc.get(qc_key):
                    copy_data[copy_key] = qc[qc_key]

            # 组装sampleId
            start_time = time.strftime("%Y%m%d", time.localtime(sample["seq_start_time"] / 1000))
            sample_num = "-".join(sample["sub_no"].replace("L", "D").split("-")[:-1]) + "_" + str(sample["bar_code"])
            copy_data["sampleId"] = sample_num + "_" + sample["seq_no"] + "_L0" + str(sample["lane_id"]) + "_" \
                                           + sample["pooling_no"] + "_" + sample["sub_no"] + "_" + start_time + "_" \
                                           + sample["slide_no"] + "_SZ_" + sample["seq_platform"] + "_" + sample["product_code"]
            self.result_data.append(copy_data)
        logger.info("信息分析:开始推送分析结果到nifty！")
        logger.info(f"请求data：{str(self.result_data)}")
        send_result_data = {"bean": json.dumps(self.result_data, ensure_ascii=False)}
        count = 0
        while True:
            count += 1
            send_result_respone = self.nifty_api_res.post_request("/fritillary-nifty/saveNiftyResult", data=urlencode(send_result_data)).json()
            # {"retCode": 1, "success": false, "retInfo": "找不到信息分析任务单,无法完成本工序任务!", "result": ""}
            if send_result_respone["retCode"] != 1 and send_result_respone["retInfo"] != "找不到信息分析任务单,无法完成本工序任务!":
                break
            if count >= 10:
                break
            time.sleep(60)
        if send_result_respone["retCode"] == 0 and send_result_respone["success"] is True:
            logger.info("信息分析:推送分析结果到nifty成功！梧桐分析平台任务单号：{}", self.sequencing_task_code)
            logger.debug("send_result_respone:{}", send_result_respone)
            return self.result_data
        else:
            logger.error("信息分析:推送分析结果到nifty失败！")
            logger.debug("send_result_respone:{}", send_result_respone)
            raise Exception("信息分析:推送分析结果到nifty失败！")

    def generate_report(self,sample=None):
        """
        造数工具生成产前报告
        :return:
        """
        time.sleep(15)
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        count = 0
        while True:
            time.sleep(2)
            count += 1
            data = {
                "task": {
                    "zsample": sampleid,
                    "zsfxd": "",
                    "username": ""
                },
                "pageNumber": "1",
                "pageSize": "500",
                "token": self.token,
                "menuId": "ReportCenter_NiftyManage_querylist"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyResult", data=urlencode(data))
            if response.json()["code"] == "200":
                if response.json()["msg"] == "success":
                    break
                if "样例不在搜索范围" in response.json()["msg"]:
                    msg_text = response.json()["msg"]
                    year = datetime.now().strftime('%y')
                    pattern = year+r'B\d+'
                    result = re.findall(pattern, msg_text)
                    sample_list = list(filter(lambda x: x not in result, sampleid.split(',')))
                    sampleid = ','.join(sample_list)
            if count >= 4:
                break
        response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyResult",data=urlencode(data))
        if response.json()["code"]=="200" and response.json()["msg"] == "success":
            logger.info("生成产前报告未完成列表查询成功")
            data_list = response.json()["data"]
            if self.abnormal_input == '1':
                data = {
                   "werks": "A020",
                   "zgxbh": "RT",
                   "token": self.token,
                   "menuId": "ReportCenter_NiftyManage_querylist"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                      data=urlencode(data))
                if response.json()["code"] == "200" and len(response.json()["data"]) > 0:
                    logger.info("报告异常查询成功！")
                    exception_data = response.json()["data"][0]
                    exce_data = {
                        "_exception": [
                            {
                                "zreason": exception_data["zreason"],
                                "name_xl": exception_data["name_xl"],
                                "yc_level": "",
                                "yc_dl": exception_data["yc_dl"],
                                "yc_ms": exception_data["yc_ms"],
                                "name_dl": exception_data["name_dl"],
                                "yc_xl": exception_data["yc_xl"],
                                "_id": 1,
                                "zcjnam": "auto-test",
                                "zcjdat": self.date_str,
                                "zcjuzt": self.time_str
                            }
                        ]
                   }
                    for j in range(0,len(data_list),25):
                        task = []
                        for i in range(len(data_list[j:j+25])):
                            data_dic = {k: "" if v is None else v for k, v in data_list[j:j+25][i].items()}
                            data_dic.update(exce_data)
                            data_dic["_key"] = i + 1
                            data_dic["_id"] = i + 1
                            data_dic["zstatus"] = "EXC"
                            data_dic["zreportid"] = "DUMMY"
                            data_dic["background_job"] = ""
                            data_dic["username"] = "auto-test"
                            data_dic["zstpch"] = data_dic["zjob_code"]
                            task.append(data_dic)
                        data = {
                           "datas": task,
                           "token": self.token,
                           "menuId": "ReportCenter_NiftyManage_querylist"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=saveNiftyResultExceptionFlag",
                                                              data=urlencode(data))
                        if response.json()["code"] == "200" and response.json()["msg"] == "success":
                           logger.info("报告异常登记成功！")
                        else:
                           logger.error(f"报告异常登记失败，返回结果：{response}")
                           raise Exception(f"报告异常登记失败，返回结果：{response}")
                else:
                    logger.error(f"报告异常查询失败，返回结果：{response}")
                    raise Exception(f"报告异常查询失败，返回结果：{response}")
            else:
                if all("reason" not in d.keys() for d in data_list):
                    for j in range(0, len(data_list), 25):
                        datas = []
                        for i in range(len(data_list[j:j+25])):
                            data_dic = {k: "" if v is None else v for k, v in data_list[j:j+25][i].items()}
                            data_dic["_key"] = i + 1
                            data_dic["_id"] = i + 1
                            data_dic["zstatus"] = "GEN"
                            data_dic["zstpch"] = data_dic["zjob_code"]
                            data_dic["background_job"] = ""
                            data_dic["username"] = "auto-test"
                            datas.append(data_dic)
                        data = {
                            "datas": datas,
                            "token": self.token,
                            "menuId": "ReportCenter_NiftyManage_querylist"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=niftyResultToReport",
                                                               data=urlencode(data))
                        if response.json()["code"] == "200" and response.json()["msg"] == "操作成功~!":
                            logger.info("生成报告成功！")
                        else:
                            logger.error(f"生成报告失败，返回结果：{response}")
                            raise Exception(f"生成报告失败，返回结果：{response}")
                else:
                    logger.error("缺失报告模板！")
        else:
            logger.error(f"生成产前报告未完成列表查询失败，返回结果：{response}")
            raise Exception(f"生成产前报告未完成列表查询失败，返回结果：{response}")

    def report_confirmation(self,sample=None):
        """
        造数工具-产前报告确认
        :return:
        """
        time.sleep(30)
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        count = 0
        while True:
            time.sleep(10)
            count += 1
            data = {
                "task": {
                    "zsample": sampleid,
                    "zsfywc": "",
                    "step": "SIGN1",
                    "username": "c-xufeng"
                },
                "pageNumber": "1",
                "pageSize": "500",
                "token": self.token,
                "menuId": "ReportCenter_NiftyManage_querylist4confirm"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyReportInfo",
                                                   data=urlencode(data))
            if response.json()["code"] == "200":
                if len(response.json()["data"]) == len(sampleid.split(',')):
                    break
                # if "样例不在搜索范围" in response.json()["msg"]:
                #     msg_text = response.json()["msg"]
                #     year = datetime.now().strftime('%y')
                #     pattern = year+r'B\d+'
                #     result = re.findall(pattern, msg_text)
                #     sample_list = list(filter(lambda x: x not in result, sampleid.split(',')))
                #     sampleid = ','.join(sample_list)
            if count >= 3:
                break
        if response.json()["code"] == "200" and len(response.json()["data"]) == len(sampleid.split(',')):
            logger.info("产前报告确认未完成列表查询成功")
            data_list = response.json()["data"]
            if self.abnormal_input == '1':
                data = {
                    "werks": "A020",
                    "zgxbh": "RT",
                    "token": self.token,
                    "menuId": "ReportCenter_NiftyManage_querylist4confirm"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                       data=urlencode(data))
                if response.json()["code"] == "200" and len(response.json()["data"]) > 0:
                    logger.info("产前报告确认异常查询成功")
                    exception_data = response.json()["data"][0]
                    datas = []
                    exce_data = {
                        "_exception": [
                            {
                                "zreason": exception_data["zreason"],
                                "name_xl": exception_data["name_xl"],
                                "yc_level": "",
                                "yc_dl": exception_data["yc_dl"],
                                "yc_ms": exception_data["yc_ms"],
                                "name_dl": exception_data["name_dl"],
                                "yc_xl": exception_data["yc_xl"],
                                "_id": 1,
                                "zcjnam": "auto-test",
                                "zcjdat": self.date_str,
                                "zcjuzt": self.time_str
                            }
                        ]
                    }
                    for j in range(0, len(data_list), 25):
                        for i in range(len(data_list[j:j+25])):
                            data_dic = {k: "" if v is None else v for k, v in data_list[j:j+25][i].items()}
                            data_dic.update(exce_data)
                            data_dic["_key"] = i + 1
                            data_dic["_id"] = i + 1
                            data_dic["zstatus"] = "EXC"
                            datas.append(data_dic)
                        data = {
                            "datas": datas,
                            "token": self.token,
                            "menuId": "ReportCenter_NiftyManage_querylist4confirm"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=saveNiftyResultExceptionFlag",
                                                               data=urlencode(data))
                        if response.json()["code"] == "200" and response.json()["msg"] == "success":
                            logger.info("产前报告确认异常登记成功！")
                        else:
                            logger.error(f"产前报告确认异常登记失败，返回结果：{response}")
                            raise Exception(f"产前报告确认异常登记失败，返回结果：{response}")
                    else:
                        logger.error(f"产前报告确认异常查询失败，返回结果：{response}")
                        raise Exception(f"产前报告确认异常查询失败，返回结果：{response}")
            else:
                for j in range(0,len(data_list),50):
                    datas = []
                    for i in range(len(data_list[j:j+50])):
                        data_dic = {k: "" if v is None else v for k, v in data_list[j:j+50][i].items()}
                        data_dic["_key"] = i + 1
                        data_dic["_id"] = i + 1
                        data_dic["zstatus"] = "CONFIRM"
                        data_dic["username"] = "auto-test"
                        datas.append(data_dic)
                    data = {
                        "datas": datas,
                        "token": self.token,
                        "menuId": "ReportCenter_NiftyManage_querylist4confirm"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=updateNiftyReportInfo",
                                                           data=urlencode(data))
                    if response.json()["code"] == "200" and response.json()["msg"] == "success":
                        logger.info("产前报告确认成功！")
                    else:
                        logger.error(f"产前报告确认失败，返回结果：{response}")
                        raise Exception(f"产前报告确认失败，返回结果：{response}")
        else:
            logger.error(f"产前报告确认未完成列表查询失败，返回结果：{response}")
            raise Exception(f"产前报告确认未完成列表查询失败，返回结果：{response}")

    def report_review(self,sample=None):
        """
        造数工具-产前报告审核
        :return:
        """
        time.sleep(15)
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        count = 0
        while True:
            count += 1
            time.sleep(5)
            data = {
                "task": {
                    "zsample": sampleid,
                    "zsfywc": "",
                    "step": "SIGN2",
                    "username": "c-xufeng"
                },
                "pageNumber": "1",
                "pageSize": "500",
                "token": self.token,
                "menuId": "ReportCenter_NiftyManage_querylist4check"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyReportInfo",
                                                   data=urlencode(data))
            if response.json()["code"] == "200":
                if len(response.json()["data"]) == len(sampleid.split(',')):
                    break
                # if "样例不在搜索范围" in response.json()["msg"]:
                #     msg_text = response.json()["msg"]
                #     year = datetime.now().strftime('%y')
                #     pattern = year+r'B\d+'
                #     result = re.findall(pattern, msg_text)
                #     sample_list = list(filter(lambda x: x not in result, sampleid.split(',')))
                #     sampleid = ','.join(sample_list)
            if count >= 3:
                break
        response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyReportInfo",data=urlencode(data))
        if response.json()["code"]=="200" and len(response.json()["data"]) == len(sampleid.split(',')):
            logger.info("产前报告审核未完成列表查询成功")
            data_list = response.json()["data"]
            if self.abnormal_input == '1':
                data = {
                    "werks": "A020",
                    "zgxbh": "RT",
                    "token": self.token,
                    "menuId": "ReportCenter_NiftyManage_querylist4check"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                       data=urlencode(data))
                if response.json()["code"] == "200" and len(response.json()["data"]) > 0:
                    logger.info("产前报告审核异常查询成功")
                    exception_data = response.json()["data"][0]
                    datas = []
                    exce_data = {
                        "_exception": [
                            {
                                "zreason": exception_data["zreason"],
                                "name_xl": exception_data["name_xl"],
                                "yc_level": "",
                                "yc_dl": exception_data["yc_dl"],
                                "yc_ms": exception_data["yc_ms"],
                                "name_dl": exception_data["name_dl"],
                                "yc_xl": exception_data["yc_xl"],
                                "_id": 1,
                                "zcjnam": "auto-test",
                                "zcjdat": self.date_str,
                                "zcjuzt": self.time_str
                            }
                        ]
                    }
                    for j in range(0, len(data_list), 25):
                        for i in range(len(data_list[j:j+25])):
                            data_dic = {k: "" if v is None else v for k, v in data_list[j:j+25][i].items()}
                            data_dic.update(exce_data)
                            data_dic["_key"] = i + 1
                            data_dic["_id"] = i + 1
                            data_dic["zstatus"] = "EXC"
                            datas.append(data_dic)
                        data = {
                            "datas": datas,
                            "token": self.token,
                            "menuId": "ReportCenter_NiftyManage_querylist4check"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=saveNiftyResultExceptionFlag",
                                                               data=urlencode(data))
                        if response.json()["code"] == "200" and response.json()["msg"] == "success":
                            logger.info("产前报告审核异常登记成功！")
                        else:
                            logger.error(f"产前报告审核异常登记失败，返回结果：{response}")
                            raise Exception(f"产前报告审核异常登记失败，返回结果：{response}")
                    else:
                        logger.error(f"产前报告审核未完成列表查询失败，返回结果：{response}")
                        raise Exception(f"产前报告审核未完成列表查询失败，返回结果：{response}")
            else:
                for j in range(0, len(data_list), 50):
                    datas = []
                    for i in range(len(data_list[j:j+50])):
                        data_dic = {k: "" if v is None else v for k, v in data_list[j:j+50][i].items()}
                        data_dic["_key"] = i + 1
                        data_dic["_id"] = i + 1
                        data_dic["zstatus"] = "CHECK"
                        data_dic["username"] = "auto-test"
                        datas.append(data_dic)
                    data = {
                        "datas": datas,
                        "token": self.token,
                        "menuId": "ReportCenter_NiftyManage_querylist4check"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=updateNiftyReportInfo",
                                                           data=urlencode(data))
                    if response.json()["code"] == "200" and response.json()["msg"] == "success":
                        logger.info("产前报告审核成功！")
                    else:
                        logger.error(f"产前报告审核失败，返回结果：{response}")
                        raise Exception(f"产前报告审核失败，返回结果：{response}")
        else:
            logger.error(f"产前报告审核未完成列表查询失败，返回结果：{response}")
            raise Exception(f"产前报告审核未完成列表查询失败，返回结果：{response}")

    def report_claim(self,sample=None):
        """
        造数工具-产前报告认领
        :return:
        """
        time.sleep(15)
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        count = 0
        while True:
            count += 1
            time.sleep(5)
            data = {
                "task": {
                    "zsjdid": "",
                    "zsample": sampleid,
                    "zsfywc": "",
                    "step": "SIGN3",
                    "username": "c-xufeng"
                },
                "pageNumber": "1",
                "pageSize": "500",
                "token": self.token,
                "menuId": "ReportCenter_NiftyManage_querylist4claim"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyReportInfo",
                                                   data=urlencode(data))
            if response.json()["code"] == "200":
                if len(response.json()["data"]) == len(sampleid.split(',')):
                    break
                # if "样例不在搜索范围" in response.json()["msg"]:
                #     msg_text = response.json()["msg"]
                #     year = datetime.now().strftime('%y')
                #     pattern = year + r'B\d+'
                #     result = re.findall(pattern, msg_text)
                #     sample_list = list(filter(lambda x: x not in result, sampleid.split(',')))
                #     sampleid = ','.join(sample_list)
            if count >= 3:
                break
        response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyReportInfo",data=urlencode(data))
        if response.json()["code"]=="200" and len(response.json()["data"]) == len(sampleid.split(',')):
            logger.info("产前报告认领未完成列表查询成功")
            data_list = response.json()["data"]
            for j in range(0, len(data_list), 50):
                datas = []
                for i in range(len(data_list[j:j+50])):
                    data_dic = {k: "" if v is None else v for k, v in data_list[j:j+50][i].items()}
                    data_dic["_key"] = i + 1
                    data_dic["_id"] = i + 1
                    data_dic["zstatus"] = "CHECK"
                    data_dic["username"] = "auto-test"
                    datas.append(data_dic)
                data = {
                    "datas": datas,
                    "token": self.token,
                    "menuId": "ReportCenter_NiftyManage_querylist4claim"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=updateNiftyReportInfo",
                                                       data=urlencode(data))
                if response.json()["code"] == "200" and response.json()["msg"] == "success":
                    logger.info("产前报告认领成功！")
                else:
                    logger.error(f"产前报告认领失败，返回结果：{response}")
                    raise Exception(f"产前报告认领失败，返回结果：{response}")
        else:
            logger.error(f"产前报告认领未完成列表查询失败，返回结果：{response}")
            raise Exception(f"产前报告认领未完成列表查询失败，返回结果：{response}")

    def report_composite(self,sample=None):
        """
        造数工具-产前报告复核
        :return:
        """
        time.sleep(15)
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        count = 0
        while True:
            count += 1
            time.sleep(5)
            data = {
                "task": {
                    "zsjdid": "",
                    "zsample": sampleid,
                    "zsfywc": "",
                    "step": "SIGN4",
                    "username": "c-xufeng"
                },
                "pageNumber": "1",
                "pageSize": "500",
                "token": self.token,
                "menuId": "ReportCenter_NiftyManage_querylist4review"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyReportInfo",
                                                   data=urlencode(data))
            if response.json()["code"] == "200":
                if len(response.json()["data"]) == len(sampleid.split(',')):
                    break
                # if "样例不在搜索范围" in response.json()["msg"]:
                #     msg_text = response.json()["msg"]
                #     year = datetime.now().strftime('%y')
                #     pattern = year + r'B\d+'
                #     result = re.findall(pattern, msg_text)
                #     sample_list = list(filter(lambda x: x not in result, sampleid.split(',')))
                #     sampleid = ','.join(sample_list)
            if count >= 3:
                break
        response = self.nifty_res.post_request("/presap/webintf.do?method=queryNiftyReportInfo",data=urlencode(data))
        if response.json()["code"]=="200" and len(response.json()["data"]) == len(sampleid.split(',')):
            logger.info("产前报告复核未完成列表查询成功")
            data_list = response.json()["data"]
            if self.abnormal_input == '1':
                data = {
                    "werks": "A020",
                    "zgxbh": "RT",
                    "token": self.token,
                    "menuId": "ReportCenter_NiftyManage_querylist4review"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=query_exception_list",
                                                       data=urlencode(data))
                if response.json()["code"] == "200" and len(response.json()["data"]) > 0:
                    logger.info("产前报告复核异常查询成功")
                    exception_data = response.json()["data"][0]
                    datas = []
                    exce_data = {
                        "_exception": [
                            {
                                "zreason": exception_data["zreason"],
                                "name_xl": exception_data["name_xl"],
                                "yc_level": "",
                                "yc_dl": exception_data["yc_dl"],
                                "yc_ms": exception_data["yc_ms"],
                                "name_dl": exception_data["name_dl"],
                                "yc_xl": exception_data["yc_xl"],
                                "_id": 1,
                                "zcjnam": "auto-test",
                                "zcjdat": self.date_str,
                                "zcjuzt": self.time_str
                            }
                        ]
                    }
                    for j in range(0, len(data_list), 25):
                        for i in range(len(data_list[j:j+25])):
                            data_dic = {k: "" if v is None else v for k, v in data_list[j:j+25][i].items()}
                            data_dic.update(exce_data)
                            data_dic["_key"] = i + 1
                            data_dic["_id"] = i + 1
                            data_dic["zstatus"] = "EXC"
                            datas.append(data_dic)
                        data = {
                            "datas": datas,
                            "token": self.token,
                            "menuId": "ReportCenter_NiftyManage_querylist4review"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=saveNiftyResultExceptionFlag",
                                                               data=urlencode(data))
                        if response.json()["code"] == "200" and response.json()["msg"] == "success":
                            logger.info("产前报告复核异常登记成功！")
                        else:
                            logger.error(f"产前报告复核异常登记失败，返回结果：{response}")
                            raise Exception(f"产前报告复核异常登记失败，返回结果：{response}")
                else:
                    logger.error(f"产前报告复核未完成列表查询失败，返回结果：{response}")
                    raise Exception(f"产前报告复核未完成列表查询失败，返回结果：{response}")
            else:
                datas = []
                for i in range(len(data_list)):
                    data_dic = {k: "" if v is None else v for k, v in data_list[i].items()}
                    data_dic["_key"] = i + 1
                    data_dic["_id"] = i + 1
                    datas.append(data_dic)
                data = {
                    "datas": datas,
                    "token": self.token,
                    "menuId": "ReportCenter_NiftyManage_querylist4review"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=sign_nifty_pdf",
                                                       data=urlencode(data))
                if response.json()["code"] == "200" and response.json()["msg"] == "":
                    logger.info("产前报告已复核！")
                    for j in range(0,len(data_list),50):
                        datas = []
                        for i in range(len(data_list[j:j+50])):
                            data_dic = {k: "" if v is None else v for k, v in data_list[j:j+50][i].items()}
                            data_dic["_key"] = i + 1
                            data_dic["_id"] = i + 1
                            data_dic["zstatus"] = "REVIEW"
                            data_dic["username"] = "auto-test"
                            datas.append(data_dic)
                        data = {
                            "datas": datas,
                            "token": self.token,
                            "menuId": "ReportCenter_NiftyManage_querylist4review"
                        }
                        response = self.nifty_res.post_request("/presap/webintf.do?method=updateNiftyReportInfo",
                                                               data=urlencode(data))
                        if response.json()["code"] == "200" and response.json()["msg"] == "success":
                            logger.info("产前报告复核提交成功！")
                        else:
                            logger.error(f"产前报告复核提交失败，返回结果：{response}")
                            raise Exception(f"产前报告复核提交失败，返回结果：{response}")
        else:
            logger.error(f"产前报告复核未完成列表查询失败，返回结果：{response}")
            raise Exception(f"产前报告复核未完成列表查询失败，返回结果：{response}")

    def product_supplement(self, sample=None):
        """
        造数工具-产物补录
        :return:
        """
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        sampleid = ','.join(sampleid)
        login_data = login(user='sitest')
        inbound_apply_order_number=DataGenerate(token=login_data["token"]).lims_inbound_apply(sampleid)
        DataGenerate(token=login_data["token"]).inbound_audit(inbound_apply_order_number)
        data = {
            "task": {
                "zfinish_date": "",
                "zfinish_dateend": "",
                "zsfxd_datum": "",
                "zsfxd_datumend": "",
                "zsfks_datum": "",
                "zsfks_datumend": "",
                "zjob_code": "",
                "zcatalo": sampleid,
                "zsfwc": "X",
                "task": "query",
                "zcxlx": "0"
            },
            "pageNumber": "1",
            "zgxbh": "MAB",
            "pageSize": "1000",
            "token": self.token,
            "menuId": "MSTaskProduct_XueJiangFenLi"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=findJobMS", data=urlencode(data))
        data_list = response.json()["data"][0]
        if response.json()["code"]=="200" and response.json()["total"]>0:
            logger.info("血浆分离任务单已完成列表查询成功")
            data_dic = {k: "" if v is None else v for k, v in data_list.items()}
            data_dic["_key"] = 1
            data_dic["_id"] = 1
            data = {
                "task": data_dic,
                "token": self.token,
                "menuId": "MSTaskProduct_XueJiangFenLi"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=findJobSampleMessageMS", data=urlencode(data))
            if response.json()["code"] == "200" and response.json()["total"] > 0:
                logger.info("血浆分离任务单已完成详情查询成功")
                data = {
                    "task": {
                        "zjob_code": data_list["zjob_code"]
                    },
                    "token": self.token,
                    "menuId": "MSTaskProduct_XueJiangFenLi"
                }
                response = self.nifty_res.post_request("/presap/webintf.do?method=tsk_chanw_query_bl", data=urlencode(data))
                chanwu_data = response.json()["data"]
                if response.json()["code"] == "200" and response.json()["total"] > 0:
                    logger.info("任务单产物补录详情查询成功")
                    datas = []
                    for i in range(len(chanwu_data)):
                        data_dic = {k: "" if v is None else v for k, v in chanwu_data[i].items()}
                        data_dic["_id"] = i + 1
                        datas.append(data_dic)
                    data = {
                        "task": datas,
                        "token": self.token,
                        "menuId": "MSTaskProduct_XueJiangFenLi"
                    }
                    response = self.nifty_res.post_request("/presap/webintf.do?method=tsk_chanw_save_bl",
                                                           data=urlencode(data))
                    if response.json()["code"] == "200" and response.json()["msg"] == "success":
                        logger.info("产物补录成功")
                    else:
                        logger.error(f"产物补录失败，返回结果：{response}")
                        raise Exception(f"产物补录失败，返回结果：{response}")
                else:
                    logger.error(f"任务单产物补录详情查询失败，返回结果：{response}")
                    raise Exception(f"任务单产物补录详情查询失败，返回结果：{response}")
            else:
                logger.error(f"血浆分离任务单已完成详情查询失败，返回结果：{response}")
                raise Exception(f"血浆分离任务单已完成详情查询失败，返回结果：{response}")
        else:
            logger.error(f"血浆分离任务单已完成列表查询失败，返回结果：{response}")
            raise Exception(f"血浆分离任务单已完成列表查询失败，返回结果：{response}")

    def repeat_controller(self,sample=None):
        """
        造数工具-重复质控品
        :return:
        """
        if sample:
            sampleid = sample
        else:
            sampleid = self.sample
        # nifty = NiftydataGenerate(area_code=self.area_code,token=self.token)
        self.product_supplement(sampleid)
        sampleid = ','.join(sampleid)
        data = {
            "task": {
                "zsampling_datum": "",
                "zsampling_datumend": "",
                "zreceiveddate": "",
                "zreceiveddateend": "",
                "zsample": sampleid,
                "zsfbd": "X"
            },
            "pageNumber": "1",
            "pageSize": "1000",
            "token": self.token,
            "menuId": "MSTaskProduct_RepeatController"
        }
        response = self.nifty_res.post_request("/presap/webintf.do?method=get_repeat_controllers",
                                               data=urlencode(data))
        data_list = response.json()["data"]
        if response.json()["code"]=="200":
            logger.info("重复质控品列表查询成功")
            datas = []
            for i in range(len(data_list)):
                data_dic = {k: "" if v is None else v for k, v in data_list[i].items()}
                data_dic["_key"] = i + 1
                data_dic["_id"] = i + 1
                data_dic["zsfbd"] = "X"
                data_dic["_zsjdid"] = {
                    "type": "a",
                    "key": "",
                    "ref": "",
                    "_owner": "",
                    "props": {
                        "children": data_dic["zsjdid"]
                    }
                }
                datas.append(data_dic)
            data = {
                "task": datas,
                "token": self.token,
                "menuId": "MSTaskProduct_RepeatController"
            }
            response = self.nifty_res.post_request("/presap/webintf.do?method=confirm_repeat_controllers",
                                                   data=urlencode(data))
            if response.json()["code"] == "200":
                logger.info("重复质控品保存成功")
            else:
                logger.error(f"重复质控品保存失败，返回结果：{response}")
                raise Exception(f"重复质控品保存失败，返回结果：{response}")
        else:
            logger.error(f"重复质控品列表查询失败，返回结果：{response}")
            raise Exception(f"重复质控品列表查询失败，返回结果：{response}")



if __name__ == '__main__':
    test = NiftydataGenerate(token="a02e2ab0-9ea1-4e3d-a894-74a6d33499ed",area_code="A020")
    # test.sequencing("TSK25000000278")
    # test.report_composite(["24B12180187"])
    # t = test.sumbit_sample()
    t= test.generate_report(["25B01169299","25B01166382","25B01168240","25B01167155"])
    t = test.report_review(["25B01169299","25B01166382","25B01168240","25B01167155"])
    # time.sleep(2)
    # t = test.report_claim(["25B01169299","25B01166382","25B01168240","25B01167155"])
    # time.sleep(2)
    # t = test.report_composite(["25B01169299","25B01166382","25B01168240","25B01167155"])
    # t = test.quality_inspection_products("TSK24000002419")
    # t = test.makednb500order(["MGISEQ-2000BLA0202500040"])
    # t = test.data_review(["20B6712258，25B01085570，25B01089805"])
    # t = test.repeat_controller(["25B01089805"])
    # t = test.product_supplement(["25B01083049"])

